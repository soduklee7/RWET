
'''

Disclaimer of Liability
With respect to documents available from the EPA website, neither the United States Government nor any of their employees, 
makes any warranty, express or implied, including the warranties of merchantability and fitness for a particular purpose, 
or assumes any legal liability or responsibility for the accuracy, completeness, or usefulness of any information, apparatus, product, 
or process disclosed, or represents that its use would not infringe privately owned rights.

The views expressed in this code are those of the authors and do not necessarily reflect the views or policies of 
the U.S. Environmental Protection Agency.

This code, RWET (Real-World Emission Tool) was developed to systematically process EPA's inhouse PEMS test data.
The RWET code was not cleaned, and it still under development.

No technical support and this RWET (Real World Emission Tool) code will be updated/modified without any notices.

This open-source code is used for the RDE "On-Road PEMS Test Data Analysis and Light-duty Vehicle In-use Emissions Development" data analysis 
published in SAE International Journal of Alternative Powertrains.   

1. Open the RWET python code by using a python code editor like PyCharm, Visual Studio code, etc.
2. Click the "Run" button from the python code editor.
3. Select the RDE_settings template Excel file
4. Select the file folder directory which contains the PEMS test data .csv file
5. The example PEMS test data were collected during US EPA inhouse tests.

'''

import os
import os.path
import sys
import time
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter
import glob
import datetime
#from datetime import time
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl
# import seaborn as sns
# sns.set_style("white") #for aesthetic purpose only
import numpy.polynomial.polynomial as poly
import tkinter as tk
from tkinter import *
from tkinter import filedialog, dialog

SI_units = True
RDE_UN_GTR = False
OBD = False
LAB_WLTC_NOx_Fit = False
MAW_CO2_plot_color = 'y'
#MAW_CO2_plot_color = 'gray'
MAW_vehspd_0_kph = 1
MAW_urban_vehspd_kph = 45
MAW_rural_vehspd_kph = 80
US_max_vehspd_mph = 75
US_max_vehspd_kph = 120
US_max_vehspd_mph = 90
US_max_vehspd_kph = 145
vehspd_0_kph = 0
vehspd_1_kph = 1
vehspd_urban_kph = 60
vehspd_rural_kph = 90
vehspd_motorway_kph = 145
EU_peak_vehspd_mph = 90
EU_vehspd_100kph = 100
EU_peak_vehspd_kph = 145
max_eng_stop_time = 180
max_grade_pct = 5.0
T3Bin160_NOx = 0.16
T3Bin160_NOx_hALT = 0.16
T3Bin160_CO = 4.2
T3Bin125_NOx = 0.125
T3Bin125_NOx_hALT = 0.16
T3Bin125_CO = 2.1
T3Bin110_NOx = 0.110
T3Bin110_NOx_hALT = 0.16
T3Bin110_CO = 2.1
T3Bin70_NOx = 0.110
T3Bin70_NOx_hALT = 0.16
T3Bin70_CO = 2.1
T3Bin85_NOx = 0.085
T3Bin70_NOx_hALT = 0.16
T3Bin70_CO = 2.1

T2Bin4_NOx = 0.04
T2Bin4_NMOG = 0.07
T2Bin4_CO = 2.1
T2Bin4_HCHO = 0.011
T2Bin4_PM = 0.01
T2Bin5_NOx = 0.07
T2Bin5_NMOG = 0.09
T2Bin5_CO = 4.2
T2Bin5_HCHO = 0.018
T2Bin5_PM = 0.01

calc_NOx_gbhp_hr = False
MILES_to_KM = 1.60934

def select_file(prompt):
    root = Tk()
    # root.update()
    root.focus_force()
    # Cause the root window to disappear milliseconds after calling the filedialog.
    root.after(100, root.withdraw)
    # root.withdraw()
    # here = os.path.dirname(os.path.realpath("__file__"))
    # currentdir= os.getcwd()

    #root.filename =  filedialog.askopenfilename(initialdir = currentdir, title ='Select '  + prompt, filetypes = [("All files","*.*")])
    #application_window = tk.Tk()
    #myfiletypes = [('all files', '.*'), ('text files', '.xlsx'), ('text files', '.csv')]
    myfiletypes = [('Excel files', '.xlsx'), ('CSV files', '.csv')]
    root.filename =  filedialog.askopenfilename(initialdir=os.getcwd(), title ='Select ' + prompt, filetypes = myfiletypes)
    print(root.filename)
    
    select_file_name = root.filename
    root.destroy()
    root.quit()
    return select_file_name

def save_folder(prompt):
    root = Tk()
    #root.withdraw()
    root.focus_force()
    # Cause the root window to disappear milliseconds after calling the filedialog.
    root.after(100, root.withdraw)
    #here = os.path.dirname(os.path.realpath("__file__"))
    currentdir= os.getcwd()
    root.directory = filedialog.askdirectory(initialdir = currentdir, title = 'Select ' + prompt)
    save_folder_name = root.directory
    root.destroy()
    root.quit()
    return save_folder_name

class takeInput(object):
    '''
    for taking a text input, for a run ID or user provided text string for use within a script
    '''

    def __init__(self, requestMessage):
        self.root = Tk()
        self.string = ''
        self.frame = Frame(self.root)
        self.frame.pack()
        self.acceptInput(requestMessage)

    def acceptInput(self, requestMessage):
        r = self.frame

        k = Label(r,text=requestMessage)
        k.pack(side='left')
        self.e = Entry(r,text='Name')
        self.e.pack(side='left')
        self.e.focus_set()
        b = Button(r,text='Enter',command=self.gettext)
        b.pack(side='right')

    def gettext(self):
        self.string = self.e.get()
        self.root.destroy()
        self.root.quit()

    def getString(self):
        return self.string

    def waitForInput(self):
        self.root.mainloop()

def getText(requestMessage):
    msgBox = takeInput(requestMessage)
    #loop until the user makes a decision and the window is destroyed
    msgBox.waitForInput()
    return msgBox.getString()

def polyfit(x, y, degree):
    results = {}

    coeffs = np.polyfit(x, y, degree)

     # Polynomial Coefficients
    results['polynomial'] = coeffs.tolist()

    # r-squared
    p = np.poly1d(coeffs)
    # fit values, and mean
    yhat = p(x)                         # or [p(z) for z in x]
    ybar = np.sum(y)/len(y)          # or sum(y)/len(y)
    ssreg = np.sum((yhat-ybar)**2)   # or sum([ (yihat - ybar)**2 for yihat in yhat])
    sstot = np.sum((y - ybar)**2)    # or sum([ (yi - ybar)**2 for yi in y])
    results['determination'] = ssreg / sstot

    return results

def vehmove_stop(veh_move_stop, iENG_SPEED_cname, engineRPM):
    engine_stop = 0
    for j in range (veh_move_stop-1, 1, -1):
        if iENG_SPEED_cname != '' and float(engineRPM[j-1]) >= 500 and float(engineRPM[j]) < 500:
            engine_stop = j
            break      

    return engine_stop

#pylab.ylabel('Example', fontsize=40).
#ax.set_ylabel('Example', fontsize=40) or afterwards with ax.yaxis.label.set_size(40).

def plt_bgf_format(pp1, fig, plt, ax1, ax2_label, vehicle_title, xlabel, ylabel):
    title=plt.title(vehicle_title)
    plt.xlabel(xlabel, fontsize=14)
    plt.ylabel(ylabel, fontsize=14)
    legend = ax1.legend(loc='best')
    ax1.grid(which='major', axis='both', linestyle='-', color='k', linewidth=0.25)
    ax1.tick_params(axis='both', which='major', labelsize=14)
    ax1.tick_params(axis='both', which='minor', labelsize=14)

    if xlabel == '': ax1.set_xticks([])
    
    if ax2_label != '':
        ax2 = ax1.twinx()
        ax2.set_ylabel(ax2_label, color='b', fontsize=14)
        ax2.set_yticks([])

    if pp1 != '': pp1.savefig(fig, dpi=600)
    
def etime_resetting(etime, vehspd):
    netime = len(etime)
    etime1 = np.zeros(netime)
    etime1[0] = etime[0]
    dt = 0
    
    vehspd1 = np.zeros(len(vehspd))
    time_span = 600
    tvehspd_zero = istop = istart = 0
    stvehspd_zero = 0
    ip1 = 0
    for ik in range (1, netime-1):
        dt = (etime[ik] - etime[ik-1])
        if ik > istart and vehspd[ik-1] > vehspd_1_kph and vehspd[ik] < vehspd_1_kph and vehspd[ik+1] < 0.5:
                
            istop = ik
            stvehspd_zero = 0
            for j in range(istop+1, netime-1):
                stvehspd_zero += vehspd[j]
                if  (vehspd[j-1] < 0.5 and vehspd[j] < 0.5 and vehspd[j+1] > 1) or stvehspd_zero > 1:
                    istart = j+1
                    tvehspd_zero = istart - istop
                    if tvehspd_zero > time_span:
                        sum_tvehspd_zero = np.sum(vehspd[istop+1:istart-1])
                        print('istart ', istart, ' istop ', istop, ' tvehspd_zero ', tvehspd_zero, ' sum_tvehspd_zero ', stvehspd_zero)
                        istop = istart + 1
                    break

        if dt > time_span or tvehspd_zero > time_span:
            etime1[ik] = etime1[ik-1] + 30
            print('inside time reset, dt = ', dt, ' tvehspd_zero ', tvehspd_zero )
            tvehspd_zero = stvehspd_zero = 0
        else:
           etime1[ik] = etime1[ik-1] + dt
           
    if etime1[ik+1] < 0.1 and etime[netime-1] > 1: etime1[ik+1] = etime1[ik]+1
    
    etime = etime1 
    del etime1, vehspd1
    return etime, tvehspd_zero    

def subplot_shared_x_axes(sf, pp1, splot_mode, title, x1, y1, y1c, xlabel1, ylabel1, legend1, x2, y2, y2c, xlabel2, ylabel2, legend2, \
                          x3, y3, y3c, xlabel3, ylabel3, legend3, x3r, y3r, y3rc, ylabel3r, legend3r):
    # fig = plt.figure()

    if (len(x2) > 0 and len(y2) > 0):
        if (splot_mode == 'share'):
            fig, axes = plt.subplots(3, 1, sharex=True)  # , sharey=True)
        else:
            fig, axes = plt.subplots(3, 1)

        subplot1 = axes[0]
        subplot2 = axes[1]
        subplot3 = axes[2]
    elif len(x2) == 0 and len(y2) == 0 and len(x3) > 0 and len(y3) > 0:
        if (splot_mode == 'share'):
            fig, axes = plt.subplots(2, 1, sharex=True)  # , sharey=True)
        else:
            fig, axes = plt.subplots(2, 1)

        subplot1=axes[0]
        subplot3 =axes[1]
    else:
        fig, axes = plt.subplots(1, 1) #, sharex=True)  # , sharey=True)
        subplot1=axes[0]

    fig.set_size_inches(11*sf, 8.5*sf, forward=True)

    subplot1.plot(x1, y1, y1c, label=legend1)
    subplot1.set_xlim(0, max(x1))
    subplot1.grid()
    subplot1.set_title(title)
    subplot1.set_ylabel(ylabel1)
    # subplot1.legend(loc='best')
    if len(x2) == 0 and len(x3) == 0: subplot1.set_xlabel(xlabel1)

    if len(x2) > 0 and len(y2) > 0:
        subplot2.plot(x2, y2, y2c, label=legend2)
        subplot2.set_xlim(0, max(x2))
        if min(y2) < 0 and max(y2) <= 30: 
            subplot2.set_ylim(np.round(min(y2)), np.round(max(y2)*1.5))
        elif min(y2) >= 0  and max(y2) <= 30:
            subplot2.set_ylim(0, np.round(max(y2)+5))
        subplot2.grid()
        subplot2.set_ylabel(ylabel2)
        subplot2.legend(loc='best')
      
    if len(x3) > 0 and len(y3) > 0:
        lns1 = subplot3.plot(x3, y3, y3c, label= legend3)
        subplot3.set_xlim(0, max(x3))
        subplot3.grid()
        subplot3.set_xlabel(xlabel3)
        subplot3.set_ylabel(ylabel3)
    if splot_mode == 'share' and len(x3r) > 0 and len(y3r) > 0:
        ax4 = subplot3.twinx()
        lns2 = ax4.plot(x3r, y3r, y3rc, label=legend3r)
        ax4.set_ylabel(ylabel3r, color='b')
        lns = lns1+lns2
        labs = [l.get_label() for l in lns]
        ax4.legend(lns, labs, loc=0)
    elif len(x3) > 0 and len(y3) > 0:
        subplot3.legend(loc='best')

    pp1.savefig(fig, dpi=600)
    
def RDE_NOx_table(RDE_FTPbin_fname, vehicle, RDE_NOx, RDEu_NOx, STD_NOx, BIN_label, FTP_NOxc, US_NOxc, HwFET_NOx):
    
    wb = openpyxl.load_workbook(engineDB_ifile)
    wb1 = openpyxl.load_workbook(engineDB_ofile)
    sheet = wb["Marine SI"]
    sheet1 = wb1["Marine SI"]
    
#    sheet1.insert_rows(idx=1)
    
    header_row = 1
    ncols = 95
    for i in range(1, header_row+1):
        for j in range(1, ncols):
            sheet1.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value
    
    yellow_background = PatternFill(bgColor=colors.YELLOW)
    diff_style = DifferentialStyle(fill=yellow_background)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["$H1<3"]
    sheet1.conditional_formatting.add("A1:CQ1", rule)
    
    times_bold_font = Font(name='Times New Roman', bold=True)
    big_red_text = Font(color=colors.RED, size=20)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                right=double_border_side,
                bottom=double_border_side,
                left=double_border_side)
    
    for cell in sheet1["1:1"]:
        cell.font = times_bold_font
    wb1.close(engineDB_ofile)
    # wb1.save(engineDB_ofile)

'''
    engine torque from engine load * engine max torque at the engspd = f(engine speed, engine load)
>>> xp = [1, 2, 3]
>>> fp = [3, 2, 0]
>>> np.interp(2.5, xp, fp)
1.0
'''

def timePEMS_to_second(timestr):
    [h, m, s] = timestr.split(':')
    s = s.split('.')
    s = s[0]
    return int(h) * 3600 + int(m) * 60 + int(s)
    
def timeOBD_to_second(hms_str):
    str1 = hms_str
    if len(str1) == 6:
        hh = str1[0:2]
        mm =  str1[2:4]
        ss = str1[4:6]
    else:
        hh = str1[0:1]
        mm = str1[1:3]
        ss = str1[3:5]

    return hh, mm, ss, (int(hh) * 3600 + int(mm) * 60 + int(ss))
       
def vehspd_obd_sync(ndoe, nsamples_obd_sync, time_start, etime2, vehspd_pems, otime2, vehspd_obd):
    itime_start = time_start
    ntime = len(otime)
    timeoffset_obd=0
    timeoffset_pems = 0
    
    rms_vehspd_pems = np.zeros(ndoe)
    rms_vehspd_obd = np.zeros(ndoe)
    k = 0
    for j in range (ndoe):
        tmp = 0
        otmp = 0
        for i in range (itime_start, itime_start + nsamples_obd_sync): 
            tmp += (abs(vehspd_pems[j+i] - vehspd_obd[i])) 
            otmp += (abs(vehspd_pems[i] - vehspd_obd[j+i])) 

        rms_vehspd_pems[k] = (tmp**2/nsamples_obd_sync)**0.5
        rms_vehspd_obd[k] = (otmp**2/nsamples_obd_sync)**0.5
        k += 1
        
    min_rms_vehspd_pems = np.min(rms_vehspd_pems)
    min_rms_vehspd_obd = np.min(rms_vehspd_obd)
    if min_rms_vehspd_pems <   min_rms_vehspd_obd:   
        df_sync = pd.DataFrame(rms_vehspd_pems)
        timeoffset_obd = df_sync.loc[df_sync.idxmin()].index[0]
        # otime = otime + timeoffset1
    else:
        df_sync = pd.DataFrame(rms_vehspd_obd)
        timeoffset_pems = df_sync.loc[df_sync.idxmin()].index[0]
        # etime = etime + timeoffset       
    return timeoffset_pems, timeoffset_obd

def add_df(idf, cname, title, unit, istart, iend, idata):  
    idf[cname] = ''
    idf[cname].columns = cname
    icname = title
    idf[cname][1] = unit
    idf[cname][istart:iend+1] = idata

def Relative_Positive_Acceleration(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, dVMTu, dVMTr, dVMTm,  Vu_ACC, Vr_ACC, Vm_ACC, a_pos_limit):
    RPAu = 0
    RPAr = 0
    RPAm = 0
    if len(Vu_ACC[Vu_ACC > a_pos_limit]) > 0:
        sum_vehspd_accel_pos = np.sum(vehspd_ACCu[Vu_ACC > a_pos_limit])# 
        RPAu = sum_vehspd_accel_pos / np.sum(dVMTu)
        
    if len(Vr_ACC[Vr_ACC > a_pos_limit]) > 0:
        sum_vehspd_accel_pos = np.sum(vehspd_ACCr[Vr_ACC > a_pos_limit])# 
        RPAr = sum_vehspd_accel_pos / np.sum(dVMTr)

    if len(Vm_ACC[Vm_ACC > a_pos_limit]) > 0:
        sum_vehspd_accel_pos = np.sum(vehspd_ACCm[Vm_ACC > a_pos_limit])# 
        RPAm = sum_vehspd_accel_pos / np.sum(dVMTm)
        
    return RPAu, RPAr, RPAm

def VehSpd_Accel_95pct(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, Vu_ACC, Vr_ACC, Vm_ACC, a_pos_limit):
    vehspd_ACCu_95pct = 0
    vehspd_ACCr_95pct = 0
    vehspd_ACCm_95pct = 0
    if len(Vu_ACC[Vu_ACC > a_pos_limit]) > 0: vehspd_ACCu_95pct= np.percentile(vehspd_ACCu[Vu_ACC > a_pos_limit], 95)  
    if len(Vr_ACC[Vr_ACC > a_pos_limit]) > 0: vehspd_ACCr_95pct= np.percentile(vehspd_ACCr[Vr_ACC > a_pos_limit], 95)  
    if len(Vm_ACC[Vm_ACC > a_pos_limit]) > 0: vehspd_ACCm_95pct= np.percentile(vehspd_ACCm[Vm_ACC > a_pos_limit], 95)  
    
    return vehspd_ACCu_95pct, vehspd_ACCr_95pct, vehspd_ACCm_95pct

def Vehspd_mean(vehspd_u, vehspd_r, vehspd_m):
    vehspd_u_mean = 0
    vehspd_r_mean = 0
    vehspd_m_mean = 0
    if len(vehspd_u) > 0: vehspd_u_mean= np.mean(vehspd_u) # 
    if len(vehspd_r) > 0: vehspd_r_mean= np.mean(vehspd_r) # 
    if len(vehspd_m) > 0: vehspd_m_mean= np.mean(vehspd_m)# 
    
    return vehspd_u_mean, vehspd_r_mean, vehspd_m_mean

def RDE_report (rpti, filename_pems, filename_obd, output_folder, now_hms, suffix, trip_distance, trip_duration, cold_start_duration, urban_distance, rural_distance, motorway_distance,
                urban_distance_share, rural_distance_share, motorway_distance_share, urban_avg_speed, rural_avg_speed, motorway_avg_speed, total_trip_avg_speed,
                motorway_speed_above_145kph, motorway_speed_above_100kph, urban_stop_time_pct, elev_abs_diff_start_end, total_cumul_elev_pos_gain, urban_cumul_elev_pos_gain,
                init_idle_duration, cold_start_avg_speed, cold_start_max_speed, cold_stop_time, vehspd_avg, RPA_avg, VA_avg_95pct,
                nVu_ACC, nVr_ACC, nVm_ACC, nMAWt, nMAWu, nMAWr, nMAWm, CO_RDE_mgpkm, NOx_RDE_mgpkm, CO_RDEu_mgpkm, NOx_RDEu_mgpkm, 
                FE_mpg, CO2_gpm_test, engNOx_mgpm_test, CO_mgpm_test):

    rpt = rpti.copy()
    del rpti
    
    print("Total trip distance [km] = ", trip_distance)
    print("Total trip duration min. [90-120] = ", trip_duration)
    print("Cold start duration min. [<=5] = ", cold_start_duration)
    
    print("Urban distance, km [>16] = ", urban_distance)
    print("Rural distance, km [>16] = ", rural_distance)
    print("Motorway distance, km [>16] = ", motorway_distance)
    
    print("Urban distance share, % [29-44] = ", np.round(urban_distance_share, 1))
    print("Rural distance share, % [23-43] = ", np.round(rural_distance_share, 1))
    print("Motorway distance share, % [23-43] = ", np.round(motorway_distance_share, 1))
    
    print("Urban average speed [kph] = ", urban_avg_speed)
    print("Rural average speed [kph] = ", rural_avg_speed)
    print("Motorway average speed [kph] = ", motorway_avg_speed)
    print("Total trip average speed [kph] = ", total_trip_avg_speed)
    
    print("Motorway speed above 145 km/h, % [<3% mot. time] = ", motorway_speed_above_145kph)
    print("Motorway speed above 100 km/h, min [>=5] = ", motorway_speed_above_100kph)
    print("Urban stop time, % [6-30] = ", urban_stop_time_pct)
    
    print("Start and end points elevation absolute difference, m [<=100m] = ", elev_abs_diff_start_end)    #print("Start and end points elevation absolute difference averaged, m [<=100m] = ", round((d_alt1 + d_elev)/2, 2))
    print("Total cumulative positive elevation gain, m/100km [<1200m/100km] = ", total_cumul_elev_pos_gain)
    print("Urban cumulative positive elevation gain, m/100km [<1200m/100km] = ", urban_cumul_elev_pos_gain)
    
    if time_engine_idle >= 300:
        print("Idling event(s) longer than 300 s, Yes/No = ", "Yes")
        engine_idle_gt_5min = 'Yes'
    else:
        print("Idling event(s) longer than 300 s, Yes/No = ", "No")
        engine_idle_gt_5min = 'Yes'
        
    print("Initial idling duration, s [<=15] = ", init_idle_duration)
    
    print("Cold start average speed, km/h [15-40] = ", cold_start_avg_speed)
    print("Cold start maximum speed, km/h [< 60] = ", cold_start_max_speed)
    print("Cold start stop time, s [<=90] = ", cold_stop_time)
    
    irow = 0
    rpt.loc[irow, 'DATA QC'] = filename_pems + '.xlsx'
    irow = irow + 1
    rpt.loc[irow, 'DATA QC'] = filename_obd
    irow = irow + 1
    rpt.loc[irow, 'DATA QC'] = ''
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = FE_mpg
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = CO2_gpm_test
    irow = irow + 1

    irow = 7
    rpt.loc[irow, 'RWET'] = 1.60934*CO_RDE_mgpkm
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = 1.60934*NOx_RDE_mgpkm
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = ''
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = 1.60934*CO_RDEu_mgpkm
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = 1.60934*NOx_RDEu_mgpkm
    irow = irow + 1

    irow = 19
    rpt.loc[irow, 'RWET'] = trip_distance
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = trip_duration
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = cold_start_duration
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = urban_distance
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = rural_distance
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = motorway_distance
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = np.round(urban_distance_share, 1)
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = np.round(rural_distance_share, 1)
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = np.round(motorway_distance_share, 1)
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = urban_avg_speed
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = rural_avg_speed
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = motorway_avg_speed
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = total_trip_avg_speed
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = motorway_speed_above_145kph
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = motorway_speed_above_100kph
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = urban_stop_time_pct
    irow = irow + 1
    
    rpt.loc[irow, 'RWET'] = elev_abs_diff_start_end
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = total_cumul_elev_pos_gain
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = urban_cumul_elev_pos_gain
    irow = irow + 1
    
    rpt.loc[irow, 'RWET'] = engine_idle_gt_5min
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = init_idle_duration
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = cold_start_avg_speed
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = cold_start_max_speed
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = cold_stop_time
    irow = irow + 1
    
    irow = 48
    rpt.loc[irow, 'RWET'] = round(GHG_WLTC_threshold/1000, 3)
    irow = irow + 1
    
    irow = 56
    rpt.loc[irow, 'RWET'] = nMAWt
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nMAWt
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = 0
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nMAWu
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nMAWr
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nMAWm
    irow = irow + 1

    rpt.loc[irow, 'RWET'] = nMAWu
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nMAWr
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nMAWm
    irow = irow + 1

    rpt.loc[irow, 'RWET'] = 100
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = 100
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = 100
    irow = irow + 1
    
    irow = 70
    rpt.loc[irow, 'RWET'] = RPA_avg[0]
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = RPA_avg[1]
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = RPA_avg[2]
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = VA_avg_95pct[0]
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = VA_avg_95pct[1]
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = VA_avg_95pct[2]
    irow = irow + 1

    rpt.loc[irow, 'RWET'] = nVu_ACC
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nVr_ACC
    irow = irow + 1
    rpt.loc[irow, 'RWET'] = nVm_ACC
    irow = irow + 1

    outFilename = os.path.join(output_folder, os.path.basename(filename_pems))
    outFilename= os.path.splitext(outFilename)[0] + '_rpt' + now_hms + suffix
    rpt.to_html(outFilename)
    return rpt

'''
https://stackoverflow.com/questions/38511373/change-the-color-of-text-within-a-pandas-dataframe-html-table-python-using-style
'''        

t = time.time()
path = os.getcwd() + '/'

'''
MAW CO2 curve page 156
1) p1  to p2
    Mco2 = a1 x vk + b1
    where, a1 = (MCo2 at p2 - Mcos2 at p1)/(vk at p2 - vk at p1)
    b1 = Mco2 at p1 - a1 x vk at p1
    
2) p2 to p3
    a2 = (Mco2 at p3 - Mco2 at p2) / (vk at p3 - vk at p2) 
    b2 = Mco2 at p2 - a2 * vk at p2    
    
    Mco2 = a2 vk + b2    
'''
engine_peak_trq = 0
engtrq_RPM = 0
engtrqWOT_NM = 0
     
RDE_report_settings_file = select_file('RDE Report & Settings Excel file')
xls = pd.ExcelFile(RDE_report_settings_file)
rpt = pd.read_excel(xls, sheet_name='RDE REPORT', index_col=None, header=0, skiprows=1, dtype=object)
rpt = rpt.replace(np.nan, '', regex=True)

df = pd.read_excel(xls, sheet_name='Settings', index_col=None, header=0, skiprows=0, dtype=object)
df = df.fillna('')

settings_MAw_CO2_pct = 50
etime_reset = FALSE
vehicle_type = 1
OBD = FALSE
LAB_WLTC_NOx_Fit = FALSE
Exhaust_Emissions_Standard = 'T2Bin4'
WLTC_GHG_gpkm = ''
WLTC_GHGu_gpkm = ''

if df.loc[0, 'PASS-FAIL CALCULATIONS'] == 'CO2 window [%]': settings_MAw_CO2_pct = df['Default'][0]
if df.loc[1, 'PASS-FAIL CALCULATIONS'] == 'etime_reset': etime_reset = df['Default'][1]
if df.loc[2, 'PASS-FAIL CALCULATIONS'] == 'VEH_TYPE': vehicle_type = df['Default'][2]
if df.loc[3, 'PASS-FAIL CALCULATIONS'] == 'sync_OBD': OBD = df['Default'][3]
if df.loc[4, 'PASS-FAIL CALCULATIONS'] == 'LAB_WLTC_NOx_Fit': LAB_WLTC_NOx_Fit = df['Default'][4]
if df.loc[5, 'PASS-FAIL CALCULATIONS'] == 'Exhaust_Emissions_Standard': Exhaust_Emissions_Standard = df['Default'][5]
if ('RDE Mode (LD/HD)') in df.loc[6, 'PASS-FAIL CALCULATIONS']: RDE_Mode = df['Default'][6]
if ('Calc NOx (g/bhp-hr)') in df.loc[7, 'PASS-FAIL CALCULATIONS']: calc_NOx_gbhp_hr = df['Default'][7]
if ('Peak Torque (NM)') in df.loc[8, 'PASS-FAIL CALCULATIONS']: engine_peak_trq = float(df['Default'][8])
if ('WLTC Urban CO2') in df.loc[9, 'PASS-FAIL CALCULATIONS']: WLTC_GHGu_gpkm = float(df['Default'][9])/1.60934

engtrq_RPM_cols = [col for col in df.columns if 'engtrq_RPM' in col]
engtrqWOT_cols = [col for col in df.columns if 'engtrqWOT_NM' in col]
if len(engtrq_RPM_cols) > 0: engtrq_RPM = df['engtrq_RPM'].values
if len(engtrqWOT_cols) > 0: engtrqWOT_NM = df['engtrqWOT_NM'].values

WLTC_CO2_gpkm = df['WLTC_CO2_gpmi'][0:4].values/1.60934
WLTC_GHG_gpkm = WLTC_CO2_comp_gpkm = float(df.loc[4, 'WLTC_CO2_gpmi'])/1.60934
WLTC_CO2_gpkm[0] = WLTC_CO2_gpkm[0] 
WLTC_CO2_gpkm[1] = WLTC_CO2_gpkm[1] 
WLTC_CO2_gpkm[2] = WLTC_CO2_gpkm[2] 
WLTC_CO2_gpkm[3] = WLTC_CO2_gpkm[3] 

WLTC_vehspd_kph = df['vehspd_WLTC_kph'][0:4].values
WLTC_vehspd_comp_kph = float(df['vehspd_WLTC_kph'][4])# WLTC vehicle speed composite
WLTC_dist_km = df['WLTC_dist_km'][0:4].values

settings_avg_speed = df['MAW_vehspd_kph'][0:16].values
settings_CO2= df['MAW_CO2_gpkm'][0:16].values
settings_TOL1_minus_pct = df['TOL1-'][0:16].values
settings_TOL1_plus_pct = df['TOL1+'][0:16].values

vehspd_FTP3bags_kph = df['vehspd_FTP3bags_kph'][0:3].values
vehspd_US06_kph = df['vehspd_US06_kph'][0:2].values
vehspd_FTP3bags_comp_kph = float(df['vehspd_FTP3bags_kph'][3])
vehspd_US06_comp_kph = float(df['vehspd_US06_kph'][2])
vehspd_HWFET_kph = df['vehspd_HWFET_kph'][0:1].values

FTP3_CO2_gpmi = df['CO2_FTP3bags_gpmi'][0:3].values
US06_CO2_gpmi = df['CO2_US06_gpmi'][0:2].values
HwFET_CO2_gpmi = df['CO2_HwFET_gpmi'][0:1].values
FTP3_CO2_comp_gpmi = float(df['CO2_FTP3bags_gpmi'][3])
US06_CO2_comp_gpmi = float(df['CO2_US06_gpmi'][2])

WLTC_NOx_gpmi = df['NOx_WLTC_gpmi'][0:4].values
FTP3_NOx_gpmi = df['NOx_FTP3bags_gpmi'][0:3].values
US06_NOx_gpmi = df['NOx_US06_gpmi'][0:2].values
HwFET_NOx_gpmi = df['NOx_HwFET_gpmi'][0:1].values
WLTC_NOx_comp_gpmi = float(df['NOx_WLTC_gpmi'][4])
FTP_NOx_comp_gpmi = float(df['NOx_FTP3bags_gpmi'][3])
US06_NOx_comp_gpmi = float(df['NOx_US06_gpmi'][2])

WLTC_CO_gpmi = df['CO_WLTC_gpmi'][0:4].values
FTP3_CO_gpmi = df['CO_FTP3bags_gpmi'][0:3].values
US06_CO_gpmi = df['CO_US06_gpmi'][0:2].values
HwFET_CO_gpmi = df['CO_HwFET_gpmi'][0:1].values
WLTC_CO_comp_gpmi = float(df['CO_WLTC_gpmi'][4])
FTP_CO_comp_gpmi = float(df['CO_FTP3bags_gpmi'][3])
US06_CO_comp_gpmi = float(df['CO_US06_gpmi'][2])

if (Exhaust_Emissions_Standard == 'T2Bin4') or (Exhaust_Emissions_Standard == 'T2BIN4'): 
    NOx_STD_mgpmi = T2Bin4_NOx*1000
    CO_STD_gpmi = T2Bin4_CO
elif (Exhaust_Emissions_Standard =='T2Bin5') or (Exhaust_Emissions_Standard =='T2BIN5'): 
    NOx_STD_mgpmi = T2Bin5_NOx*1000
    CO_STD_gpmi = T2Bin5_CO
elif (Exhaust_Emissions_Standard =='T3Bin70') or (Exhaust_Emissions_Standard =='T3BIN70'): 
    NOx_STD_mgpmi = T3Bin70_NOx*1000
    CO_STD_gpmi = T3Bin70_CO
elif (Exhaust_Emissions_Standard =='T3Bin85') or (Exhaust_Emissions_Standard =='T3BIN85'): 
    NOx_STD_mgpmi = T3Bin85_NOx*1000
    CO_STD_gpmi = T3Bin85_CO
elif (Exhaust_Emissions_Standard =='T3Bin125') or (Exhaust_Emissions_Standard =='T3BIN125'): 
    NOx_STD_mgpmi = T3Bin125_NOx*1000
    CO_STD_gpmi = T3Bin125_CO
elif (Exhaust_Emissions_Standard == 'T3Bin160') or (Exhaust_Emissions_Standard == 'T3BIN160'): 
    NOx_STD_mgpmi = T3Bin160_NOx*1000
    CO_STD_gpmi = T3Bin160_CO

df_setting = df
del xls

GHG_WLTC_threshold = (settings_MAw_CO2_pct/100 + 0.0001)*np.sum(WLTC_dist_km * WLTC_CO2_gpkm)

script_inputs_file = ''
infile_folder = save_folder('Select PEMS and OBD input file folder')

suffix = ".html"

now=datetime.datetime.now()
hhmmss=datetime.time(now.hour, now.minute, now.second) #'{:02d}:{:02d}:{:02d}'.format(now.hour, now.minute, now.second)
hms = str(hhmmss).split(':')
now_hms = str(hms[0]) + str(hms[1]) + str(hms[2])

odir = os.getcwd()    
os.chdir(infile_folder)

elapsed = time.time() - t
print("\n*************\nElapsed Time: {:5.0f} Seconds after selecting the template and file folder" .format(elapsed))

extension = 'csv'
pp_filenames = [i for i in glob.glob('pp_*.{}'.format(extension))]
csv_filenames = [i for i in glob.glob('*.{}'.format(extension))]
nel_df1 = 0
vehicle_title = ''

pp_files = pd.DataFrame(pp_filenames, columns=['fname']).sort_values('fname')
pp_files = pp_files.reset_index(drop = True)

sTIME_cname = ''
sDATE_cname =''
sSTATUS_PATH_cname =''
iVEH_SPEED_cname = ''
imVEH_SPEED_cname = ''
iGPS_GROUND_SPEED_cname=''
iGPS_ALT_cname=''
iENG_SPEED_cname=''
iENG_LOAD_cname =''
Lambda_cname = ''
iCOOL_TEMP_cname = ''
imCOOL_TEMP_cname =''
TP_B_cname =''
APP_D_cname=''
APP_P_cname =''
APP_E_cname =''
P_ACCEL_POSTN_cname =''
iWfgps_cname =''
iSCB_LAP_cname=''
Tamb_cname=''
CatalystTemp_cname=''
iCALCRT_CO2m_cname=''
iCALCRT_COm_cname = ''
iCALCRT_NOxm_cname = ''
iCALCRT_kNOxm_cname = ''
npp_files = len(pp_files)

for i in range (npp_files):
    pp_file = str(pp_files.fname[i])
    if i == 0:
        cols = []
        units = []
        icol_ECT = ''
        sTIME_unit = ''
        iWfgps_unit = ''
        ECT_unit = ''
        iVEH_SPEED_unit =''
        imVEH_SPEED_unit =''
        iGPS_GROUND_SPEED_unit =''
        Tamb_unit =''
        Tamb_cname=''
        df = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape', low_memory=False)
        df.dropna(axis=1, how='all', inplace=True)
        #, names = [f'col{i+1}' for i in range(286)])# add "flot" in the SG_fuel
        # df_units = pd.read_csv(pp_file, skiprows=2, escapechar='\\', encoding = 'unicode_escape')# add "flot" in the SG_fuel
        columns = list(df.columns)
        units = list(df.loc[0, :])
        # units = []
        # for i in range(len(df_units.columns)):
        #     if ('.' in df_units.columns[i]) and (df_units.columns[i] != 'hh:mm:ss.xxx'):
        #         units.append(df_units.columns[i].split('.')[0])
        
        try: 
            for text in columns:
                if 'sTIME' in text: 
                    sTIME_cname = 'sTIME'
                    icol_sTIME = columns.index(sTIME_cname)
                    if ('hh:mm:ss.xxx' in units[icol_sTIME]) or ('hh:mm:ss' in units[icol_sTIME]):
                        sTIME_unit = 'hhmmss'
                    else:
                        sTIME_unit = 's'

                if 'iCOOL_TEMP' in text: 
                    iCOOL_TEMP_cname = text
                    icol_ECT = columns.index(iCOOL_TEMP_cname)
                    if ('deg F' in units[icol_ECT]) or ('degF' in units[icol_ECT]):
                        ECT_unit = 'deg F'
                    else:
                        ECT_unit = 'deg C'

                if 'imCOOL_TEMP' in text: 
                    imCOOL_TEMP_cname = 'imCOOL_TEMP'
                    imcol_ECT = columns.index(imCOOL_TEMP_cname)
                    if ('deg F' in units[imcol_ECT]) or ('degF' in units[imcol_ECT]):
                        ECT_unit = 'deg F'
                    else:
                        ECT_unit = 'deg C'

                if 'iWfgps' in text: 
                    iWfgps_cname = 'iWfgps'
                    icol_iWfgps = columns.index(iWfgps_cname)
                    if ('gal/s' in units[icol_iWfgps]):
                        iWfgps_unit = 'gal/s'
                    else:
                        iWfgps_unit = 'g/s'
                    
                if ('CATEMP11' or 'iCATEMP11') in text: 
                    CatalystTemp_cname = text
                    icol_CatalystTemp = columns.index(CatalystTemp_cname)
                    if ('Deg F' in units[icol_CatalystTemp]) or ('degF' in units[icol_CatalystTemp]) or ('deg F' in units[icol_CatalystTemp]):
                        CatalystTemp_unit = 'deg F'
                    else: 
                        CatalystTemp_unit = 'deg C'

                if 'imVEH_SPEED' in text: 
                    imVEH_SPEED_cname = 'imVEH_SPEED'
                    icol_VEH_SPEED = columns.index(imVEH_SPEED_cname)
                    if ('kph' in units[icol_VEH_SPEED]) or ('km/h' in units[icol_VEH_SPEED]):
                        imVEH_SPEED_unit = 'kph'
                    else: 
                        imVEH_SPEED_unit = 'mph'
                if 'iVEH_SPEED' in text: 
                    iVEH_SPEED_cname = 'iVEH_SPEED'
                    icol_VEH_SPEED = columns.index(iVEH_SPEED_cname)
                    if ('mph' in units[icol_VEH_SPEED]) or ('mi/h' in units[icol_VEH_SPEED]):
                        iVEH_SPEED_unit = 'mph'
                    else: 
                        iVEH_SPEED_unit = 'kph'

                if 'iGPS_GROUND_SPEED' in text:
                    iGPS_GROUND_SPEED_cname = 'iGPS_GROUND_SPEED'
                    icol_iGPS_GROUND_SPEED = columns.index(iGPS_GROUND_SPEED_cname)
                    if ('mph' in units[icol_iGPS_GROUND_SPEED]) or ('MPH' in units[icol_iGPS_GROUND_SPEED]):
                        iGPS_GROUND_SPEED_unit = 'mph'
                    elif ('kph' in units[icol_iGPS_GROUND_SPEED]) or ('KPH' in units[icol_iGPS_GROUND_SPEED]):
                        iGPS_GROUND_SPEED_unit = 'kph'
                    else:
                        iGPS_GROUND_SPEED_unit = 'mph'

                if 'iENG_SPEED' in text: iENG_SPEED_cname = 'iENG_SPEED'
                if 'iENG_LOAD' in text: iENG_LOAD_cname = 'iENG_LOAD'               
                if ('iSCB_LAT' or 'AAT' or 'iAAT' or 'AmbientTemp') in text: 
                    Tamb_cname = text.split('.')[0]
                    icol_Tamb = columns.index(Tamb_cname)
                    if ('Deg F' in units[icol_Tamb]) or ('degF' in units[icol_Tamb]) or ('deg F' in units[icol_Tamb]):
                        Tamb_unit = 'deg F'
                    else:
                        Tamb_unit = 'deg C'
                    
        except ImportError:
            iVEH_SPEED_cname = ''
            iGPS_GROUND_SPEED_cname = ''
            iENG_SPEED_cname = ''
            iENG_LOAD_cname = ''
                
        # del df_columns, df_units
        
        # df = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape')# add "flot" in the SG_fuel
        df = df.drop(df.index[0])
        df = df.reset_index(drop = True)
        df = df.replace(np.nan, '', regex=True)

        itmp = df[df['sDATE'].str.match('Vehicle Description:')].index[0]# df[df['model'].str.contains('ac')]
        vehicle_title = df.sTIME[itmp] + ', '
        itmp = df[df['sDATE'].str.match('Summary Information:')].index[0] - 3 # df[df['model'].str.contains('ac')]
        
        nel_df1 = len(df.sTIME)
        df = df.drop(df.index[itmp : nel_df1])
        df.dropna(axis=1, how='all', inplace=True)
        df = df.reset_index(drop = True)
        nel_df1 = len(df.sTIME)
        df1 = df
        df1_file = pp_file
    else:
        # dft = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape')#encoding = "ISO-8859-1")
        dft = df
        dft = dft.drop(dft.index[0])
        dft = dft.reset_index(drop = True)
        dft = dft.replace(np.nan, '', regex=True)

        itmp = dft[dft['sDATE'].str.match('Summary Information:')].index[0] - 3 # df[df['model'].str.contains('ac')]        
        nel_dft = len(dft.sTIME)

        dft = dft.drop(df.index[itmp : nel_dft])
        df.dropna(axis=1, how='all', inplace=True)
        dft = dft.drop(dft[dft.sSTATUS_PATH == ''].index)
        dft = dft.drop(dft[dft[iGPS_GROUND_SPEED_cname] == ''].index)
        dft = dft.drop(dft[dft.sSTATUS_PATH == 'STANDBY'].index)
        dft = dft.drop(dft[dft.sSTATUS_PATH == 'CALIBRATION'].index)
        #dft = dft.sort_values(['sTIME'], ascending=[True])
        dft = dft.reset_index(drop = True)
        dft.index = range(len(df), len(df)  + len(dft))
        if i == 1:
            df2 = dft
        df2_file = pp_file
            
        df = pd.concat([df, dft])

df = df.drop(df[df.sSTATUS_PATH == 'STANDBY'].index)
df = df.drop(df[df.sSTATUS_PATH == 'CALIBRATION'].index)
df = df.reset_index(drop=True)
nel = len(df)

if (iENG_SPEED_cname != ''): 
    df = df.drop(df[df.iENG_SPEED == ''].index)
    df = df.reset_index(drop = True)
    df.iENG_SPEED = df.iENG_SPEED.astype(float)

if (imVEH_SPEED_cname != ''): 
    df = df.drop(df[df.imVEH_SPEED == ''].index)
    df = df.reset_index(drop = True)
    if imVEH_SPEED_unit == 'kph': df.imVEH_SPEED = df.imVEH_SPEED.astype(float)
    
if (iVEH_SPEED_cname != ''): 
    df = df.drop(df[df.iVEH_SPEED == ''].index)
    df = df.reset_index(drop = True)
    if iVEH_SPEED_unit == 'mph': df.iVEH_SPEED = 1.60934*df.iVEH_SPEED.astype(float)
    
vehspd_ecu = df.iVEH_SPEED = df.iVEH_SPEED.astype(float)
vehspd_move_start = 0
for j in range (nel):
    if j >= 9 and vehspd_move_start  == 0 and vehspd_ecu[j] > 0.5 and vehspd_ecu[j-1] < 0.1 and vehspd_ecu[j-2] < 0.01 and np.min(vehspd_ecu[j: j+15]) > 0 and np.max(vehspd_ecu[j: j+30]) <= vehspd_urban_kph :
        vehspd_move_start = j-10

if vehspd_move_start > 0: 
    df = df.drop(df.index[:vehspd_move_start])
    df = df.reset_index(drop = True)
        
vehspd_ecu = df.iVEH_SPEED = df.iVEH_SPEED.astype(float)
df_US_peak_vehspd  = df[df.iVEH_SPEED > US_max_vehspd_kph]
nindex = 0j
index = 0
index1 = 0
index2 = 0
for i in range (1, len(df_US_peak_vehspd)):
    if (df_US_peak_vehspd.index[i] - df_US_peak_vehspd.index[i-1]) > 1:
        nindex += 1
        index1 = df_US_peak_vehspd.index[i-1] 
        index2 = df_US_peak_vehspd.index[i] 
    else:
        nindex = 0
        index1 = df_US_peak_vehspd.index[0]
        index2 = df_US_peak_vehspd.index[len(df_US_peak_vehspd)-1]

vehspd_move_stop = 0
for j in range(index1, 1, -1):
    if iENG_SPEED_cname != '' and vehspd_ecu[j] > 0 and vehspd_ecu[j-1] <= 0 and df.iENG_SPEED[j-1] >= 50:
        vehspd_move_stop = j
        break
    elif iENG_SPEED_cname == '' and vehspd_ecu[j] > 0 and vehspd_ecu[j-1] <= 0:
        for k in range (500, 1, -1):
            if vehspd_ecu[j-k] <= 0 and vehspd_ecu[j-k-1] > 0.5: 
                vehspd_move_stop = j
                break

if vehspd_move_stop > 0: 
    df = df.drop(df.index[vehspd_move_stop:])
    df = df.reset_index(drop = True)
        
vehspd_ecu = df.iVEH_SPEED.astype(float)
        
# if max(df.iVEH_SPEED) > EU_peak_vehspd_kph: df = df.drop(df_US_peak_vehspd.index)
# vehspd_ecu = df.iVEH_SPEED
if (iGPS_GROUND_SPEED_cname != ''): 
    df = df.drop(df[df.iGPS_GROUND_SPEED == ''].index)
    df = df.reset_index(drop = True)
    if iGPS_GROUND_SPEED_unit == 'mph': df.iGPS_GROUND_SPEED = 1.60934 * df.iGPS_GROUND_SPEED.astype(float)
    else: df.iGPS_GROUND_SPEED = df.iGPS_GROUND_SPEED.astype(float)      
    # df_US_peak_vehspd  = df[df.iGPS_GROUND_SPEED > US_max_vehspd_kph].index
    # if iGPS_GROUND_SPEED_cname != '' and max(df.iGPS_GROUND_SPEED) > EU_peak_vehspd_kph: df = df.drop(df_US_peak_vehspd.index)

df = df.drop(df[df.sSTATUS_PATH == ''].index)
df = df.reset_index(drop=True)
nel = len(df)
    
if iGPS_GROUND_SPEED_cname != '':
    j = 0
    for i in range (nel):
        if df[iGPS_GROUND_SPEED_cname][i] == '': j += 1       
    if j >= 0.5 * nel : iGPS_GROUND_SPEED_cname = ''
    
if iGPS_GROUND_SPEED_cname != '': 
    df = df.drop(df[df.iGPS_GROUND_SPEED == ''].index)
    df = df.reset_index(drop=True)

cols = df.columns
df = df.replace(np.nan, '', regex=True)
rows = df.shape[0]
cols = df.shape[1]
if cols > 185:
    max_cols = 185
else:
    max_cols = cols
for icol in range(max_cols-1):
    cname = df.columns[icol]
    if cname == 'sDATE': sDATE_cname = cname
    elif cname == 'sSTATUS_PATH': sSTATUS_PATH_cname = cname
    elif cname == 'iGPS_ALT' : iGPS_ALT_cname = cname
    elif cname == ('iCALCRT_CO2m' or 'iCALCRT_CO2m'): iCALCRT_CO2m_cname = cname
    elif cname == 'iCALCRT_COm': iCALCRT_COm_cname = cname
    elif cname == 'iCALCRT_NOxm': iCALCRT_NOxm_cname = cname # instantenous NOx
    elif cname == 'iCALCRT_kNOxm': iCALCRT_kNOxm_cname = cname # use the corrected instantenous NOx
    elif cname == 'iSCB_LAP': iSCB_LAP_cname = cname # local ambient pressure
    elif cname == 'Lambda': Lambda_cname = cname # local ambient temperature
    elif cname[0:8] == 'Unnamed:':
        df = df.drop(columns=cname)
        cols -= 1
        iUnnamed = 1

sample_start = 0

sample_end = -1
df = df.drop(df[df.sTIME == ''].index)
df = df.reset_index(drop = True)
nel = len(df[sTIME_cname])

if iCALCRT_CO2m_cname != '':
    j = 0
    for i in range (1, nel):
        if df[iCALCRT_CO2m_cname][i] == '': j += 1
    if j > 0.1 * nel: iCALCRT_CO2m_cname = ''
if iENG_SPEED_cname != '':
    j = 0
    for i in range (1, nel):
        if df[iENG_SPEED_cname][i] == '': j += 1
    if j > 0.1 * nel: iENG_SPEED_cname = ''

if iCALCRT_CO2m_cname != '': 
    df = df.drop(df[df.iCALCRT_CO2m == ''].index)
    df = df.reset_index(drop = True)

if sample_end == -1: sample_end = len(df)   

if script_inputs_file == '': script_inputs_file = pp_file
    
pp_files = pd.DataFrame(pp_filenames, columns=['fname']).sort_values('fname')
pp_files = pp_files.reset_index(drop = True)
pp_file1 = pp_files.fname[0]
pp_filen = pp_files.fname[len(pp_files.fname)-1]
pp_filem = max(pp_filenames, key=len)
if len(pp_filem) > len(pp_filen): pp_filen = pp_filem
        
troute = pp_file1.split('.')[0]
route1 = (troute.split('-')[0]).split('_')
route1 = route1[len(route1)-1]

troute = pp_filen.split('.')[0]
route2 = troute.split('-')[1]
route = route1 + '-' + route2

testdate =''
if df.sDATE[0] != '':
    testdate = str(df.sDATE[0])#'03/29/2018'
else:
 	testdate = str(df.sDATE[1])#'03/29/2018'

vehicle_title = vehicle_title + route + ' routes @ ' + testdate

SAE_vehicle_title = ''
vehicle_title = SAE_vehicle_title

filename_combined = pp_file1.split('-')[0] + '-' + route2
if OBD == True: 
	filename_combined_obd = 'OBD_' + pp_file1.split('-')[0] + '-' + route2 + '.xlsx'
else:
	filename_combined_obd = ''

del troute, route1, route2, pp_file1, pp_filen

df = df.reset_index(drop = True)
nel = len(df[sDATE_cname])
sample_end = len(df[sDATE_cname])

if (iVEH_SPEED_cname == '') and (iGPS_GROUND_SPEED_cname != ''):
    vehspd_ecu = df.iGPS_GROUND_SPEED.astype(float)
    iGPS_GROUND_SPEED_mode = 1
elif iVEH_SPEED_cname != '': 
    vehspd_ecu = df.iVEH_SPEED.astype(float)
    iGPS_GROUND_SPEED_mode = 0

if iCALCRT_CO2m_cname != '':
    for i in range (sample_start, nel-3):
        if df[iCALCRT_CO2m_cname][i] != '' and (vehspd_ecu[i] < US_max_vehspd_mph and vehspd_ecu[i+1] > 0 and vehspd_ecu[i+2] > 0 and vehspd_ecu[i+3] > 0) and \
        (vehspd_ecu[i+1] < US_max_vehspd_mph and vehspd_ecu[i+2] < US_max_vehspd_mph and vehspd_ecu[i+3] < US_max_vehspd_mph):
            sample_start = i
            break

if np.max(vehspd_ecu) > EU_peak_vehspd_kph:
    for i in range(sample_end-2, sample_start+1, -1):
        if vehspd_ecu[i] > EU_peak_vehspd_kph and  vehspd_ecu[i-15] <10 and vehspd_ecu[i-16] < 0.5:
            for j in range (500, 0, -1):
                if vehspd_ecu[i-j] == 0 and vehspd_ecu[i-j-1] > 0.5 and  vehspd_ecu[i-j-2] > 1.5:
                    sample_end = i-j+2
                    break
                
if sample_start > 15: nel = sample_end - sample_start + 1

if iVEH_SPEED_cname != '': 
    vehspd_ecu = df.iVEH_SPEED.astype(float)
elif iGPS_GROUND_SPEED_cname != '':
    vehspd_ecu = df.iGPS_GROUND_SPEED.astype(float)
if iENG_SPEED_cname != '': engine_RPM = (df.iENG_SPEED)
    
for icol in range(nel):
    if iCOOL_TEMP_cname != '' and df.iCOOL_TEMP.all() == '': iCOOL_TEMP_cname = '' # no ECT data

etime = np.zeros(nel)
VMT = np.zeros(nel)
vehspd_flt = np.zeros(nel)
veh_accel = np.zeros(nel)
RPA = np.zeros(nel)
eVeh_accelD = np.zeros(nel)
eVeh_accelE = np.zeros(nel)
eVeh_accelP = np.zeros(nel)

Altitude = np.zeros(nel)
ECT = np.zeros(nel)
Tamb = np.zeros(nel)
Texh = np.zeros(nel)
engine_RPM = np.zeros(nel)
engine_load = np.zeros(nel)
eng_Combustion_ON = np.zeros(nel)
fuelflow_gals = np.zeros(nel)
Inst_Mass_GHG = np.zeros(nel)
Inst_Mass_CO = np.zeros(nel)
Inst_Mass_NOx = np.zeros(nel)
Corrected_Inst_Mass_NOx = np.zeros(nel)
Cumul_Mass_GHG = np.zeros(nel)
Lambda = np.zeros(nel)
CatalystTemp = np.zeros(nel)
AAT = np.zeros(nel)

if len(vehspd_ecu) > nel: 
    tmp_vehspd = vehspd_ecu
    del vehspd_ecu
    vehspd_ecu = np.zeros(nel)
    for i in range(nel): vehspd_ecu[i] = tmp_vehspd[i]
    del tmp_vehspd
   
if len(df) > nel: ntime = len(df)  
else: ntime = nel
    
if sTIME_unit != 'hhmmss': 
    for i in range (ntime):
        df.sTIME[i] = time.strftime('%H:%M:%S', time.gmtime(np.int(df.sTIME[i])))

[last_h, last_m, last_s] =str(df.sTIME[len(df.sTIME)-1]).split(':')
last_s = float(last_s.split('.')[0])
     
for i in range (nel):
    tmp_etime = timePEMS_to_second(str(df.sTIME[i]))
        
    if i == 0: etime0= tmp_etime
    etime[i] = tmp_etime - etime0
    
    if iCALCRT_CO2m_cname == '': 
        Inst_Mass_GHG[i]= 0
    elif df[iCALCRT_CO2m_cname][i] != '':
        Inst_Mass_GHG[i]= float(df[iCALCRT_CO2m_cname][i])
    if iENG_LOAD_cname != '': 
        if df[iENG_LOAD_cname][i] == '': 
            engine_load[i]= 0
        else:
            engine_load[i] = float(df[iENG_LOAD_cname][i])
    if iENG_SPEED_cname != '': 
        if df[iENG_SPEED_cname][i] == '':
            engine_RPM[i]= 0
        else:
            engine_RPM[i]= float(df[iENG_SPEED_cname][i])
    if engine_RPM[i]>= 500: eng_Combustion_ON[i]= 1
    if imCOOL_TEMP_cname != '' and df[imCOOL_TEMP_cname][i] != '':
        if ECT_unit == 'deg F': 
            ECT[i]= 5/9*(float(df[iCOOL_TEMP_cname][i]) - 32)
        else:
            ECT[i]= float(df[imCOOL_TEMP_cname][i])
    if Lambda_cname != '' and df[Lambda_cname][i] != '':
        Lambda[i] = float(df[Lambda_cname][i])
    if CatalystTemp_cname != '' and df[CatalystTemp_cname][i] != '':
        if CatalystTemp_unit == 'deg F': 
            CatalystTemp[i] = 5/9*(float(df[CatalystTemp_cname][i]) - 32)
        else: CatalystTemp[i] = float(df[CatalystTemp_cname][i])
    if Tamb_cname != '' and df[Tamb_cname][i] != '': 
        if Tamb_unit == 'deg F': 
            Tamb[i]= 5/9*(float(df[Tamb_cname][i]) - 32)
        else:
            Tamb[i]= float(df[Tamb_cname][i])
            
    if iWfgps_cname != '': 
        if df[iWfgps_cname][i] == '': 
            fuelflow_gals[i]= 0
        elif iWfgps_unit == 'gal/s':
            fuelflow_gals[i]= float(df[iWfgps_cname][i])
        elif iWfgps_unit == 'g/s':
            fuelflow_gals[i]= float(df[iWfgps_cname][i])/2834.89
            
    Inst_Mass_GHG[i]= float(df[iCALCRT_CO2m_cname][i])
    Inst_Mass_CO[i]= float(df[iCALCRT_COm_cname][i])
    Inst_Mass_NOx[i]= float(df[iCALCRT_NOxm_cname][i])
    if iCALCRT_kNOxm_cname != '': Corrected_Inst_Mass_NOx[i]= float(df[iCALCRT_kNOxm_cname][i])
    
    if i > 0: Cumul_Mass_GHG[i]= Cumul_Mass_GHG[i-1]+ Inst_Mass_GHG[i]
    Altitude[i]= float(df[iGPS_ALT_cname][i])

    if i > 0 and ((iENG_SPEED_cname != '' and float(df[iENG_SPEED_cname][i]) < 50) or vehspd_ecu[i] < 0.2): Altitude[i]= float(Altitude[i-1])

    if iGPS_GROUND_SPEED_cname != '' and df[iGPS_GROUND_SPEED_cname][i] <= 0.5 and vehspd_ecu[i]> 10: vehspd_ecu[i]= 0
    if vehicle_type == 1 and iENG_SPEED_cname != '' and OBD == FALSE and iGPS_GROUND_SPEED_mode == 1 and engine_RPM[i]< 50 and vehspd_ecu[i]> 1: 
        vehspd_ecu[i]= 0.0
        
    if i > 3 and i < nel - 2 and vehspd_ecu[i]< vehspd_1_kph and vehspd_ecu[i-1]< vehspd_1_kph and vehspd_ecu[i+1] < vehspd_1_kph and vehspd_ecu[i+2] < vehspd_1_kph: vehspd_ecu[i]= 0
       
enum_MAW_vehspd = i

if etime[len(etime) - 1] <=  etime[len(etime) - 2]: etime[len(etime) - 1] = etime[len(etime) - 2] + 1
for j in range (nel):
    if j > 3 and j < nel -1 and vehspd_ecu[j] < 0.01 and Inst_Mass_GHG[j] > 0.5 and vehspd_ecu[j-1] > 0.5 and vehspd_ecu[j+1] > 0.5:
        vehspd_ecu[j] = 0.5*(vehspd_ecu[j-1] + vehspd_ecu[j+1])
        
for j in range (len(etime)-1):
    if ((etime[j+1] - etime[j]) > 120):
        if (iENG_SPEED_cname != '' and (etime[j+1] - etime[j]) > 300 and engine_RPM[j] > 0): engine_RPM[j] = 0
        if iENG_LOAD_cname != '' and (etime[j+1] - etime[j]) > 300 and engine_load[j+1] >= 0:
            engine_load[j+1] = 0
            engine_load[j] = 0
        if iCOOL_TEMP_cname != '' and (etime[j+1] - etime[j]) > 300 and ECT[j+1] > 10:
            ECT[j+1] = 0
            ECT[j] = 0
        if iCOOL_TEMP_cname != '' and (etime[j+1] - etime[j]) > 300 and ECT[j] > 10 and ECT[j+1] < 10:
            ECT[j+1] = 0
            ECT[j] = 0
        if iENG_SPEED_cname != '' and (etime[j+1] - etime[j]) > 120 and engine_RPM[j+1] == 0  and engine_RPM[j] > 0: engine_RPM[j] = 0
        if iENG_SPEED_cname != '' and (etime[j+1] - etime[j]) > 120 and engine_RPM[j] == 0  and engine_RPM[j+1] > 0: engine_RPM[j+1] = 0

df = df.reset_index(drop = True)
df.to_excel(filename_combined + '.xlsx', index=False) #encoding='utf8') # 'utf-8-sig')

'''
engtrq = np.interp(engspd, engtrq_RPM, engtrqWOT_NM)
'''
engtrq_cols = [col for col in df.columns if 'iENG_TORQ' in col]
if len(engtrq_cols) > 0: 
    engtrq_NM = df.iENG_TORQ = df['iENG_TORQ'].astype(float)
    for i in range(len(df.iENG_TORQ)):
        if df.iENG_TORQ[i] < 0: engtrq_NM[i] = 0

if calc_NOx_gbhp_hr == True and len(engtrq_RPM) > 1 and len(engtrqWOT_NM) > 1: 
    engtrq_NM = 0.01 * engine_load * np.interp(engine_RPM, engtrq_RPM, engtrqWOT_NM)
elif calc_NOx_gbhp_hr == True and np.max(engine_load) > 0:
    engtrq_NM = 0.01 * engine_load * engine_peak_trq

if (calc_NOx_gbhp_hr == True) or (len(engtrq_cols) > 0):    
    engpwr_KW = engtrq_NM * engine_RPM * np.pi/30/1000
    engpwr_HP = engpwr_KW/0.7457
    Inst_Mass_NOx_gbhp_hr = 3600*Inst_Mass_NOx/engpwr_HP
    for k in range(len(engpwr_HP)):
        if Inst_Mass_NOx_gbhp_hr[k] < 0: Inst_Mass_NOx_gbhp_hr[k] = 0
        
raw_Altitude = Altitude
Altitude = savgol_filter(raw_Altitude, window_length=61, polyorder=3)
VMT = np.zeros(nel)
dAltitude = np.zeros(nel)
    
for j in range (1, nel-1):
    VMT[j]= VMT[j-1]+ vehspd_ecu[j]/(3.6)
    veh_accel[j]= (vehspd_ecu[j+1]- vehspd_ecu[j-1])/(2*3.6)
    if veh_accel[j]> 0.1: RPA[j]= veh_accel[j]    
    if vehspd_ecu[j] > MAW_vehspd_0_kph: dAltitude[j] = (Altitude[j] - Altitude[j-1])
    
if VMT[nel-1] < VMT[nel-2]: VMT[nel-1] = VMT[nel-2]

if OBD == TRUE:
    [h2, m2, s2] = str(df2.sTIME[df2.index[0]]).split(':')
    h2 = float(h2)
    m2 = float(m2)
    s2 = float(s2.split('.')[0])
    
    dff_obd = pd.DataFrame(csv_filenames, columns = ['OBD'])
    dff_obd = dff_obd.sort_values('OBD')
    dff_obd = dff_obd.reset_index(drop=True)
    nobd = len (dff_obd)
    for i in range (nobd):
        if i <= len (dff_obd) :
            if (dff_obd.OBD[i][0:3] == 'TA_'):
                dff_obd = dff_obd.drop(dff_obd.index[i])
                dff_obd = dff_obd.reset_index(drop = True)
            elif (dff_obd.OBD[i][0:3] == 'pp_'):
                dff_obd = dff_obd.drop(dff_obd.index[i:len(dff_obd)])
                if dff_obd.OBD[i-1][0:3] == 'pp_': dff_obd = dff_obd.drop(dff_obd.index[i-1])
                dff_obd = dff_obd.reset_index(drop = True)

    for i in range (len(dff_obd)):
        obd_file = str(dff_obd.OBD[i])
        if i == 0:
            df_obd = pd.read_csv(obd_file, skiprows=0, encoding = 'unicode_escape')#encoding = "ISO-8859-1")
            df_obd = df_obd.replace(np.nan, '', regex=True)
            if i == 0:
                df1_obd = df_obd
            df1_obd_file = obd_file

            nel_obd1 = len(df_obd)
        else:
            dft = pd.read_csv(obd_file, skiprows=0, encoding = 'unicode_escape')#encoding = "ISO-8859-1")
            dft = dft.replace(np.nan, '', regex=True)            
            dft.index = range(len(df_obd), len(df_obd)  + len(dft))
            df_obd = pd.concat([df_obd, dft])
            if i == 1:
                df2_obd = dft
            df2_obd_file = obd_file
    
    df_obd = df_obd.sort_values(['Clock_T(hms)'], ascending=[True])
    df_obd = df_obd.reset_index(drop=True)
        
    ovehspd = df_obd['SPD(km/h)'].astype(float)
    oengtrq = df_obd['TQRFIN(Nm)'].astype(float)
    oengrpm = df_obd['NE(rpm)'].astype(float)
    oengect = df_obd['THW(C)'].astype(float)
    oambairT= df_obd['THA(C)'].astype(float)
    oexhtmp = df_obd['THEXH2(C)'].astype(float)
    hms_obd = df_obd['Clock_T(hms)']

    [hh0, mm0, ss0, otime0] = timeOBD_to_second(str(hms_obd[0]))

    last_obd_found = 0
    nhms_obd = len(hms_obd)
    last_obd_index = nhms_obd
    otime = np.zeros(nhms_obd)
    for i in range (nhms_obd):
        [hh, mm, ss, otimei] = timeOBD_to_second(str(hms_obd[i]))
        otime[i] = otimei - otime0
        
        if (last_obd_found == 0) and float(hh) > float(last_h) and (float(mm) +30) >= float(last_m) and float(ss) >= float(last_s):
            last_obd_index = i
            last_obd_found = 1

    if nhms_obd > last_obd_index: df_obd = df_obd.drop(df_obd.index[last_obd_index:nhms_obd])

    vehspd1 = vehspd_ecu
    etime1 = etime
    ovehspd1 = ovehspd
    otime1 = otime
            
    timeoffset_obd = 0
    timeoffset_pems = 0
    itime_start = 0
    ndoe = 500
    nsamples_obd_sync = 600
    [timeoffset_pems, timeoffset_obd] = vehspd_obd_sync(ndoe, nsamples_obd_sync, itime_start, etime, vehspd_ecu, otime, ovehspd.values)

    timeoffset_pems0 = timeoffset_pems
    timeoffset_obd0 = timeoffset_obd
    
    vehspd2 = 1.60394*df2.iGPS_GROUND_SPEED.astype(float)       
    ovehspd2 = df2_obd['SPD(km/h)'].astype(float)

    etime10 = timePEMS_to_second(str(df1.sTIME[df1.index[0]]))
    etime11 = timePEMS_to_second(str(df1.sTIME[df1.index[len(df1)-1]]))
    etime20 = timePEMS_to_second(str(df2.sTIME[df2.index[0]]))
    etime21 = timePEMS_to_second(str(df2.sTIME[df2.index[len(df2)-1]]))
    delta_etime21 = etime20 - etime10

    [hh10, mm10, ss10, otime10] = timeOBD_to_second(str(df1_obd['Clock_T(hms)'][df1_obd.index[0]]))
    [hh11, mm11, ss11, otime11] = timeOBD_to_second(str(df1_obd['Clock_T(hms)'][df1_obd.index[len(df1_obd)-1]]))
    [hh20, mm20, ss20, otime20] = timeOBD_to_second(str(df2_obd['Clock_T(hms)'][df2_obd.index[0]]))
    [hh21, mm21, ss21, otime21] = timeOBD_to_second(str(df2_obd['Clock_T(hms)'][df2_obd.index[len(df2_obd)-1]]))
    delta_otime21 = otime20 - otime10
            
    etime = etime + timeoffset_pems
    otime = otime + timeoffset_obd
    
    vehspd2 = vehspd2.reset_index(drop=True)
    ovehspd2 = ovehspd2.reset_index(drop=True)

    iovehspd_GT0 = 0
    ivehspd_GT0 = 0
    for i in range (len(ovehspd2)):
        if ovehspd2[i] > 0 and iovehspd_GT0 == 0: iovehspd_GT0 = i
    for i in range (len(vehspd2)):
        if vehspd2[i] > 0 and ivehspd_GT0 == 0: ivehspd_GT0 = i
    
    ivehspd_ecu = 0       
    for i in range (nel_df1-500, nel_df1+1000):
        if ivehspd_ecu == 0 and abs(vehspd_ecu[i]) > 0.5 and abs(vehspd_ecu[i+1]) > 0.5 and abs(vehspd_ecu[i+2]) > 0.5:
            ivehspd_ecu = i
            
    ovehspd = df_obd['SPD(km/h)'].astype(float)
    oengtrq = df_obd['TQRFIN(Nm)'].astype(float)
    oengrpm = df_obd['NE(rpm)'].astype(float)
    oengect = df_obd['THW(C)'].astype(float)
    if max(oengect) > 300: oengect-= 273.5
    oambairT= df_obd['THA(C)'].astype(float)
    oexhtmp = df_obd['THEXH2(C)'].astype(float)
    hms_obd = df_obd['Clock_T(hms)'].astype(float)
    otime = otime[0:len(df_obd)]

    for j in range (len(otime)-1):
        if ((otime[j+1] - otime[j]) > 120):
            if oengtrq[j+1] > 0: oengtrq[j+1] = 0
            if (otime[j+1] - otime[j]) > 300 and oengrpm[j] > 0 : oengrpm[j] = 0
            if (otime[j+1] - otime[j]) > 300 and oengect[j+1] > 10 :
                oengect[j+1] = 0
                oengect[j] = 0
            if (otime[j+1] - otime[j]) > 120 and oengrpm[j+1] == 0  and oengrpm[j] > 0: oengrpm[j] = 0
            if (otime[j+1] - otime[j]) > 120 and oengrpm[j] == 0  and oengrpm[j+1] > 0: oengrpm[j+1] = 0

    df_obd['OBD_time(s)'] = otime
    df_obd['OBD_engrpm'] = oengrpm
    df_obd['OBD_engtrq_NM'] = oengtrq
    df_obd['OBD_ECT_C'] = oengect
    df_obd.to_excel(filename_combined_obd, index=False) #, encoding='utf-8-sig')
    
for i in range (1, nel): 
	if VMT[i] < VMT[i-1]: VMT[i] = VMT[i-1]
###################

elapsed = time.time() - t
print("\n*************\nElapsed Time: {:5.0f} Seconds before plotting and saving to PDF" .format(elapsed))

#t = time.time()
here = os.path.dirname(os.path.realpath("__file__"))
                            
nVu = 0
nVr = 0
nVm = 0
vehmove_start_ini = 0
time_engine_start = 0
time_engine_idle_int = 0
time_cold_start = 0
time_cold_start_end = 0

cold_start= 1 
cold_start_end=0     
for i in range(nel):
    if (iENG_SPEED_cname != ''):
        # if time_engine_start == 0 and eng_Combustion_ON[i] == 1 and eng_Combustion_ON[i-1] == 0 and vehmove_start_ini == 0:
        #     time_engine_start = i

        if i > 0 and time_engine_idle_int == 0 and vehspd_ecu[i] > 0.5 and vehspd_ecu[i-1] <= 0.01 and vehmove_start_ini == 0 and i > time_engine_start:
#            time_engine_idle_int = i - time_engine_start
            time_engine_idle_int = i

        if abs(ECT[i] - (Tamb[i])) <= 7.0 and ECT[i] <= 35 and vehmove_start_ini > 0 and time_cold_start == 0 and cold_start==1:
            cold_start = 0
            time_cold_start = i
    
        if cold_start_end == 0 and ECT[i] >= 70 and i > 5:
            cold_start_end = 1
            time_cold_start_end =i
    else:
        time_engine_start = 1
        time_engine_idle_int = 0
        cold_start_end = 1
        time_cold_start_end =300

    if i > 0 and vehspd_ecu[i] > 0.5 and vehspd_ecu[i-1] <= 0.01 and vehmove_start_ini == 0 and i > time_engine_start:
        vehmove_start_ini = i

    if (time_cold_start_end - time_cold_start+1) > 300 and (cold_start_end == 1): 
        time_cold_start_end = 300 + time_cold_start-1        
# Urban, Rural and Motorway vehicle speeds at EU regulation 2017/1151 page 165
    if vehspd_ecu[i] <= vehspd_urban_kph and vehspd_ecu[i] >= 0: # > 0.03
        nVu = nVu +1
    elif vehspd_ecu[i] > vehspd_rural_kph:
        nVm = nVm + 1
    elif vehspd_ecu[i] > vehspd_urban_kph and vehspd_ecu[i] <= vehspd_rural_kph:
        nVr = nVr + 1
if iCOOL_TEMP_cname == '': time_cold_start_end = 300
                
Vu = np.zeros(nVu)
Vr = np.zeros(nVr)
Vm = np.zeros(nVm)
Vuf = np.zeros(nVu)
Vrf = np.zeros(nVr)
Vmf = np.zeros(nVm)
Vui = np.zeros(nVu).astype(int)
Vri = np.zeros(nVr).astype(int)
Vmi = np.zeros(nVm).astype(int)
vehspd_accel = np.zeros(nel)
dVMT = np.zeros(nel)
veh_ACC = np.zeros(nel)

vehspd_ACCu = np.zeros(nVu)
vehspd_ACCr = np.zeros(nVr)
vehspd_ACCm = np.zeros(nVm)
vehspd_ACCu1 = np.zeros(nVu)
vehspd_ACCr1 = np.zeros(nVr)
vehspd_ACCm1 = np.zeros(nVm)
veh_ACCu = np.zeros(nVu)
Vr_ACC = np.zeros(nVr)
Vm_ACC = np.zeros(nVm)
Vu_ACC = np.zeros(nVu)
Vr_ACC = np.zeros(nVr)
Vm_ACC = np.zeros(nVm)
Vu_ACC1 = np.zeros(nVu)
Vr_ACC1 = np.zeros(nVr)
Vm_ACC1 = np.zeros(nVm)
dVMTu = np.zeros(nVu)
dVMTr = np.zeros(nVr)
dVMTm = np.zeros(nVm)
TIMEu = np.zeros(nVu)
TIMEr = np.zeros(nVr)
TIMEm = np.zeros(nVm)

pACCu_jel = np.zeros(nVu)
pACCr_jel = np.zeros(nVr)
pACCm_jel = np.zeros(nVm)
Nku_iel = np.zeros(nVu)
Nkr_iel = np.zeros(nVr)
Nkm_iel = np.zeros(nVm)
pACCu = np.zeros(nVu)
pACCr = np.zeros(nVr)
pACCm = np.zeros(nVm)

iVu = 0
iVr = 0
iVm = 0
vehspd_145kph = 0
vehspd_100kph = 0
vehspd_zero = 0

time_engine_idle=0

for i in range(1, nel-1):
    et = etime[i+1] - etime[i-1]
    veh_ACC[i] = (vehspd_ecu[i+1] - vehspd_ecu[i-1])/(et*3.6)# 1Hz 1 second time inteval
    vehspd_accel[i] = vehspd_ecu[i] * veh_ACC[i]/3.6
            
    dVMT[i] = (VMT[i+1] - VMT[i])
    if vehspd_ecu[i] <=  vehspd_urban_kph and vehspd_ecu[i] >= 0: # for EMROAD vehspd > =1 and vehspd < 60 kph
        Vu[iVu]=vehspd_ecu[i]
        dVMTu[iVu] = dVMT[i]
        TIMEu[iVu] = etime[i]
        iVu = iVu +1
        if i == (nel-2): TIMEu[iVu] = etime[i] + 1
        if vehspd_ecu[i] <= vehspd_1_kph/20:    # for EMROAD vehspd > =1 and vehspd < 60 kph, otherwise <= 0.2
            vehspd_zero = vehspd_zero + 1
            time_engine_idle = time_engine_idle + (etime[i+1] - etime[i])
    elif vehspd_ecu[i] > vehspd_rural_kph:
        Vm[iVm]=vehspd_ecu[i]
        dVMTm[iVm] = dVMT[i]
        TIMEm[iVm] = etime[i]
        iVm = iVm + 1
        if vehspd_ecu[i] > EU_vehspd_100kph:
            vehspd_100kph = vehspd_100kph + 1
        if vehspd_ecu[i] > EU_peak_vehspd_kph:
            vehspd_145kph = vehspd_145kph + 1        
    elif vehspd_ecu[i] > vehspd_urban_kph and vehspd_ecu[i] <= vehspd_rural_kph:
        Vr[iVr]=vehspd_ecu[i]
        dVMTr[iVr] = dVMT[i]
        TIMEr[iVr] = etime[i]
        iVr = iVr + 1

if TIMEu[nVu-1] == 0: TIMEu[nVu-1] = TIMEu[nVu-2] + 1

# EU JRC EMROAD tool, 2-second time interval central difference method
et2 = 2
for i in range(1, nVu-1):
    #et2 = TIMEu[i+1] - TIMEu[i-i]
    Vu_ACC[i] = (Vu[i+1] - Vu[i-1])/(et2*3.6)
    if Vu[i] < 1: Vu_ACC[i] = 0
    vehspd_ACCu[i] = Vu[i] * Vu_ACC[i]/3.6

for i in range(1, nVr-1):
    #et2 = TIMEr[i+1] - TIMEr[i-1]
    Vr_ACC[i] = (Vr[i+1] - Vr[i-1])/(et2*3.6)
    vehspd_ACCr[i] = Vr[i]* Vr_ACC[i]/3.6

for i in range(1, nVm-1):
    #et2 = TIMEm[i+1] - TIMEm[i-1]
    Vm_ACC[i] = (Vm[i+1] - Vm[i-1])/(et2*3.6)
    vehspd_ACCm[i] = Vm[i] * Vm_ACC[i]/3.6

# 1-second time interval PKE method
et1 = 1
for i in range(nVu-1):
    #et1 = TIMEu[i+1] - TIMEu[i]
    Vu_ACC1[i] = (Vu[i+1] - Vu[i])/(et1*3.6)
    if Vu[i] < 1: Vu_ACC1[i] = 0
    vehspd_ACCu1[i] = (Vu[i+1] + Vu[i])/2 * Vu_ACC1[i]/3.6
    
for i in range(nVr-1):
    #et1 = TIMEr[i+1] - TIMEr[i]
    Vr_ACC1[i] = (Vr[i+1] - Vr[i])/(et1*3.6)
    vehspd_ACCr1[i] = (Vr[i+1] + Vr[i])/2 * Vr_ACC1[i]/3.6

for i in range(nVm-1):
    #et1 = TIMEm[i+1] - TIMEm[i]
    Vm_ACC1[i] = (Vm[i+1] - Vm[i])/(et1*3.6)
    vehspd_ACCm1[i] = (Vm[i+1] + Vm[i])/2 * Vm_ACC1[i]/3.6
    
#plt.plot(Inst_Mass_GHG[vehspd_ecu< MAW_vehspd_0_kph])
Inst_Mass_GHG[vehspd_ecu<=MAW_vehspd_0_kph] = 0
         
# plt.plot(vehspd_ACCu, 'b-')   
# plt.plot(vehspd_ACCu1, 'r-.')   
# plt.plot(vehspd_ACCm, 'b-')   
# plt.plot(vehspd_ACCm1, 'r-.')   

# # EU JRC EMROAD tool 2-second time interval centeral difference method, veh_accel > 0.1
# np.percentile(vehspd_ACCu[Vu_ACC > 0.1], 95)
# np.percentile(vehspd_ACCr[Vr_ACC > 0.1], 95)
# np.percentile(vehspd_ACCm[Vm_ACC > 0.1], 95)
# # PKE method, 1-second time interval 1Hz data for faster acceleration, veh_accel > 0
# np.percentile(vehspd_ACCu1[Vu_ACC1 > 0], 95)
# np.percentile(vehspd_ACCr1[Vr_ACC1 > 0], 95)
# np.percentile(vehspd_ACCm1[Vm_ACC1 > 0], 95)

num_stops=1
for i in range(vehmove_start_ini+1, nel):
    if vehspd_ecu[i] > 0 and vehspd_ecu[i-1] <= 0 and np.sum(Inst_Mass_GHG[i:nel-1]) > GHG_WLTC_threshold:
        num_stops = num_stops + 1
    
vehmove_start = np.zeros(num_stops)
vehmove_start[0] = vehmove_start_ini
vehmove_end = np.zeros(num_stops)

j=0
k=0
for i in range(vehmove_start_ini, nel):
    if vehspd_ecu[i] > 0 and vehspd_ecu[i-1] <= 0:
        if np.sum(Inst_Mass_GHG[i:nel-1]) > GHG_WLTC_threshold :
            vehmove_start[j]=i
            j = j + 1
    elif vehspd_ecu[i] <= 0 and vehspd_ecu[i-1] > 0:
        if np.sum(Inst_Mass_GHG[i:nel-1]) > GHG_WLTC_threshold :
            vehmove_end[k]=i-1
            k = k+1
                       
if max(vehmove_start) > max(vehmove_end):
    for m in range (int(max(vehmove_start))+1, nel-1):
        if sum(Inst_Mass_GHG[m:nel-1]) < GHG_WLTC_threshold and vehspd_ecu[m] > 0:
            vehmove_end[k]=m-1
            break

num_vehmove_start = j-1
num_vehmove_end = k

GHG_start = np.zeros(nel)
GHG_end = np.zeros(nel)
GHG_exclude_pts = np.zeros(nel)

GHG_MAW = np.zeros(nel)  
CO_MAW = np.zeros(nel)  
NOx_MAW = np.zeros(nel)  
tMAW_engRPM = np.zeros(nel)
tMAW_vehspd = np.zeros(nel)
tMAW_GHG_gpkm = np.zeros(nel)
tMAW_GHG_g = np.zeros(nel)
tMAW_CO_gpkm = np.zeros(nel)
tMAW_NOx_gpkm = np.zeros(nel)
tMAW_ECT = np.zeros(nel)
tMAW_Texh = np.zeros(nel)
tMAW_VA_95pct = np.zeros(nel)

if max(vehspd_ecu) < 2: 
    print('max(vehspd) < 2')
    sys.exit(0)

kel = 0      
for i in range(0, num_vehmove_start+1):
    k = int(vehmove_start[i])
    l = int(vehmove_end[i])
    for iel in range(k, l+1):
        start = GHG_start[iel] = iel
        total_Inst_Mass_GHG = 0
        total_Inst_Mass_CO = 0
        total_Inst_Mass_NOx = 0
        for jel in range (start, nel):      
            if vehspd_ecu[jel] >= 0.1 and jel >= time_cold_start_end:
                total_Inst_Mass_GHG = total_Inst_Mass_GHG + Inst_Mass_GHG[jel]
                total_Inst_Mass_CO = total_Inst_Mass_CO + Inst_Mass_CO[jel]
                total_Inst_Mass_NOx = total_Inst_Mass_NOx + Inst_Mass_NOx[jel]
                if total_Inst_Mass_GHG > GHG_WLTC_threshold :
                    GHG_MAW[iel] = total_Inst_Mass_GHG
                    CO_MAW[iel] = total_Inst_Mass_CO
                    NOx_MAW[iel] = total_Inst_Mass_NOx
                    if jel == GHG_end[iel-1]:
                        end = GHG_end[iel] = jel+1
                    else:
                        end = GHG_end[iel] = jel                    
                        break
        if max(GHG_end) == 0: 
            print('Max (GHG_end) = 0 or max(vehspd) < 10')
            sys.exit(0)
            
        for ipts in range (start, end):
            if vehspd_ecu[ipts] <= 0: npts_exclude = GHG_exclude_pts[iel] = GHG_exclude_pts[iel] +1

        if end < start or end==0:
            GHG_start[iel] = 0
            GHG_exclude_pts[iel] = 0
        elif iel >= time_cold_start_end and np.sum(vehspd_ecu[start:end]) > 0:   
            tMAW_vehspd[kel] = np.sum(vehspd_ecu[start:end])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_engRPM[kel] = np.sum(engine_RPM[start:end])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_ECT[kel] = np.sum(ECT[start:end])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_Texh[kel] = np.sum(Texh[start:end])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_GHG_gpkm[kel] = GHG_MAW[iel]*1000/ np.sum(dVMT[start:end])
            tMAW_CO_gpkm[kel] = CO_MAW[iel]*1000/ np.sum(dVMT[start:end])
            tMAW_NOx_gpkm[kel] = NOx_MAW[iel]*1000/ np.sum(dVMT[start:end])
            tMAW_GHG_g[kel] = total_Inst_Mass_GHG
            kel += 1

num_MAW_vehspd = kel
MAW_vehspd = tMAW_vehspd[:num_MAW_vehspd]
MAW_GHG_gpkm = tMAW_GHG_gpkm[:num_MAW_vehspd]
MAW_CO_gpkm = tMAW_CO_gpkm[:num_MAW_vehspd]
MAW_NOx_gpkm = tMAW_NOx_gpkm[:num_MAW_vehspd]

MAWu_vehspd = np.zeros(num_MAW_vehspd)
MAWu_GHG_gpkm = np.zeros(num_MAW_vehspd)
MAWu_CO_gpkm = np.zeros(num_MAW_vehspd)
MAWu_NOx_gpkm = np.zeros(num_MAW_vehspd)
MAWr_vehspd = np.zeros(num_MAW_vehspd)
MAWr_GHG_gpkm = np.zeros(num_MAW_vehspd)
MAWr_CO_gpkm = np.zeros(num_MAW_vehspd)
MAWr_NOx_gpkm = np.zeros(num_MAW_vehspd)

j=0
nMAWu = 0
nMAWr = 0
nMAWm = 0
for i in range(num_MAW_vehspd):
    if MAW_vehspd[i] <= MAW_urban_vehspd_kph:
        MAWu_vehspd[nMAWu] = MAW_vehspd[i]
        MAWu_GHG_gpkm[nMAWu] = MAW_GHG_gpkm[i]
        MAWu_CO_gpkm[nMAWu] = MAW_CO_gpkm[i]
        MAWu_NOx_gpkm[nMAWu] = MAW_NOx_gpkm[i]
        nMAWu = nMAWu + 1
    elif MAW_vehspd[i] > MAW_urban_vehspd_kph and MAW_vehspd[i] <= MAW_rural_vehspd_kph:
        MAWr_vehspd[nMAWr] = MAW_vehspd[i]
        MAWr_GHG_gpkm[nMAWr] = MAW_GHG_gpkm[i]
        MAWr_CO_gpkm[nMAWr] = MAW_CO_gpkm[i]
        MAWr_NOx_gpkm[nMAWr] = MAW_NOx_gpkm[i]        
        nMAWr = nMAWr + 1
    elif MAW_vehspd[i] > MAW_rural_vehspd_kph:
        nMAWm = nMAWm + 1

TIMEt = nVu + nVr + nVm
MAWu_vehspd = MAWu_vehspd[:nMAWu]
MAWu_GHG_gpkm = MAWu_GHG_gpkm[:nMAWu]
MAWu_CO_gpkm = MAWu_CO_gpkm[:nMAWu]
MAWu_NOx_gpkm = MAWu_NOx_gpkm[:nMAWu]

MAWr_vehspd = MAWr_vehspd[:nMAWr]
MAWr_GHG_gpkm = MAWr_GHG_gpkm[:nMAWr]
MAWr_CO_gpkm = MAWr_CO_gpkm[:nMAWr]
MAWr_NOx_gpkm = MAWr_NOx_gpkm[:nMAWr]
mean_Vu = np.mean(vehspd_ecu[vehspd_ecu <= vehspd_urban_kph])

rms_dAltitude = 'RMS dAltitude = ' + str(round(np.sqrt(np.mean(dAltitude**2)), 2)) + 'm'

rgrade = np.zeros(nel)
slope = np.zeros(nel)

for i in range(0, nel):
    if i == 0: 
        slope[0] = 0
        rgrade[0] = 0
    else:
        if dVMT[i] < 0.01:
            slope[i] = 0
            rgrade[i] = 0
        else:
            slope[i] = dAltitude[i]/dVMT[i]
            if vehspd_ecu[i] < 0.1: slope[i] = 0
            
            if slope[i] > 0.7:
                slope[i] = 0.7
            elif slope[i] < -0.7:
                slope[i] = -0.7
        
            rgrade[i] = 100* np.tan(np.arcsin(slope[i]))
            if rgrade[i] > max_grade_pct:
                rgrade[i] = max_grade_pct
            elif rgrade[i] < -max_grade_pct:
                rgrade[i] = -max_grade_pct

#rms = np.sqrt(np.mean(y**2))
rgrade_savgol = savgol_filter(rgrade, window_length=61, polyorder=3)
rms_road_grade = 'RMS road raw grade = ' + str(round(np.sqrt(np.mean(rgrade**2)), 2)) + '%'
rms_road_grade_pos = 'RMS road positive raw grade = ' + str(round(np.sqrt(np.mean(rgrade[rgrade>0]**2)), 2)) + '%'
  
rms_road_grade_savgol  =  'RMS Road Grade filtered' + str(round(np.sqrt(np.mean(rgrade_savgol**2)), 3)) + '%'
rms_road_grade_savgol_pos =  'RMS Road Positive Grade filtered' + str(round(np.sqrt(np.mean(rgrade_savgol[rgrade_savgol>0]**2)), 2)) + '%'

ehspd_ACCu_95pct = 0
vehspd_ACCr_95pct = 0
vehspd_ACCm_95pct = 0
Vu_ACC_95pct = 0
Vr_ACC_95pct = 0
Vm_ACC_95pct = 0
RPAu = 0
RPAr = 0
RPAm = 0

a_pos_limit_2sec = 0.1
a_pos_limit_PKE = 0.0

#if nVu > 0 and nVr > 0 and nVm > 0: 
[Vu_mean, Vr_mean, Vm_mean] = Vehspd_mean(Vu, Vr, Vm)
[vehspd_ACCu_95pct, vehspd_ACCr_95pct, vehspd_ACCm_95pct] = VehSpd_Accel_95pct(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, Vu_ACC, Vr_ACC, Vm_ACC, a_pos_limit_2sec)
[vehspd_ACCu1_95pct, vehspd_ACCr1_95pct, vehspd_ACCm1_95pct] = VehSpd_Accel_95pct(vehspd_ACCu1, vehspd_ACCr1, vehspd_ACCm1, Vu_ACC1, Vr_ACC1, Vm_ACC1, a_pos_limit_PKE)
[Vu_ACC_95pct, Vr_ACC_95pct, Vm_ACC_95pct] = VehSpd_Accel_95pct(Vu*Vu_ACC/3.6, Vr*Vr_ACC/3.6, Vm*Vm_ACC/3.6, Vu_ACC, Vr_ACC, Vm_ACC, a_pos_limit_2sec)
[RPAu, RPAr, RPAm] = Relative_Positive_Acceleration(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, dVMTu, dVMTr, dVMTm, Vu_ACC, Vr_ACC, Vm_ACC, a_pos_limit_2sec)
[RPAu1, RPAr1, RPAm1] = Relative_Positive_Acceleration(vehspd_ACCu1, vehspd_ACCr1, vehspd_ACCm1, dVMTu, dVMTr, dVMTm, Vu_ACC1, Vr_ACC1, Vm_ACC1, a_pos_limit_PKE)

vehspd_avg = [Vu_mean, Vr_mean, Vm_mean]
VA_avg_95pct = [vehspd_ACCu_95pct, vehspd_ACCr_95pct, vehspd_ACCm_95pct]
VA1_avg_95pct = [vehspd_ACCu1_95pct, vehspd_ACCr1_95pct, vehspd_ACCm1_95pct]
VA_urm_95pct = [Vu_ACC_95pct, Vr_ACC_95pct, Vm_ACC_95pct]

RPA_avg = [RPAu, RPAr, RPAm]
RPA1_avg = [RPAu1, RPAr1, RPAm1]

trip_distance = round(VMT[nel-1]/1000, 2)
trip_duration = round(TIMEt/60, 2)
if iCOOL_TEMP_cname == '':
    cold_start_duration = 300/60
else:
    cold_start_duration = round((time_cold_start_end-time_engine_start)/60, 2)

urban_distance = round(np.sum(dVMTu)/1000, 2)
rural_distance = round(np.sum(dVMTr)/1000, 2)
motorway_distance = round(np.sum(dVMTm)/1000, 2)

urban_distance_share = round(urban_distance/trip_distance*100, 2)
rural_distance_share = round(rural_distance/trip_distance*100, 2)
motorway_distance_share = round(motorway_distance/trip_distance*100, 2)

Vt = (np.sum(Vu) + np.sum(Vr) + np.sum(Vm))/(nVu + nVr + nVm)
Vtm = round(np.mean(vehspd_ecu[vehspd_ecu >= vehspd_0_kph]), 2)

urban_avg_speed = round(Vu_mean, 2)
rural_avg_speed = round(Vr_mean, 2)
motorway_avg_speed = round(Vm_mean, 2)
total_trip_avg_speed = round(Vt, 2)

if nVm == 0: motorway_speed_above_145kph = 0
else: motorway_speed_above_145kph = round(vehspd_145kph/nVm*100, 2)
    
if nVr == 0: motorway_speed_above_100kph = 0
else: motorway_speed_above_100kph = round(vehspd_100kph/60, 2)
urban_stop_time_pct = round(vehspd_zero/nVu*100, 2)

delta_altitude_raw = abs(raw_Altitude[nel-1]-raw_Altitude[0])
delta_altitude = abs(Altitude[nel-1]-Altitude[0])
delta_altitude_start_end = min(delta_altitude, delta_altitude_raw)
elevation_pos = np.sum(dAltitude[dAltitude>0])
pu = dAltitude * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph)
elevation_pos_u = np.sum(pu[pu>0])

elev_abs_diff_start_end = round(delta_altitude_start_end, 2)
total_cumul_elev_pos_gain_100km = round(elevation_pos*(100/trip_distance), 2)
urban_cumul_elev_pos_gain_100km = round(elevation_pos_u*(100/urban_distance), 2)

init_idle_duration = round(time_engine_idle_int, 2)
if time_engine_start > time_cold_start_end and time_cold_start_end == 300: 
    time_cold_start_end = time_cold_start_end + time_engine_start
cold_start_avg_speed = round(np.mean(vehspd_ecu[ time_engine_start:time_cold_start_end]), 2)
cold_start_max_speed = round(np.max(vehspd_ecu[ time_engine_start:time_cold_start_end]), 2)
j = 0
nstop_during_coldstart = 0
for k in range (int(time_cold_start_end)):
	if vehspd_ecu[k] <= MAW_vehspd_0_kph: j = j + 1# for EMROAD zero speed definition
nstop_during_coldstart = j

cold_stop_time = round(nstop_during_coldstart)

nVu_ACC = len(Vu_ACC[Vu_ACC > 0.1])
nVr_ACC = len(Vr_ACC[Vr_ACC > 0.1])
nVm_ACC = len(Vm_ACC[Vm_ACC > 0.1])

nMAWt = len(MAW_vehspd)
    
CO_RDE_mgpkm = np.sum(Inst_Mass_CO) *1000/(np.sum(dVMT)/1000)
NOx_RDE_mgpkm = np.sum(Inst_Mass_NOx) *1000/(np.sum(dVMT)/1000)
    
CO_RDEu_mgpkm = np.sum(Inst_Mass_CO[vehspd_ecu <= vehspd_urban_kph]) *1000/(np.sum(dVMT[vehspd_ecu <= vehspd_urban_kph]/1000))
NOx_RDEu_mgpkm = np.sum(Inst_Mass_NOx[vehspd_ecu <= vehspd_urban_kph]) *1000/(np.sum(dVMT[vehspd_ecu <= vehspd_urban_kph]/1000))

total_dist_miles = 0.000621371 * VMT[nel-1]
total_fuel_gals = np.sum(fuelflow_gals)
if total_fuel_gals <= 0.1: 
    FE_mpg = 'N/A'
    CO2_gpm = 'N/A'
else:
    FE_mpg = total_dist_miles / total_fuel_gals
    CO2_mgpm = 8887/FE_mpg
    
CO_mgpm_test = 1000 * np.sum(Inst_Mass_CO) / total_dist_miles
CO2_gpm_test = np.sum(Inst_Mass_GHG) / total_dist_miles
engNOx_mgpm_test = 1000 * np.sum(Inst_Mass_NOx) / total_dist_miles

label_CO_test = str(round(CO_mgpm_test/1000,1)) + " g/mile"
label_CO2_gpm_test = str(round(CO2_gpm_test,1)) + " g/mile"
label_NOx_test = str(round(engNOx_mgpm_test,1)) + " mg/mile"

if FE_mpg == 'N/A':
    print (vehicle_title, ": ", "%0.1f" %GHG_WLTC_threshold, "gCO2 MAW, FE = N/A", label_CO2_gpm_test, label_NOx_test)
else:
    print (vehicle_title, ": ", "%0.1f" %GHG_WLTC_threshold, "gCO2 MAW, FE = ", "%0.1f" %FE_mpg, "MPG, GHG = ", label_CO2_gpm_test, ", NOx = ", label_NOx_test)

rpt = RDE_report (rpt, filename_combined, filename_combined_obd, infile_folder, now_hms, suffix, trip_distance, trip_duration, cold_start_duration, urban_distance, rural_distance, motorway_distance,
            urban_distance_share, rural_distance_share, motorway_distance_share, urban_avg_speed, rural_avg_speed, motorway_avg_speed, total_trip_avg_speed,
            motorway_speed_above_145kph, motorway_speed_above_100kph, urban_stop_time_pct, elev_abs_diff_start_end, total_cumul_elev_pos_gain_100km, urban_cumul_elev_pos_gain_100km,
            init_idle_duration, cold_start_avg_speed, cold_start_max_speed, cold_stop_time, vehspd_avg, RPA_avg, VA_avg_95pct, nVu_ACC, nVr_ACC, nVm_ACC,
            nMAWt, nMAWu, nMAWr, nMAWm, CO_RDE_mgpkm, NOx_RDE_mgpkm, CO_RDEu_mgpkm, NOx_RDEu_mgpkm, FE_mpg, CO2_gpm_test, engNOx_mgpm_test, CO_mgpm_test)

outFilename_inp = os.path.join(infile_folder, os.path.basename(filename_combined))
outFilename = os.path.splitext(outFilename_inp)[0] + '_fig' + now_hms + '.pdf'

Duration = np.zeros(nel)
shp = (nel, 24)
data = np.zeros(shp)
    
data[:,0]=etime
data[:,1]=vehspd_ecu
data[:,2]=VMT
data[:,3]=Altitude
if (Altitude[nel-1] >= 0) == False: Altitude[nel-1] = Altitude[nel-2]
data[:,4]=engine_RPM
data[:,5]=engine_load
data[:,6]=ECT
data[:,7]=Tamb
data[:,8]=Inst_Mass_GHG
data[:,9]=Inst_Mass_NOx

data[:,10]=Inst_Mass_CO
data[:,11]=fuelflow_gals
data[:,12]= CatalystTemp
for k in range (nel):
    if GHG_end[k] == GHG_start[k] and GHG_exclude_pts[k] == 0:
        Duration[k] = 0
    else:
        Duration[k] = GHG_end[k] - GHG_start[k] - GHG_exclude_pts[k] + 1
data[:,13] = Duration
data[:,14]=GHG_start
data[:,15]=GHG_end
data[:,16]=GHG_exclude_pts

data[:,17]=tMAW_GHG_gpkm
data[:,18]=tMAW_CO_gpkm      
data[:,19]=tMAW_NOx_gpkm      
data[:,20]=tMAW_vehspd      
data[:,21]=veh_accel     
data[:,22]=RPA     
data[:,23]=Lambda     

df_columns = ['TIME', 'Vehicle Speed', 'Cumul_Distance', 'GPS Altitude', 'Engine RPM', 'Engine Load', 'Coolant Temp.', 'Ambient Temperature', 'Instantaneous Mass CO2', 
                  'Instantaneous Mass NOx', 'Instantaneous Mass CO', 'Fuel Flow', 'CATEMP11',
                  'Duration', 'Start', 'End', 'Excluded points', 'CO2', 'CO', 'NOx', 'Mean Vehicle Speed',  'veh_accel',  'RPA',  'lambda']

df_RDE = pd.DataFrame(data, index=None, columns=df_columns)        
df_RDE['CATEMP11']=df_RDE['CATEMP11'].replace(0, '', regex=True)
df_RDE['Duration']=df_RDE['Duration'].replace(0, '', regex=True)
df_RDE['Start']=df_RDE['Start'].replace(0, '', regex=True)
df_RDE['End']=df_RDE['End'].replace(0, '', regex=True)
df_RDE['Excluded points']=df_RDE['Excluded points'].replace(0, '', regex=True)

df_RDE['CO2']=df_RDE['CO2'].replace(0, '', regex=True)
df_RDE['CO']=df_RDE['CO'].replace(0, '', regex=True)
df_RDE['NOx']=df_RDE['NOx'].replace(0, '', regex=True)

df_RDE['Mean Vehicle Speed']=df_RDE['Mean Vehicle Speed'].replace(0, '', regex=True)
df_RDE['veh_accel']=df_RDE['veh_accel'].replace(0, '', regex=True)
df_RDE['RPA']=df_RDE['RPA'].replace(0, '', regex=True)
df_RDE['lambda']=df_RDE['lambda'].replace(0, '', regex=True)

curve_CO2 = np.zeros(len(settings_avg_speed))
for j in range (len(settings_avg_speed)):
    tmp = float(settings_avg_speed[j])
    if tmp < WLTC_vehspd_kph [0]: # extra-polation
       curve_CO2[j] =  WLTC_CO2_gpkm[1] + (tmp - WLTC_vehspd_kph[1])/(WLTC_vehspd_kph[0] - WLTC_vehspd_kph[1]) * (WLTC_CO2_gpkm[0] - WLTC_CO2_gpkm[1])
    elif tmp >= WLTC_vehspd_kph [0] and  tmp <= WLTC_vehspd_kph[1]:
       a1 =  (WLTC_CO2_gpkm[1] - WLTC_CO2_gpkm[0])/(WLTC_vehspd_kph[1] - WLTC_vehspd_kph[0])
       curve_CO2[j] = a1 * tmp + (WLTC_CO2_gpkm[0] - a1 * WLTC_vehspd_kph[0])
    elif tmp >= WLTC_vehspd_kph [1] and  tmp <= WLTC_vehspd_kph[2]:
       a1 =  (WLTC_CO2_gpkm[2] - WLTC_CO2_gpkm[1])/(WLTC_vehspd_kph[2] - WLTC_vehspd_kph[1])
       curve_CO2[j] = a1 * tmp + (WLTC_CO2_gpkm[1] - a1 * WLTC_vehspd_kph[1])
    elif tmp >= WLTC_vehspd_kph [2] and  tmp <= WLTC_vehspd_kph[3]:
       a1 =  (WLTC_CO2_gpkm[3] - WLTC_CO2_gpkm[2])/(WLTC_vehspd_kph[3] - WLTC_vehspd_kph[2])
       curve_CO2[j] = a1 * tmp + (WLTC_CO2_gpkm[2] - a1 * WLTC_vehspd_kph[2])
    elif tmp > WLTC_vehspd_kph [3]:
       curve_CO2[j] =  WLTC_CO2_gpkm[2] + (tmp - WLTC_vehspd_kph[2])/(WLTC_vehspd_kph[3] - WLTC_vehspd_kph[2]) * (WLTC_CO2_gpkm[3] - WLTC_CO2_gpkm[2])

settings_TOL1_minus = (1 + settings_TOL1_minus_pct/100) * settings_CO2
settings_TOL1_plus = (1 + settings_TOL1_plus_pct/100) * settings_CO2
#settings_TOL2_minus = (1 + settings_TOL2_minus_pct/100) * settings_CO2
#settings_TOL2_plus = (1 + settings_TOL2_plus_pct/100) * settings_CO2

settings_TOL1_minus = (1 + settings_TOL1_minus_pct/100) * curve_CO2
settings_TOL1_plus = (1 + settings_TOL1_plus_pct/100) * curve_CO2

VA_vehspd_mean = [0,	10,	20,	30,	40,	50,	60,	75,	80,	90,	100, 110, 120, 130, 140, 150]

VA_95pct_limit = np.zeros(len(VA_vehspd_mean))
for j in range (len(VA_vehspd_mean)):
    tmp = float(VA_vehspd_mean[j])
    if tmp < 74.6:
        VA_95pct_limit[j] =  0.136*tmp+14.44
    else:
        VA_95pct_limit[j] =  0.0742*tmp+18.966

RPA_limit = np.zeros(len(VA_vehspd_mean))
for j in range (len(VA_vehspd_mean)):
    tmp = float(VA_vehspd_mean[j])
    if tmp < 94.05:
        RPA_limit[j] =  -0.0016 *tmp+0.1755
    else:
        RPA_limit[j] =  0.025

pp1 = PdfPages(outFilename)

etime0 = etime

if etime_reset == True and OBD != True:
    [etime, tvehspd_zero] = etime_resetting(etime, vehspd_ecu)
#    if OBD == True: [otime, tvehspd_zero] = etime_resetting(otime, ovehspd)
MAW_GHG_gpmi = 1.60934* MAW_GHG_gpkm

sMAWu_GHG_gpkm = MAW_GHG_gpkm[MAW_vehspd<45]
sMAWu_GHG_gpkm = np.mean(sMAWu_GHG_gpkm[sMAWu_GHG_gpkm > 0])
sMAWr_GHG_gpkm = np.mean(MAWr_GHG_gpkm)
sMAWm_GHG_gpkm = np.mean(MAW_GHG_gpkm[MAW_vehspd >= 80])
sMAW_GHG_gpkm = np.mean(MAW_GHG_gpkm)
sMAWu_NOx_gpkm = np.mean(MAW_NOx_gpkm[MAW_vehspd<45])
sMAWr_NOx_gpkm = np.mean(MAWr_NOx_gpkm)
sMAWm_NOx_gpkm = np.mean(MAW_NOx_gpkm[MAW_vehspd >= 80])
sMAW_NOx_gpkm = np.mean(MAW_NOx_gpkm)
sMAW_Vu = np.mean(MAWu_vehspd)
sMAW_Vr = np.mean(MAWr_vehspd)
sMAW_Vm = np.mean(MAW_vehspd[MAW_vehspd>=80])
sMAW_V = np.mean(MAW_vehspd)

MAW_vehspd_urm = [sMAW_Vu, sMAW_Vr, sMAW_Vm]
MAW_GHG_urm = [sMAWu_GHG_gpkm, sMAWr_GHG_gpkm, sMAWm_GHG_gpkm]

df_vehspd = pd.DataFrame(MAW_vehspd_urm, columns=['MAW_vehspd_urm'])
#df_vehspd['MAW_vehspd_urm'] = MAW_vehspd_urm
df_vehspd['MAW_GHG_urm'] = MAW_GHG_urm
df_vehspd['VehSpd_urm'] = vehspd_avg
df_vehspd['VA_95p_urm'] = VA_urm_95pct
df_vehspd['RPA_urm'] = RPA_avg
#df_vehspd.to_excel( filename_combined + '-MawRPA.xlsx', index=False, encoding='utf-8-sig')

fname = os.path.splitext(outFilename_inp)[0] + '_out' + now_hms + '.xlsx'

writer = pd.ExcelWriter(fname, engine = 'xlsxwriter')
df_RDE.to_excel(writer, index=None, header=True, sheet_name = 'Sheet1')
df_vehspd.to_excel(writer, index=None, header=True, sheet_name = 'VehSpd')
# writer.save()
writer.close()

if RDE_UN_GTR == False:
    MAW_GHG_gpkm1 = MILES_to_KM * MAW_GHG_gpkm
    WLTC_CO2_gpkm1 = MILES_to_KM * WLTC_CO2_gpkm
    sMAWu_GHG_gpkm1 = MILES_to_KM * sMAWu_GHG_gpkm
    sMAWr_GHG_gpkm1 = MILES_to_KM * sMAWr_GHG_gpkm
    sMAWm_GHG_gpkm1 = MILES_to_KM * sMAWm_GHG_gpkm

    MAW_vehspd1 = MAW_vehspd/MILES_to_KM
    WLTC_vehspd_kph1 = WLTC_vehspd_kph/MILES_to_KM
    sMAW_Vu1 = sMAW_Vu/MILES_to_KM
    sMAW_Vr1 = sMAW_Vr/MILES_to_KM
    sMAW_Vm1 = sMAW_Vm/MILES_to_KM

sf=1.0
fig = plt.figure(facecolor=(1, 1, 1))
ax = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)

label_MAW_CO2 = 'MAW w/ ' + str(settings_MAw_CO2_pct) + '% WLTC CO2'
# if RDE_UN_GTR == True:
if RDE_UN_GTR == False:
    plt.scatter(MAW_vehspd1, MAW_GHG_gpkm1, s=100, linewidth=1, marker="o", facecolors='none', edgecolors='gray',
                label='MAW CO2')
else:
    plt.scatter(MAW_vehspd, MAW_GHG_gpkm, s=100, linewidth=1, marker="o", facecolors='none', edgecolors='gray', label='MAW CO2')
# if RDE_UN_GTR == True:
#     plt.scatter(WLTC_vehspd_kph, WLTC_CO2_gpkm, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='c', label='WLTC CO2')
if RDE_UN_GTR == False:
    plt.scatter(sMAW_Vu1, sMAWu_GHG_gpkm1, s=200, linewidth=3, marker="s", facecolors='none', edgecolors='b', label = 'Urban Mean CO2')
    plt.scatter(sMAW_Vr1, sMAWr_GHG_gpkm1, s=250, linewidth=3, marker="<", facecolors='none', edgecolors='r', label = 'Rural Mean CO2')
    plt.scatter(sMAW_Vm1, sMAWm_GHG_gpkm1, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='g', label = 'Motorway Mean CO2')
else:
    plt.scatter(WLTC_vehspd_kph, WLTC_CO2_gpkm, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='c', label='WLTC CO2')
    plt.scatter(sMAW_Vu, sMAWu_GHG_gpkm, s=200, linewidth=3, marker="s", facecolors='none', edgecolors='b', label = 'Urban Mean CO2')
    plt.scatter(sMAW_Vr, sMAWr_GHG_gpkm, s=250, linewidth=3, marker="<", facecolors='none', edgecolors='r', label = 'Rural Mean CO2')
    plt.scatter(sMAW_Vm, sMAWm_GHG_gpkm, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='g', label = 'Motorway Mean CO2')

if RDE_UN_GTR == False:
    plt.plot([45/MILES_to_KM, 45/MILES_to_KM],[int(MILES_to_KM*100),  int(MILES_to_KM*max(MAW_GHG_gpkm))], 'k--', linewidth=2, label = 'Urban | Rural')
    plt.plot([80/MILES_to_KM, 80/MILES_to_KM], [int(MILES_to_KM*100), int(MILES_to_KM*max(MAW_GHG_gpkm))], 'r--', linewidth=2, label = 'Rural | Motorway')
else:
    plt.plot([45, 45], [100, int(max(MAW_GHG_gpkm))], 'k--', linewidth=2, label = 'Urban | Rural')
    plt.plot([80, 80], [100, int(max(MAW_GHG_gpkm))], 'r--', linewidth=2, label = 'Rural | Motorway')

if RDE_UN_GTR == True:
    plt.plot(settings_avg_speed, curve_CO2, 'k-', linewidth=3, label = 'Threshold Curve')
    plt.plot(settings_avg_speed, settings_TOL1_minus, 'y-.', linewidth=3, label = 'Low-Threshold')
    plt.plot(settings_avg_speed, settings_TOL1_plus, 'm-.', linewidth=3, label = 'High Threshold')

if RDE_UN_GTR == False:
    if max(MAW_vehspd1) < 70:
        plt.xlim(0, 75)
    elif max(MAW_vehspd1) < 80:
        plt.xlim(0, 85)
    else:
        plt.xlim(0, int(140/MILES_to_KM))
else:
    plt.xlim(0, 140)

if RDE_UN_GTR == False:
    plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [mile/h]', 'CO2 Emission [g/mile]')
else:
    plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [km/h]', 'CO2 Emission [g/km]')

fig = plt.figure(facecolor=(1, 1, 1))
ax = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
plt.scatter(vehspd_avg, VA_avg_95pct, s=150, linewidth=4, marker="s", facecolors='none', edgecolors='r', label='VA_urm')
plt.scatter(vehspd_avg, VA1_avg_95pct, s=150, linewidth=4, marker="d", facecolors='none', edgecolors='g', label='VA_urm /w PKE')
plt.plot(VA_vehspd_mean, VA_95pct_limit, 'y-.', linewidth=3, label ='Limited Curve')
plt.xlim(0.0, 150)
plt.ylim(0.0, 35)
plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [km/h]', '95 pecentile Speed x Positive Acceleration [m2/s3]')

fig = plt.figure(facecolor=(1, 1, 1))
ax = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
plt.plot(VA_vehspd_mean, RPA_limit, 'y-.', linewidth=3, label ='Limited Curve')
plt.scatter(vehspd_avg, RPA_avg, s=150, linewidth=4, marker="s", facecolors='none', edgecolors='r', label='RPA_urm')
plt.scatter(vehspd_avg, RPA1_avg, s=150, linewidth=4, marker="d", facecolors='none', edgecolors='g', label='RPA_urm /w PKE')
plt.xlim(0.0, 150)
plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [km/h]', 'RPA [m/s2]')

print('Mean Vehicle Speed [km/h] = ', np.mean(vehspd_ecu))
print('Vehicle Accel [m/s2] = ', np.sqrt(np.mean(veh_accel**2)))
print('Positive Accel [m/s2] = ', np.sqrt(np.mean(veh_accel[veh_accel>0.1]**2)))
print('Negative Accel [m/s2] = ', np.sqrt(np.mean(veh_accel[veh_accel<0]**2)))
print('Mean Urban Vehicle Speed > 1 km/h [km/h] = ', np.mean(Vu[Vu >= vehspd_0_kph]))
print('Mean Urban Vehicle Speed [km/h] = ', np.mean(Vu))
print('Mean Rural Vehicle Speed [km/h] = ', np.mean(Vr))
print('Mean Motorway Vehicle Speed [km/h] = ', np.mean(Vm))
print('RMS Road Grade [%] = ', round(np.sqrt(np.mean(rgrade**2)), 1))
print('RMS Road Grade Filtered [%] = ', round(np.sqrt(np.mean(rgrade_savgol**2)),1))
print('RMS Catalyst Temperature [C] = ', np.round(np.sqrt(np.mean(CatalystTemp**2)),1))
print('*********************')
CO2 = Inst_Mass_GHG * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
CO2u = Inst_Mass_GHG * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NOx = Inst_Mass_NOx * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
NOxu = Inst_Mass_NOx * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NOxr = Inst_Mass_NOx * np.logical_and(vehspd_ecu > vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph)
NOxm = Inst_Mass_NOx * np.logical_and(vehspd_ecu > vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph)
COu = Inst_Mass_CO * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph)
urban_distance_km = np.sum(dVMT * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph))/1000
rural_distance_km = np.sum(dVMT * np.logical_and(vehspd_ecu>vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph))/1000
highway_distance_km = np.sum(dVMT[np.logical_and(vehspd_ecu>vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph)])/1000
distance_km = np.sum(dVMT * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= 120))/1000
CO2_gpkm = np.sum(CO2)/distance_km
CO2u_gpkm = np.sum(CO2u)/urban_distance_km
NOx_gpmi = 1000 * np.sum(NOx)/distance_km * 1.60934
NOxu_gpmi = 1000 * np.sum(NOxu)/urban_distance_km * 1.60934

MAW_GHGu = np.mean(tMAW_GHG_gpkm [np.logical_and(vehspd_ecu>=1, vehspd_ecu <= vehspd_urban_kph)])
print('MAW CO2 [g/km] = ', MAW_GHGu)
print('CO2 RDE [g/km] = ', CO2_gpkm)
print('CO2 RDE Urban [g/km] = ', CO2u_gpkm)
print('CO2 WLTC [g/km] = ', WLTC_GHG_gpkm)
print('CO2 WLTC Urban [g/km] = ', WLTC_GHGu_gpkm)
print('NOx RDE [mg/mi] = ', NOx_gpmi)
print('NOx RDE Urban [mg/mi] = ', NOxu_gpmi)

RF1 = 1.2
RF2 = 1.25
RF1 = 1.3
RF2 = 1.5
a1 = (RF2 - 1) / (RF2*(RF1 - RF2))
b1 = 1 - a1 * RF1

CO2f = CO2_gpkm / WLTC_GHG_gpkm
CO2uf = CO2u_gpkm/WLTC_GHGu_gpkm
MAW_GHGu = np.mean(tMAW_GHG_gpkm[np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph)])
mCO2uf = MAW_GHGu/WLTC_GHGu_gpkm
if CO2f <= RF1:
    RFt = 1.0
elif CO2f > RF1 and CO2f <= RF2:
    RFt = a1 * CO2f + b1
else:
    RFt = 1/CO2f

if CO2uf <= RF1:
    RFu = 1.0
elif CO2uf > RF1 and CO2uf <= RF2:
    RFu = a1 * CO2uf + b1
else:
    RFu = 1/CO2uf

print('CO2 ratiosRDE/WLTC = ', np.round(CO2f, 2), ' RDE/WLTC Urban = ', np.round(CO2uf, 2), ' MAW/WLTC Urban = ', np.round(mCO2uf, 2))    
print('a1 = ', np.round(a1, 2), ' b1 = ', np.round(b1, 3), ' RF urban = ', np.round(RFu, 2), ' RFt = ', RFt)
   
NOx_x_ranges = [min(MAW_vehspd), max(MAW_vehspd)]
FTP_vehspd_ranges = [min(MAW_vehspd), 91.25] # mean = 21.2 mph (34.12kph, max = 56.7 mph (91.25 kph)
vehspd_FTP3bags_mph = 1/1.60934*vehspd_FTP3bags_kph
vehspd_US06_mph = 1/1.60934*vehspd_US06_kph
vehspd_HWFET_mph = 1/1.60934*vehspd_HWFET_kph

MAW_GHG_gpmi = 1.60934 * MAW_GHG_gpkm
MAW_vehspd_mph = 1/1.60934 * MAW_vehspd
MAW_CO2_ranges = [HwFET_CO2_gpmi, np.max(MAW_GHG_gpmi)*2/3]
MAWu_CO2_ranges = [HwFET_CO2_gpmi, np.max(MAW_GHG_gpmi)*2/3]
MAW_NOx_mgpmi = 1.60934*1000*MAW_NOx_gpkm
NOx_RDEu_mgpmi = 1.60934*NOx_RDEu_mgpkm
NOx_RDE_mgpmi = 1.60934*NOx_RDE_mgpkm
FTP_NOx_mgpmi = 1000*FTP_NOx_comp_gpmi
FTP3_NOx_mgpmi = 1000* FTP3_NOx_gpmi
US06_NOx_mgpmi = 1000* US06_NOx_gpmi
HwFET_NOx_mgpmi = 1000* HwFET_NOx_gpmi
MAW_CO_gpmi = 1.60934*MAW_CO_gpkm
CO_RDEu_gpmi = 1.60934*CO_RDEu_mgpkm/1000
CO_RDE_gpmi = 1.60934*CO_RDE_mgpkm/1000

label_NOx_title = vehicle_title + ', NOx Standards: ' + str(int(NOx_STD_mgpmi)) + ' mg/mile'

label_RDEu_NOx = 'RDE Urban: ' + str(round(NOx_RDEu_mgpmi, 1)) + ' mg/mile'
label_RDE_NOx = 'RDE NOx: ' + str(round(NOx_RDE_mgpmi, 1)) + ' mg/mile'
label_NOx_standards = 'NOx Standards: ' + str(int(NOx_STD_mgpmi)) + ' mg/mile'

label_RDEu_CO = 'RDE Urban: ' + str(round(CO_RDEu_gpmi, 1)) + ' g/mile'
label_RDE_CO = 'RDE CO: ' + str(round(CO_RDE_gpmi, 1)) + ' g/mile'
label_CO_standards = 'CO Standards: ' + str(round(CO_STD_gpmi, 1)) + ' g/mile'
vehspd_WLTC_mph = WLTC_vehspd_kph/1.6034
WLTC_CO2_gpmi = WLTC_CO2_gpkm*1.6034
WLTC_NOx_mgpmi = 1000*WLTC_NOx_gpmi

NOx_RDE_mgpmi = 1000 * np.sum(Inst_Mass_NOx)/(VMT[nel-1]/1000)*1.6034
GHG_RDE_gpmi = np.sum(Inst_Mass_GHG)/(VMT[nel-1]/1000)*1.6034
CO_RDE_gpmi = np.sum(Inst_Mass_CO)/(VMT[nel-1]/1000)*1.6034
CO_RDEu_gpmi = 1000*np.sum(Inst_Mass_NOx[vehspd_ecu <= vehspd_urban_kph])/(np.sum(dVMT[vehspd_ecu <= vehspd_urban_kph])/1000)*1.6034
NOx_RDEu_gpmi = 1000*np.sum(Inst_Mass_NOx[vehspd_ecu <= vehspd_urban_kph])/(np.sum(dVMT[vehspd_ecu <= vehspd_urban_kph])/1000)*1.6034
CO_RDEu_gpmi = np.sum(Inst_Mass_CO[vehspd_ecu <= vehspd_urban_kph])/(np.sum(dVMT[vehspd_ecu <= vehspd_urban_kph])/1000)*1.6034

mean_vehspd_RDE = np.mean(vehspd_ecu)
mean_vehspd_RDEu = np.mean(vehspd_ecu[vehspd_ecu <= vehspd_urban_kph])

# 2016 Malibu Composite Emissions
# Composite_vehspd_mph = np.array([34.12, 53.5, 77.9])/1.60934
# Composite_CO2_gpmi = [264.5, 254.2, 301.8]
# Composite_NOx_mgpmi= [10, 6, 13]
# Composite_CO_gpmi = [0.352, 0.393, 1.269]

if (LAB_WLTC_NOx_Fit == True) or (LAB_WLTC_NOx_Fit > 0):
    Composite_vehspd_mph = np.array([vehspd_FTP3bags_comp_kph, WLTC_vehspd_comp_kph, vehspd_US06_comp_kph])/1.60934
    Composite_CO2_gpmi = np.array([FTP3_CO2_comp_gpmi, WLTC_CO2_comp_gpmi, US06_CO2_comp_gpmi])
    Composite_NOx_mgpmi = 1000 * np.array([FTP_NOx_comp_gpmi, WLTC_NOx_comp_gpmi, US06_NOx_comp_gpmi])
    Composite_CO_gpmi =  np.array([FTP_CO_comp_gpmi, WLTC_CO_comp_gpmi, US06_CO_comp_gpmi])

    x_vehspd = np.array([vehspd_FTP3bags_mph[0], vehspd_FTP3bags_mph[1], vehspd_FTP3bags_mph[2], vehspd_US06_mph[0], vehspd_US06_mph[1], vehspd_HWFET_mph[0], \
                          vehspd_WLTC_mph[0], vehspd_WLTC_mph[1], vehspd_WLTC_mph[2], vehspd_WLTC_mph[3], Composite_vehspd_mph[0], Composite_vehspd_mph[1], Composite_vehspd_mph[2]])
    x_CO2 = np.array([FTP3_CO2_gpmi[0], FTP3_CO2_gpmi[1], FTP3_CO2_gpmi[2], US06_CO2_gpmi[0], US06_CO2_gpmi[1], HwFET_CO2_gpmi[0], \
                      WLTC_CO2_gpmi[0], WLTC_CO2_gpmi[1], WLTC_CO2_gpmi[2], WLTC_CO2_gpmi[3], Composite_CO2_gpmi[0], Composite_CO2_gpmi[1], Composite_CO2_gpmi[2]])
    y_NOx = np.array([FTP3_NOx_mgpmi[0], FTP3_NOx_mgpmi[1], FTP3_NOx_mgpmi[2], US06_NOx_mgpmi[0], US06_NOx_mgpmi[1], HwFET_NOx_mgpmi[0], \
                      WLTC_NOx_mgpmi[0], WLTC_NOx_mgpmi[1], WLTC_NOx_mgpmi[2], WLTC_NOx_mgpmi[3], Composite_NOx_mgpmi[0], Composite_NOx_mgpmi[1], Composite_NOx_mgpmi[2]])
    y_CO = np.array([FTP3_CO_gpmi[0], FTP3_CO_gpmi[1], FTP3_CO_gpmi[2], US06_CO_gpmi[0], US06_CO_gpmi[1], HwFET_CO_gpmi[0], \
                      WLTC_CO_gpmi[0], WLTC_CO_gpmi[1], WLTC_CO_gpmi[2], WLTC_CO_gpmi[3], Composite_CO_gpmi[0], Composite_CO_gpmi[1], Composite_CO_gpmi[2]])
else:
    Composite_vehspd_mph = np.array([vehspd_FTP3bags_comp_kph, vehspd_US06_comp_kph])/1.60934
    Composite_CO2_gpmi = np.array([FTP3_CO2_comp_gpmi, US06_CO2_comp_gpmi])
    Composite_NOx_mgpmi = 1000 * np.array([FTP_NOx_comp_gpmi, US06_NOx_comp_gpmi])
    Composite_CO_gpmi =  np.array([FTP_CO_comp_gpmi, US06_CO_comp_gpmi])   
    x_vehspd = np.array([vehspd_FTP3bags_mph[0], vehspd_FTP3bags_mph[1], vehspd_FTP3bags_mph[2], vehspd_US06_mph[0], vehspd_US06_mph[1], vehspd_HWFET_mph[0], \
                          Composite_vehspd_mph[0], Composite_vehspd_mph[1]])
    x_CO2 = np.array([FTP3_CO2_gpmi[0], FTP3_CO2_gpmi[1], FTP3_CO2_gpmi[2], US06_CO2_gpmi[0], US06_CO2_gpmi[1], HwFET_CO2_gpmi[0], \
                      Composite_CO2_gpmi[0], Composite_CO2_gpmi[1]])
    y_NOx = np.array([FTP3_NOx_mgpmi[0], FTP3_NOx_mgpmi[1], FTP3_NOx_mgpmi[2], US06_NOx_mgpmi[0], US06_NOx_mgpmi[1], HwFET_NOx_mgpmi[0], \
                      Composite_NOx_mgpmi[0], Composite_NOx_mgpmi[1]])
    y_CO = np.array([FTP3_CO_gpmi[0], FTP3_CO_gpmi[1], FTP3_CO_gpmi[2], US06_CO_gpmi[0], US06_CO_gpmi[1], HwFET_CO_gpmi[0], \
                      Composite_CO_gpmi[0], Composite_CO_gpmi[1]])

coefs_NOx_GHG = poly.polyfit(MAW_GHG_gpmi, MAW_NOx_mgpmi, 1)
x_new = np.linspace(np.min(MAW_GHG_gpmi), np.max(MAW_GHG_gpmi), 100)
ffit = poly.polyval(x_new, coefs_NOx_GHG)

print('MAW NOx/GHG Slope: ', coefs_NOx_GHG[1], ' intercept: ', coefs_NOx_GHG[0])
'''
ffit = coefs[0] + coefs[1] * x_new + coefs[2] * x_new**2
plt.plot(x_new, ffit, 'b-', x_new, ffit1, 'r-.')
'''

dfx = pd.DataFrame(x_CO2, columns = ['MAW_CO2'])
dfx['MAW_NOx'] = y_NOx
dfx = dfx.sort_values(['MAW_CO2'], ascending=[True])
dfx = dfx.reset_index(drop = True)
coefs_ftp = poly.polyfit(dfx['MAW_CO2'], dfx['MAW_NOx'], 1)
x_ftp = np.linspace(np.min(dfx['MAW_CO2']), np.max(dfx['MAW_CO2']), 100)
ffit_ftp = poly.polyval(x_ftp, coefs_ftp)
print('FTP NOx/GHG Slope: ', coefs_ftp[1], ' intercept: ', coefs_ftp[0])

fig = plt.figure(facecolor=(1, 1, 1))
ax1 = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
mean_GHG_gpmi = 1.6034*np.sum(Inst_Mass_GHG)*1000/np.sum(dVMT)
Inst_Mass_GHGu = Inst_Mass_GHG[vehspd_ecu <= vehspd_urban_kph]
mean_GHGu_gpmi = 1.6034*np.sum(Inst_Mass_GHGu)*1000/np.sum(dVMT[vehspd_ecu <= vehspd_urban_kph])

plt.scatter(MAW_GHG_gpmi, MAW_NOx_mgpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW NOx')
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW NOx fit')
if RDE_UN_GTR == True:
    plt.plot(x_ftp, ffit_ftp,  'm--', linewidth=3, label = 'FTP NOx fit')
plt.scatter(mean_GHG_gpmi, NOx_RDE_mgpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_NOx)
plt.scatter(mean_GHGu_gpmi, NOx_RDEu_mgpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_NOx)
if RDE_UN_GTR == True:
    plt.scatter(FTP3_CO2_gpmi, FTP3_NOx_mgpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 NOx')
    plt.scatter(US06_CO2_gpmi, US06_NOx_mgpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 NOx')
    plt.scatter(HwFET_CO2_gpmi, HwFET_NOx_mgpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET NOx')
    plt.scatter(Composite_CO2_gpmi, Composite_NOx_mgpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite NOx')
    if (LAB_WLTC_NOx_Fit > 0): plt.scatter(WLTC_CO2_gpmi, WLTC_NOx_mgpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC NOx')

if np.min(MAW_NOx_mgpmi) < 0: 
    plt.ylim(np.min(MAW_NOx_mgpmi)-30, np.max(MAW_NOx_mgpmi)+30)
elif LAB_WLTC_NOx_Fit > 0 and np.max(WLTC_NOx_mgpmi) > np.max(MAW_NOx_mgpmi):
    plt.ylim(0, np.max(WLTC_NOx_mgpmi)+1)
elif NOx_RDEu_mgpmi > np.max(MAW_NOx_mgpmi):
    plt.ylim(0, NOx_RDEu_mgpmi+5)
elif np.min(MAW_NOx_mgpmi) < 10 :
    plt.ylim(0, np.max(MAW_NOx_mgpmi)+5)
elif np.min(MAW_NOx_mgpmi) < dfx['MAW_NOx'][0]:
    plt.ylim(np.min(MAW_NOx_mgpmi)-30, np.max(MAW_NOx_mgpmi)+30)
else:
    plt.ylim(dfx['MAW_NOx'][0]-30, np.max(MAW_NOx_mgpmi)+30)
plt_bgf_format(pp1, fig, plt, ax1, label_NOx_standards, vehicle_title, 'CO2 Emission [g/mile]', 'NOx Emission [mg/mile]')

coefs_NOx_vehspd = poly.polyfit(MAW_vehspd_mph, MAW_NOx_mgpmi, 1)
x_new = np.linspace(np.min(MAW_vehspd_mph), np.max(MAW_vehspd_mph), 100)
ffit = poly.polyval(x_new, coefs_NOx_vehspd)
print('MAW NOx/VehSpd Slope: ', coefs_NOx_vehspd[1], ' intercept: ', coefs_NOx_vehspd[0])

dfv = pd.DataFrame(x_vehspd, columns = ['vehspd_mph'])
dfv['MAW_NOx'] = y_NOx
dfv = dfv.sort_values(['vehspd_mph'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_ftp_vehspd = poly.polyfit(dfv['vehspd_mph'], dfv['MAW_NOx'], 1)
x_ftp_vehspd = np.linspace(np.min(dfv['vehspd_mph']), np.max(dfv['vehspd_mph']), 100)
ffit_ftp_vehspd = poly.polyval(x_ftp_vehspd, coefs_ftp_vehspd)
print('FTP NOx/VehSpd Slope: ', coefs_ftp_vehspd[1], ' intercept: ', coefs_ftp_vehspd[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_vehspd_mph, MAW_NOx_mgpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW NOx')
plt.scatter(mean_vehspd_RDE, NOx_RDE_mgpkm*1.6094, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_NOx)
plt.scatter(mean_vehspd_RDEu, NOx_RDEu_mgpkm*1.6094, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_NOx)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW NOx fit')
if RDE_UN_GTR == True:
    plt.scatter(vehspd_FTP3bags_mph, FTP3_NOx_mgpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 NOx')
    plt.scatter(vehspd_US06_mph, US06_NOx_mgpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 NOx')
    plt.scatter(vehspd_HWFET_mph, HwFET_NOx_mgpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET NOx')
    plt.scatter(Composite_vehspd_mph, Composite_NOx_mgpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite NOx')
    if (LAB_WLTC_NOx_Fit > 0): plt.scatter(vehspd_WLTC_mph, WLTC_NOx_mgpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC NOx')
    plt.plot(x_ftp_vehspd, ffit_ftp_vehspd,  'm--', linewidth=3, label = 'FTP NOx fit')
plt_bgf_format(pp1, fig, plt, ax1, label_NOx_standards, vehicle_title, 'Vehicle Speed [km/h]', 'NOx Emission [mg/mile]')

coefs_CO_GHG = poly.polyfit(MAW_GHG_gpmi, MAW_CO_gpmi, 1)
x_new = np.linspace(np.min(MAW_GHG_gpmi), np.max(MAW_GHG_gpmi), 100)
ffit = poly.polyval(x_new, coefs_CO_GHG)
print('MAW CO/CO2 Slope: ', coefs_CO_GHG[1], ' intercept: ', coefs_CO_GHG[0])

dfv = pd.DataFrame(x_CO2, columns = ['MAW_CO2'])
dfv['MAW_CO'] = y_CO
dfv = dfv.sort_values(['MAW_CO2'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_CO_ftp_GHG = poly.polyfit(dfv['MAW_CO2'], dfv['MAW_CO'], 1)
x_ftp_GHG = np.linspace(np.min(dfv['MAW_CO2']), np.max(dfv['MAW_CO2']), 100)
ffit_ftp_GHG = poly.polyval(x_ftp_GHG, coefs_CO_ftp_GHG)
print('FTP CO/CO2 Slope: ', coefs_CO_ftp_GHG[1], ' intercept: ', coefs_CO_ftp_GHG[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_GHG_gpmi, MAW_CO_gpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW CO')
plt.scatter(mean_GHG_gpmi, CO_RDE_gpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_CO)
plt.scatter(mean_GHGu_gpmi, CO_RDEu_gpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_CO)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW CO fit')
if RDE_UN_GTR == True:
    plt.scatter(FTP3_CO2_gpmi, FTP3_CO_gpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 CO')
    plt.scatter(US06_CO2_gpmi, US06_CO_gpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 CO')
    plt.scatter(HwFET_CO2_gpmi, HwFET_CO_gpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET CO')
    plt.scatter(Composite_CO2_gpmi, Composite_CO_gpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite CO')
    if (LAB_WLTC_NOx_Fit > 0): plt.scatter(WLTC_CO2_gpmi, WLTC_CO_gpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC CO')
    plt.plot(x_ftp_GHG, ffit_ftp_GHG,  'm--', linewidth=3, label = 'FTP CO fit')
plt_bgf_format(pp1, fig, plt, ax1, label_CO_standards, vehicle_title, 'CO2 Emission [g/mile]', 'CO Emission [g/mile]')

coefs_CO_vehspd = poly.polyfit(MAW_vehspd_mph, MAW_CO_gpmi, 1)
x_new = np.linspace(np.min(MAW_vehspd_mph), np.max(MAW_vehspd_mph), 100)
ffit = poly.polyval(x_new, coefs_CO_vehspd)
print('MAW CO/VehSpd Slope: ', coefs_CO_vehspd[1], ' intercept: ', coefs_CO_vehspd[0])

dfv = pd.DataFrame(x_vehspd, columns = ['vehspd_mph'])
dfv['MAW_CO'] = y_CO
dfv = dfv.sort_values(['vehspd_mph'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_CO_ftp_vehspd = poly.polyfit(dfv['vehspd_mph'], dfv['MAW_CO'], 1)
x_ftp_vehspd = np.linspace(np.min(dfv['vehspd_mph']), np.max(dfv['vehspd_mph']), 100)
ffit_ftp_vehspd = poly.polyval(x_ftp_vehspd, coefs_CO_ftp_vehspd)
print('FTP CO/VehSpd Slope: ', coefs_CO_ftp_vehspd[1], ' intercept: ', coefs_CO_ftp_vehspd[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_vehspd_mph, MAW_CO_gpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW CO')
plt.scatter(mean_vehspd_RDE, CO_RDE_gpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_CO)
plt.scatter(mean_vehspd_RDEu, CO_RDEu_gpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_CO)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW CO fit')
if RDE_UN_GTR == True:
    plt.scatter(vehspd_FTP3bags_mph, FTP3_CO_gpmi, s=150, linewidth=5, marker="d", facecolors='none', edgecolors='g', label='FTP3 CO')
    plt.scatter(vehspd_US06_mph, US06_CO_gpmi, s=150, linewidth=5, marker="h", facecolors='none', edgecolors='b', label='US06 CO')
    plt.scatter(vehspd_HWFET_mph, HwFET_CO_gpmi, s=150, linewidth=5, marker="o", facecolors='none', edgecolors='c', label='HwFET CO')
    plt.scatter(Composite_vehspd_mph, Composite_CO_gpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite CO')
    if (LAB_WLTC_NOx_Fit > 0): plt.scatter(vehspd_WLTC_mph, WLTC_CO_gpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC CO')
    plt.plot(x_ftp_vehspd, ffit_ftp_vehspd,  'm--', linewidth=3, label = 'FTP CO fit')
plt_bgf_format(pp1, fig, plt, ax1, label_CO_standards, vehicle_title, 'Vehicle Speed [km/h]', 'CO Emission [g/mile]')

fuelflow_gps = fuelflow_gals*2834.89 # https://www.aqua-calc.com/calculate/volume-to-weight
print('RMS Fuel Flow [g/s] = ', np.sqrt(np.mean(fuelflow_gps**2)))

coefs_fuelflow = poly.polyfit(fuelflow_gps, Inst_Mass_GHG, 1)
x_new = np.linspace(np.min(fuelflow_gps), np.max(fuelflow_gps), 100)
ffit = poly.polyval(x_new, coefs_fuelflow)
print('MAW CO2/Fuel Flow Slope: ', coefs_fuelflow[1], ' intercept: ', coefs_fuelflow[0])

'''
ffit1 = coefs[0] + coefs[1] * x_new + coefs[2] * x_new**2
plt.plot(x_new, ffit, 'b-', x_new, ffit1, 'r-.')
'''

sf = 1.0
fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
subplot1=fig.add_subplot(111)
#subplot1.scatter(fuelflow_gals*2834.89, Inst_Mass_GHG, 'b-.', label = 'GPS')
plt.plot(x_new, ffit,  'k--', linewidth=5, label = 'RDE CO2 fit')
subplot1.scatter(fuelflow_gps, Inst_Mass_GHG, s=150, linewidth=2, marker="d", facecolors='none', edgecolors='r', label = 'CO2')
plt_bgf_format(pp1, fig, plt, subplot1, '', vehicle_title, 'Fuel Flow [g/s]', 'CO2 Emission [g/s]')

# fig = plt.figure()
fig, axes = plt.subplots(2, 1, sharex=True) #, sharey=True)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
subplot1=axes[0]
subplot1.plot(etime, vehspd_ecu, 'b-', linewidth=1, label = 'Vehicle Speed')
#subplot1.set_xlim(0, len(etime))
subplot1.set_ylabel('Vehicle Speed [km/hr]')
subplot1.grid()

subplot2 = axes[1] # fig.add_subplot(212, sharex=True)
subplot2.plot(etime, rgrade, 'b-', linewidth=1, label ='Raw')
subplot2.plot(etime, rgrade_savgol, 'r--', linewidth=1, label ='w/ SG filter')
#subplot2.set_xlim(0.0, len(etime))
subplot2.set_xlabel('Elapsed Time [s]')
subplot2.set_ylabel('Road Grade [%]')
subplot2.grid()

pp1.savefig(fig, dpi=600)

sum_Corrected_Inst_Mass_NOx = np.sum(Corrected_Inst_Mass_NOx)
sum_Inst_Mass_GHG = np.sum(Inst_Mass_GHG)
sum_Inst_Mass_NOx = np.sum(Inst_Mass_NOx)
label_Inst_Mass_GHG = 'CO2: ' + str(round(sum_Inst_Mass_GHG/1000,1)) + ' Kg'
label_Corrected_Inst_Mass_NOx = 'NOx: ' + str(round(sum_Corrected_Inst_Mass_NOx,1)) + ' g'
label_Inst_Mass_NOx = 'Inst Mass NOx: ' + str(round(sum_Inst_Mass_NOx,1)) + ' g'
label_Inst_Mass_CO = 'CO: ' + str(round(np.sum(Inst_Mass_CO),1)) + ' g'

if OBD == True: 
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Inst_Mass_GHG, 'g-', '', 'CO2 [g/s]', 'Inst CO2', \
                            etime, 1000*Inst_Mass_NOx, 'r', '', 'NOx [mg/s]', 'NOx: ' + label_NOx_test, \
                            etime, 1000*Inst_Mass_CO, 'b', 'Elapsed Time [s]', 'CO [mg/s]', label_CO_test, \
                            otime, oengect, 'r--', 'Coolant Temperature [C]', 'Coolant Temperature')

elif OBD == False and iCOOL_TEMP_cname != '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Inst_Mass_GHG, 'g-', '', 'CO2 [g/s]', 'Inst CO2', \
                            etime, 1000*Inst_Mass_NOx, 'r', '', 'NOx [mg/s]', 'NOx: ' + label_NOx_test, \
                            etime, 1000*Inst_Mass_CO, 'b', 'Elapsed Time [s]', 'CO [mg/s]', label_CO_test, \
                            etime, ECT, 'r--', 'Coolant Temperature [C]', 'Coolant Temperature')

subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'b-', '', 'Vehicle Speed [km/h]', 'PEMS', \
                        etime, Altitude, 'r', '', 'Altitude [m]', rms_dAltitude, \
                        etime, rgrade, 'b-', 'Elapsed Time [s]', 'Grade [%]', 'Raw', \
                        etime, rgrade_savgol, 'r--', '', 'w/ SG filter')

if (iENG_SPEED_cname != '' and iENG_LOAD_cname != '' and iCOOL_TEMP_cname != '') or \
    (OBD == True and max(oengrpm) > 500 and max(oengtrq) > 0 and max(oengect) > 0):
    if OBD == True:
        subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, otime, oengrpm, 'b-', '', 'Engine Speed [RPM]', 'OBD', \
                                otime, oengtrq, 'r', '', 'Engine Torque [Nm]', 'OBD', \
                                etime, ECT, 'r-', 'Elapsed Time [s]', 'Engine Coolant Temperature [C]', 'ECT - PEMS', \
                                otime, oengect, 'b--', 'Coolant Temperature [C]', 'ECT - OBD')
    else:
        subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, engine_RPM, 'b-', '', 'Engine Speed [RPM]', 'PEMS', \
                                etime, engine_load, 'r', '', 'Engine Load [%]', 'PEMS', \
                                etime, ECT, 'r-', 'Elapsed Time [s]', 'Engine Coolant Temperature [C]', 'PEMS', \
                                '', '', 'r--', '', '')
        
if iENG_SPEED_cname != '' and iENG_LOAD_cname == '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, 1000*Inst_Mass_NOx, 'b-', '', 'NOx [mg/s]', 'NOx: ' + label_NOx_test, \
                            etime, 1000*Inst_Mass_CO, 'r', '', 'CO [mg/s]', 'CO: ' + label_CO_test, \
                            etime, engine_RPM, 'g-', 'Elapsed Time [s]', 'Engine RPM', 'PEMS', \
                            '', '', 'r--', '', '')
elif iENG_SPEED_cname != '' and iENG_LOAD_cname != '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, 1000*Inst_Mass_NOx, 'b-', '', 'NOx [mg/s]', 'NOx: ' + label_NOx_test, \
                            etime, 1000*Inst_Mass_CO, 'r', '', 'CO [mg/s]', 'Inst Mass CO: ' + label_CO_test, \
                            etime, engine_RPM, 'b-', 'Elapsed Time [s]', 'Engine RPM', 'Engine Speed', \
                            etime, engine_load, 'r--', 'Engine Load [%]', 'Engine Load')

if OBD == True: 
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Altitude, 'r', '', 'Altitude [m]', rms_dAltitude, \
                            etime, dAltitude, 'r', '', 'delta Altitude', 'delta Altitude @ t=0', \
                            etime, vehspd_ecu, 'b-', 'Elapsed Time [s]', 'Vehicle Speed [km/h]', 'PEMS', \
                            otime, ovehspd, 'r--', '', 'OBD')
else:
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'b', '', 'Vehicle Speed [km/h]', 'PEMS', \
                            etime, dAltitude, 'g', '', 'delta Altitude', 'delta Altitude @ t=0', \
                            etime, Altitude, 'b-.', 'Elapsed Time [s]', 'Altitude [m]', 'Filtered', \
                            etime, raw_Altitude, 'r--', '', 'Raw')

if nVu > 0 and nVr > 0 and nVm > 0:
    xVu = np.arange(nVu)
    xVr = np.arange(nVr)
    xVm = np.arange(nVm)
    subplot_shared_x_axes(sf, pp1, 'noshare', '', xVm, Vm, 'r', '', '', 'Motorway', \
                            xVr, Vr, 'g', '', 'Vehicle Speed [kph]', 'Rural', \
                            xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr > 0 and nVm == 0:
    xVu = np.arange(nVu)
    xVr = np.arange(nVr)
    subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', \
                            xVr, Vr, 'g', '', 'Vehicle Speed [kph]', 'Rural', \
                            '', '', 'b-', '', '', '', \
                            xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm > 0:
    xVu = np.arange(nVu)
    xVm = np.arange(nVm)
    subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', \
                            xVm, Vm, 'g', '', 'Vehicle Speed [kph]', 'Motorway', \
                            '', '', 'b-', '', '', '', \
                            xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm == 0:
    xVu = np.arange(nVu)
    subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', \
                            xVu, Vu, 'g', 'Sampling Points', 'Vehicle Speed [kph]', 'Urban', \
                            '', '', 'b-', '', '', '', \
                            '', '', 'b-', '', '', '', \
                            '', '', 'r--', '', '')

subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, engine_RPM, 'g-', '', 'Engine Speed [RPM]', 'Engine RPM', \
                        etime, engine_load, 'b-', '', 'Engine Load [%]', 'Engine Load', \
                        etime, fuelflow_gps, 'r-', 'Elapsed Time [s]', 'Fuel Flow [g/s]', 'Fuel Flow', \
                        '', '', 'r--', '', '')

if nVu > 0 and nVr > 0 and nVm > 0:
    subplot_shared_x_axes(sf, pp1, 'noshare', vehicle_title, xVu, Vu * Vu_ACC, 'g', '', '', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            xVr, Vr * Vr_ACC, 'b', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Rural, ' + str(round(vehspd_ACCr_95pct,1)), \
                            xVm, Vm * Vm_ACC, 'r-', 'Sampling Points', '', 'Motorway, ' + str(round(vehspd_ACCm_95pct,1)), \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm == 0:
    xVu = np.arange(nVu)
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, xVu, Vu * Vu_ACC, 'g', 'Sampling Points', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            '', '', 'b', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Rural, ', \
                            '', '', 'r-', 'Sampling Points', '', 'Motorway, ', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr > 0 and nVm == 0:
    xVu = np.arange(nVu)
    xVr = np.arange(nVr)
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, xVu, Vu * Vu_ACC, 'g', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            '', '', 'b', '', '', '', \
                            xVr, Vr * Vr_ACC, 'b', 'Sampling Points', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Rural, ' + str(round(vehspd_ACCr_95pct,1)), \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm > 0:
    xVu = np.arange(nVu)
    xVm = np.arange(nVm)
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, xVu, Vu * Vu_ACC, 'g', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            '', '', 'b', '', '', '', \
                            xVm, Vm * Vm_ACC, 'r-', 'Sampling Points', '', 'Motorway, ' + str(round(vehspd_ACCm_95pct,1)), \
                            '', '', 'r--', '', '')
    
subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'g', 'Elapsed Time [s]', 'Vehicle Speed [KPH]', '', \
                        etime, veh_ACC, 'b', '', 'Vehicle Acceleration [m/s2]', 'Veh_ACC ', \
                        etime, vehspd_ecu * veh_ACC, 'r-', 'Elapsed Time [s]', 'Vehspd * ACC', 'Vehspd x veh_ACC', \
                        '', '', 'r--', '', '')
subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'g', 'Elapsed Time [s]', 'Vehicle Speed [KPH]', '', \
                        etime, Tamb, 'b', '', 'Ambient Temperature', 'Tamb [C] ', \
                        etime, CatalystTemp, 'r-', 'Elapsed Time [s]', 'Catalyst Temperature', 'CAT. Temp [C]', \
                        '', '', 'r--', '', '')
df_tmp = pd.DataFrame(Inst_Mass_NOx_gbhp_hr[Inst_Mass_NOx_gbhp_hr != float('inf')])
df_Inst_Mass_NOx_gbhp_hr = df_tmp.dropna()

if calc_NOx_gbhp_hr == True:

    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Inst_Mass_NOx_gbhp_hr, 'g', 'Elapsed Time [s]', 'NOx [g/bhp-hr]', 'RMS NOx: ' + str(np.round(1000*np.sqrt(np.mean(df_Inst_Mass_NOx_gbhp_hr**2)), 1)) + ' mg/bhp-hr', \
                            etime, engtrq_NM, 'b', '', 'Engine Torque [NM]', 'RMS Torque: ' + str(np.round(np.sqrt(np.mean(engtrq_NM**2)),1)) + ' NM', \
                            etime, engpwr_HP, 'r-', 'Elapsed Time [s]', 'Engine Power', 'RMS Power: ' + str(np.round(np.sqrt(np.mean(engpwr_HP**2)),1)) + ' HP', \
                            '', '', 'r--', '', '')   
        
if OBD == True:   
    sf = 1.0
    fig = plt.figure()
    fig.set_size_inches(11*sf, 8.5*sf, forward=True)
    subplot1=fig.add_subplot(111)
    subplot1.plot(etime, vehspd_ecu, 'b-.', label = 'GPS')
    if OBD == True: subplot1.plot(otime, ovehspd, 'r--', label = 'OBD')
    subplot1.set_xlim(0, round(np.max(etime)))
#    subplot1.set_xlim(0, 1000)
    subplot1.set_ylim(000, 125)
    subplot1.grid()
    subplot1.set_title(vehicle_title)
    subplot1.set_ylabel('Vehicle Speed [km/h]')
    subplot1.legend(loc='best')      
    pp1.savefig(fig, dpi=600)

pp1.close()
# plt.show()
os.chdir(odir)

elapsed = time.time() - t
print("\n*************\nElapsed Time: {:5.0f} Seconds after plotting and saving to PDF" .format(elapsed))
