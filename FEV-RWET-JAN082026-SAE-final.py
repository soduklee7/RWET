
'''
page 153, EU COMMISSION REGULATION (EU) 2017/1151 of 1 June 2017
3.1. Definition of averaging windows
The instantaneous emissions calculated according to Appendix 4 shall be integrated using a moving averaging window method, based on the reference CO 2 mass. The principle of the calculation is as follows: The mass emissions are not calculated for the complete data set, but for sub-sets of the complete data set, the length of these sub-sets being determined so as to match the CO 2 mass emitted by the vehicle over the reference laboratory cycle. The moving average calculations are conducted with a time increment Δt corresponding to the data sampling frequency. These sub-sets used to average the emissions data are referred to as ‘averaging windows’. The calculation described in the present point may be run from the last point (backwards) or from the first point (forward).
The following data shall not be considered for the calculation of the CO 2 mass, the emissions and the distance of the averaging windows:
— The periodic verification of the instruments and/or after the zero drift verifications
— The cold start emissions, defined according to Appendix 4, point 4.4
— Vehicle ground speed < 1 km/h
— Any section of the test during which the combustion engine is switched off.

pu = dAltitude * np.logical_and(vehspd_ecu>1, vehspd_ecu<=vehspd_urban_kph)
elevation_pos_u = np.sum(pu[pu>0])

NOxu = Inst_Mass_NOx * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph, eng_Combustion_ON > 0)
CO2 = Inst_Mass_GHG * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
CO2u = Inst_Mass_GHG * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NOx = Inst_Mass_NOx * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
NOxu = Inst_Mass_NOx * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NOxr = Inst_Mass_NOx * np.logical_and(vehspd_ecu > vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph)
NOxm = Inst_Mass_NOx * np.logical_and(vehspd_ecu > vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph)
COu = Inst_Mass_CO * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph)
urban_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph))/1000
rural_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph))/1000
highway_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph))/1000
distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= 120))/1000
CO2_gpmi = np.sum(CO2)/distance_km * 1.60934
CO2u_gpkm = np.sum(CO2u)/urban_distance_km
NOx_mgpmi = 1000*np.sum(NOx)/distance_mi * 1.60934
NOxu_mgpmi = 1000 * np.sum(NOxu)/urban_distance_mi * 1.60934
NOxr_mgpmi = 1000 * np.sum(NOxr)/rural_distance_km * 1.60934
NOxm_mgpmi = 1000 * np.sum(NOxm)/highway_distance_km * 1.60934
urban_CO_mgpmi = 1000 * np.sum(COu)/urban_distance_km * 1.60934

np.sum(Inst_Mass_NOx)*1000/ (np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph))/1000)

'''

import os
import os.path
import sys
import time
import pandas as pd
import numpy as np
from scipy.signal import savgol_filter
from scipy import interpolate
import glob
import datetime
#from datetime import time
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import openpyxl
import seaborn as sns 
sns.set_style("white") #for aesthetic purpose only
import numpy.polynomial.polynomial as poly
import tkinter as tk
from tkinter import *
from tkinter import filedialog, dialog
import datetime

# import plot_RDE_MMHC_NOx as plot_RDE_MMHC_NOx

current_time = time.time()
print(f"Current time: {current_time}")

os.chdir("C:/Models/RWET")
SCREEN_OUTPUT = False

eTIME_DATA_PLOT = 3000
eTIME_DATA_PLOT_flag = False
engine_cold_start_ECT = 70.01

SI_units = True
RDE_UN_GTR = False

OBD = False
LAB_WLTC_NOx_Fit = False
MAW_CO2_plot_color = 'y' 
#MAW_CO2_plot_color = 'gray' 
MAW_vehspd_0_kph = 1
MAW_urban_vehspd_kph = 45
MAW_rural_vehspd_kph = 80
US_max_vehspd_mph = 75
US_max_vehspd_kph = 120
US_max_vehspd_mph = 90  # for EMROAD check
US_max_vehspd_kph = 145
vehspd_0_kph = 0
vehspd_1_kph = 1

# UPDATED urban, rural and motorway windows: 0-45 km/h (urban), 45-80 km/h (rural), and 80-145 km/h (motorway) at page 30, RDE-2020-JRC119889_01.pdf
vehspd_urban_kph = 60 # 45 # 60
vehspd_rural_kph = 90 # 80 # 90
vehspd_motorway_kph = 145
EU_peak_vehspd_mph = 90
EU_vehspd_100kph = 100
EU_peak_vehspd_kph = 145
max_eng_stop_time = 180
max_grade_pct = 5.0
T3Bin160_NOx, T3Bin160_NOx_hALT, T3Bin160_CO = 0.16, 0.16, 4.2
T3Bin125_NOx, T3Bin125_NOx_hALT, T3Bin125_CO = 0.125, 0.16, 2.1
T3Bin110_NOx, T3Bin110_NOx_hALT, T3Bin110_CO = 0.110, 0.16, 2.1
T3Bin70_NOx, T3Bin70_NOx_hALT, T3Bin70_CO = 0.110, 0.16, 2.1
T3Bin85_NOx, T3Bin70_NOx_hALT, T3Bin70_CO = 0.085, 0.16, 2.1

T2Bin4_NOx, T2Bin4_NMOG, T2Bin4_CO, T2Bin4_HCHO, T2Bin4_PM = 0.04, 0.07, 2.1, 0.011, 0.01
T2Bin5_NOx, T2Bin5_NMOG, T2Bin5_CO, T2Bin5_HCHO, T2Bin5_PM = 0.07, 0.09, 4.2, 0.018, 0.018
EURO6D_SI_NOx, EURO6D_SI_NMHC, EURO6D_SI_CO, EURO6D_SI_THC, EURO6D_SI_PM, EURO6D_SI_PN= \
    (60/1000)*1.60934, (68/1000)*1.60934, (1000/1000)*1.60934, (100/1000)*1.60934, (4.5/1000)*1.60934, 6.8*10**11*1.60934
EURO6D_CI_NOx, EURO6D_CI_NMHC, EURO6D_CI_CO, EURO6D_CI_THC, EURO6D_CI_PM, EURO6D_CI_PN= \
    (80/1000)*1.60934, (68/1000)*1.60934, (1000/1000)*1.60934, (90/1000)*1.60934, (4.5/1000)*1.60934, 6.8*10**11*1.60934
    
VA_avg_95pct_emload = [9.02, 16.66, 17.23]
RPA_avg_emload = [0.22, 0.12, 0.06]
vehspd_avg_emload = [23.98, 74.78, 102.87]

VA_vehspd_mean_emload = [0,	10,	20,	30,	40,	50,	60,	75,	80,	90,	100, 110, 120, 130, 140, 150]
VA_95pct_emload = [14, 16, 17, 19, 20, 21, 23, 25, 25, 26, 26, 27, 28, 29, 29, 30]

RPA_vehspd_mean_emload = [0, 10, 20, 30, 40, 50, 60, 70, 80, 94, 100, 110, 120, 130, 140, 150]
RPA_mean_emload = [0.1755, 0.1595,	0.1435,	0.1275,	0.1115,	0.0955,	0.0795,	0.0635,	0.0475,	0.0250,	0.0250,	0.0250,	0.0250,	0.0250,	0.0250,	0.0250]

def select_file(prompt):
    root = Tk()
    # root.update()
    root.focus_force()
    root.after(100, root.withdraw)
    # root.withdraw()
    # here = os.path.dirname(os.path.realpath("__file__"))
    # currentdir= os.getcwd()

    myfiletypes = [('Excel files', '.xlsx'), ('CSV files', '.csv')]
    root.filename = filedialog.askopenfilename(initialdir=os.getcwd(), title ='Select ' + prompt, filetypes = myfiletypes)
    print(root.filename)
    
    select_file_name = root.filename
    root.destroy()
    root.quit()
    return select_file_name

def save_folder(prompt):
    root = Tk()
    #root.withdraw()
    root.focus_force()
    root.after(100, root.withdraw)
    #here = os.path.dirname(os.path.realpath("__file__"))
    currentdir= os.getcwd()
    root.directory = filedialog.askdirectory(initialdir = currentdir, title = 'Select ' + prompt)
    save_folder_name = root.directory
    root.destroy()
    root.quit()
    return save_folder_name

class takeInput(object):
    '''
    for taking a text input, for a run ID or user provided text string for use within a script
    '''

    def __init__(self, requestMessage):
        self.root = Tk()
        self.string = ''
        self.frame = Frame(self.root)
        self.frame.pack()
        self.acceptInput(requestMessage)

    def acceptInput(self, requestMessage):
        r = self.frame

        k = Label(r,text=requestMessage)
        k.pack(side='left')
        self.e = Entry(r,text='Name')
        self.e.pack(side='left')
        self.e.focus_set()
        b = Button(r,text='Enter',command=self.gettext)
        b.pack(side='right')

    def gettext(self):
        self.string = self.e.get()
        self.root.destroy()
        self.root.quit()

    def getString(self):
        return self.string

    def waitForInput(self):
        self.root.mainloop()

def getText(requestMessage):
    msgBox = takeInput(requestMessage)
    #loop until the user makes a decision and the window is destroyed
    msgBox.waitForInput()
    return msgBox.getString()

def polyfit(x, y, degree):
    results = {}

    coeffs = np.polyfit(x, y, degree)

     # Polynomial Coefficients
    results['polynomial'] = coeffs.tolist()

    # r-squared
    p = np.poly1d(coeffs)
    # fit values, and mean
    yhat = p(x)                         # or [p(z) for z in x]
    ybar = np.sum(y)/len(y)          # or sum(y)/len(y)
    ssreg = np.sum((yhat-ybar)**2)   # or sum([ (yihat - ybar)**2 for yihat in yhat])
    sstot = np.sum((y - ybar)**2)    # or sum([ (yi - ybar)**2 for yi in y])
    results['determination'] = ssreg / sstot

    return results

def vehmove_stop(veh_move_stop, iENG_SPEED_cname, engineRPM):
    engine_stop = 0
    for j in range (veh_move_stop-1, 1, -1):
        if iENG_SPEED_cname != '' and float(engineRPM[j-1]) >= 500 and float(engineRPM[j]) < 500:
            engine_stop = j
            break       

    return engine_stop

#pylab.ylabel('Example', fontsize=40).
#ax.set_ylabel('Example', fontsize=40) or afterwards with ax.yaxis.label.set_size(40).

def plt_bgf_format(pp1, fig, plt, ax1, ax2_label, vehicle_title, xlabel, ylabel):
    title=plt.title(vehicle_title)
#    plt.getp(title, 'text')            #print out the 'text' property for title
#    plt.setp(title, color='k')
    plt.xlabel(xlabel, fontsize=14)
    plt.ylabel(ylabel, fontsize=14)
#    ax.set_facecolor("white")
#    ax.spines['bottom'].set_color('black')
#    ax.spines['top'].set_color('black')
#    ax.spines['left'].set_color('black')
#    ax.spines['right'].set_color('black')
#    ax.xaxis.label.set_color('black')
#    ax.yaxis.label.set_color('black')
#    ax.tick_params(axis='x', colors='black')
#    ax.tick_params(axis='y', colors='black')
    legend = ax1.legend(loc='best')
#    legend.get_frame().set_facecolor('white')
#    plt.setp(legend.get_texts(), color='k')
#    legend.get_frame().set_linewidth(0.25)
#    legend.get_frame().set_edgecolor("black")
#    ax.grid(which='major', axis='both', linestyle='-', color='k')
    ax1.grid(which='major', axis='both', linestyle='-', color='k', linewidth=0.25)
    ax1.tick_params(axis='both', which='major', labelsize=14)
    ax1.tick_params(axis='both', which='minor', labelsize=14)

    if xlabel == '': ax1.set_xticks([])
    
    if ax2_label != '':
        ax2 = ax1.twinx()
        ax2.set_ylabel(ax2_label, color='b', fontsize=14)
        ax2.set_yticks([])

    if pp1 != '': pp1.savefig(fig, dpi=600)
    
def etime_resetting(etime, vehspd):
    netime = len(etime)
    etime1 = np.zeros(netime)
    etime1[0] = etime[0] 
    dt = 0
    
    vehspd1 = np.zeros(len(vehspd))
    time_span = 600
#    time_span = 25
    tvehspd_zero = istop = istart = 0
    stvehspd_zero = 0
    ip1 = 0
    for ik in range (1, netime-1):
        dt = (etime[ik] - etime[ik-1])
        if ik > istart and vehspd[ik-1] > vehspd_1_kph and vehspd[ik] < vehspd_1_kph and vehspd[ik+1] < 0.5:
                
            istop = ik
            stvehspd_zero = 0
            for j in range(istop+1, netime-1):
                stvehspd_zero += vehspd[j]
                if  (vehspd[j-1] < 0.5 and vehspd[j] < 0.5 and vehspd[j+1] > 1) or stvehspd_zero > 1:
                    istart = j+1
                    tvehspd_zero = istart - istop
                    if tvehspd_zero > time_span:
                        sum_tvehspd_zero = np.sum(vehspd[istop+1:istart-1])
                        print('istart ', istart, ' istop ', istop, ' tvehspd_zero ', tvehspd_zero, ' sum_tvehspd_zero ', stvehspd_zero)
                        istop = istart + 1
                    break

        if dt > time_span or tvehspd_zero > time_span:
            etime1[ik] = etime1[ik-1] + 30
            print('inside time reset, dt = ', dt, ' tvehspd_zero ', tvehspd_zero )
            tvehspd_zero = stvehspd_zero = 0
        else:
           etime1[ik] = etime1[ik-1] + dt
           
    if etime1[ik+1] < 0.1 and etime[netime-1] > 1: etime1[ik+1] = etime1[ik]+1
    
    etime = etime1  
    # etime[201], etime[202]
    del etime1, vehspd1
    return etime, tvehspd_zero     

def subplot_shared_x_axes(sf, pp1, splot_mode, title, x1, y1, y1c, xlabel1, ylabel1, legend1, x2, y2, y2c, xlabel2, ylabel2, legend2, \
                          x3, y3, y3c, xlabel3, ylabel3, legend3, x3r, y3r, y3rc, ylabel3r, legend3r):
    fig = plt.figure(facecolor=(1, 1, 1))
    fig.set_size_inches(11*sf, 8.5*sf, forward=True)
#    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    if len(x2) > 0 and len(y2) > 0:
        subplot1=plt.subplot(311)
        subplot2=plt.subplot(312)
        subplot3 =plt.subplot(313)
    elif len(x3) > 0 and len(y3) > 0: 
        subplot1=plt.subplot(211)
        subplot3 =plt.subplot(212)
    else:
        subplot1=plt.subplot(111)
    
    subplot1.plot(x1, y1, y1c, label=legend1)
    subplot1.set_xlim(0, max(x1))
    subplot1.grid()
    subplot1.set_title(title)
    subplot1.set_ylabel(ylabel1)
    if len(legend1) > 0: subplot1.legend(loc='best')
    if len(x2) == 0 and len(x3) == 0: subplot1.set_xlabel(xlabel1)

    if len(x2) > 0 and len(y2) > 0:
        subplot2.plot(x2, y2, y2c, label=legend2)
        subplot2.set_xlim(0, max(x2))
        if min(y2) < 0 and max(y2) <= 30: 
            subplot2.set_ylim(np.round(min(y2)), np.round(max(y2)*1.5))
        elif min(y2) >= 0  and max(y2) <= 30:
            subplot2.set_ylim(0, np.round(max(y2)+5))
        subplot2.grid()
        subplot2.set_ylabel(ylabel2)
        subplot2.legend(loc='best')
      
    if len(x3) > 0 and len(y3) > 0:
        lns1 = subplot3.plot(x3, y3, y3c, label= legend3)
        subplot3.set_xlim(0, max(x3))
        subplot3.grid()
        subplot3.set_xlabel(xlabel3)
        subplot3.set_ylabel(ylabel3)
    if splot_mode == 'share' and len(x3r) > 0 and len(y3r) > 0:
        ax4 = subplot3.twinx()
        lns2 = ax4.plot(x3r, y3r, y3rc, label=legend3r)
        ax4.set_ylabel(ylabel3r, color='b')
        lns = lns1+lns2
        labs = [l.get_label() for l in lns]
        ax4.legend(lns, labs, loc=0)
    elif len(x3) > 0 and len(y3) > 0:
        subplot3.legend(loc='best')
   
    if splot_mode == 'share' and len(x2) > 0 and len(y2) > 0:
        # subplot1.get_shared_x_axes().join(subplot1, subplot2, subplot3)
        subplot2.sharex(subplot1)
        subplot3.sharex(subplot1)
        subplot1.set_xticklabels([])
        subplot2.set_xticklabels([])
    elif splot_mode == 'share' and len(x2) == 0 and len(y2) == 0 and len(x3) > 0 and len(y3) > 0:
        # subplot2.sharex(subplot1)
        subplot3.sharex(subplot1)
        # subplot1.get_shared_x_axes().join(subplot1, subplot3)
        subplot1.set_xticklabels([])
    
    pp1.savefig(fig, dpi=600)
    
def RDE_NOx_table(RDE_FTPbin_fname, vehicle, RDE_NOx, RDEu_NOx, STD_NOx, BIN_label, FTP_NOxc, US_NOxc, HwFET_NOx):
    
    wb = openpyxl.load_workbook(engineDB_ifile)
    wb1 = openpyxl.load_workbook(engineDB_ofile)
    sheet = wb["Marine SI"]
    sheet1 = wb1["Marine SI"]
    
#    sheet1.insert_rows(idx=1)
    
    header_row = 1
    ncols = 95
    for i in range(1, header_row+1):
        for j in range(1, ncols):
            sheet1.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value
    
    yellow_background = PatternFill(bgColor=colors.YELLOW)
    diff_style = DifferentialStyle(fill=yellow_background)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["$H1<3"]
    sheet1.conditional_formatting.add("A1:CQ1", rule)
    
    times_bold_font = Font(name='Times New Roman', bold=True)
    big_red_text = Font(color=colors.RED, size=20)
    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                right=double_border_side,
                bottom=double_border_side,
                left=double_border_side)
    
    for cell in sheet1["1:1"]:
        cell.font = times_bold_font
    wb1.save(engineDB_ofile)
    
'''
    engine torque from engine load * engine torque = f(engine speed, engine load)
>>> xp = [1, 2, 3]
>>> fp = [3, 2, 0]
>>> np.interp(2.5, xp, fp)
1.0
'''

#def MAW_GHG_WLTC(istart, df):
def timePEMS_to_second(timestr):
    [h, m, s] = timestr.split(':')
    s = s.split('.'); s = s[0]
    return int(h) * 3600 + int(m) * 60 + int(s)
    
def timeOBD_to_second(hms_str):
    str1 = hms_str
    if len(str1) == 6:
        hh, mm, ss = str1[0:2], str1[2:4], str1[4:6]   
    else:
        hh, mm, ss = str1[0:1], str1[1:3], str1[3:5]        
    return hh, mm, ss, (int(hh) * 3600 + int(mm) * 60 + int(ss))
       
def vehspd_obd_sync(ndoe, nsamples_obd_sync, time_start, etime2, vehspd_pems, otime2, vehspd_obd):
    itime_start = time_start
    ntime = len(otime)
    timeoffset_obd=0 
    timeoffset_pems = 0
    
    rms_vehspd_pems = np.zeros(ndoe)
    rms_vehspd_obd = np.zeros(ndoe)
    k = 0
    for j in range (ndoe):
        tmp = 0
        otmp = 0
        for i in range (itime_start, itime_start + nsamples_obd_sync): 
            tmp += (abs(vehspd_pems[j+i] - vehspd_obd[i])) 
            otmp += (abs(vehspd_pems[i] - vehspd_obd[j+i])) 

        rms_vehspd_pems[k] = (tmp**2/nsamples_obd_sync)**0.5
        rms_vehspd_obd[k] = (otmp**2/nsamples_obd_sync)**0.5
        k += 1
        
    min_rms_vehspd_pems = np.min(rms_vehspd_pems)
    min_rms_vehspd_obd = np.min(rms_vehspd_obd)
    if min_rms_vehspd_pems <   min_rms_vehspd_obd:   
        df_sync = pd.DataFrame(rms_vehspd_pems)
        timeoffset_obd = df_sync.loc[df_sync.idxmin()].index[0]
        # otime = otime + timeoffset1
    else:
        df_sync = pd.DataFrame(rms_vehspd_obd)
        timeoffset_pems = df_sync.loc[df_sync.idxmin()].index[0]
        # etime = etime + timeoffset        
    return timeoffset_pems, timeoffset_obd

def add_df(idf, cname, title, unit, istart, iend, idata):  
    idf[cname] = ''
    idf[cname].columns = cname
    icname = title
    idf[cname][1] = unit
    idf[cname][istart:iend+1] = idata

def Relative_Positive_Acceleration(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, dVMTu, dVMTr, dVMTm,  Vu_ACC, Vr_ACC, Vm_ACC):
    RPAu = RPAr = RPAm = 0
    if len(Vu_ACC[Vu_ACC > 0.1]) > 0:
        sum_vehspd_accel_pos = np.sum(vehspd_ACCu[Vu_ACC > 0.1]) # 
        RPAu = sum_vehspd_accel_pos / np.sum(dVMTu)
        
    if len(Vr_ACC[Vr_ACC > 0.1]) > 0:
        sum_vehspd_accel_pos = np.sum(vehspd_ACCr[Vr_ACC > 0.1]) # 
        RPAr = sum_vehspd_accel_pos / np.sum(dVMTr)

    if len(Vm_ACC[Vm_ACC > 0.1]) > 0:
        sum_vehspd_accel_pos = np.sum(vehspd_ACCm[Vm_ACC > 0.1])# 
        RPAm = sum_vehspd_accel_pos / np.sum(dVMTm)
        
    return RPAu, RPAr, RPAm

def VehSpd_Accel_95pct(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, Vu_ACC, Vr_ACC, Vm_ACC):
    vehspd_ACCu_95pct = vehspd_ACCr_95pct = vehspd_ACCm_95pct = 0
    if len(Vu_ACC[Vu_ACC > 0.1]) > 0: vehspd_ACCu_95pct= np.percentile(vehspd_ACCu[Vu_ACC > 0.1], 95)  
    if len(Vr_ACC[Vr_ACC > 0.1]) > 0: vehspd_ACCr_95pct= np.percentile(vehspd_ACCr[Vr_ACC > 0.1], 95)  
    if len(Vm_ACC[Vm_ACC > 0.1]) > 0: vehspd_ACCm_95pct= np.percentile(vehspd_ACCm[Vm_ACC > 0.1], 95)  
    
    return vehspd_ACCu_95pct, vehspd_ACCr_95pct, vehspd_ACCm_95pct

def Vehspd_mean(vehspd_u, vehspd_r, vehspd_m):
    vehspd_u_mean = vehspd_r_mean = vehspd_m_mean = 0
    if len(vehspd_u) > 0: vehspd_u_mean= np.mean(vehspd_u) # 
    if len(vehspd_r) > 0: vehspd_r_mean= np.mean(vehspd_r) # 
    if len(vehspd_m) > 0: vehspd_m_mean= np.mean(vehspd_m)# 
    
    return vehspd_u_mean, vehspd_r_mean, vehspd_m_mean

def RDE_report (rpt, filename_pems, filename_obd, output_folder, now_hms, suffix, trip_distance, trip_duration, cold_start_duration, urban_distance, rural_distance, motorway_distance,
                urban_distance_share, rural_distance_share, motorway_distance_share, urban_avg_speed, rural_avg_speed, motorway_avg_speed, total_trip_avg_speed,
                motorway_speed_above_145kph, motorway_speed_above_100kph, urban_stop_time_pct, elev_abs_diff_start_end, total_cumul_elev_pos_gain, urban_cumul_elev_pos_gain,
                init_idle_duration, cold_start_avg_speed, cold_start_max_speed, cold_stop_time, vehspd_avg_EPA, RPA_avg_EPA, VA_avg_95pct_EPA,
                nVu_ACC, nVr_ACC, nVm_ACC, nMAWt, nMAWu, nMAWr, nMAWm, CO_RDE_mgpkm, NOx_RDE_mgpkm, CO_RDEu_mgpkm, NOx_RDEu_mgpkm, 
                FE_mpg, CO2_gpm_test, engNOx_mgpm_test, CO_mgpm_test):
    
    print("Total trip distance [km] = ", trip_distance)
    print("Total trip duration min. [90-120] = ", trip_duration)
    print("Cold start duration min. [<=5] = ", cold_start_duration)
    
    print("Urban distance, km [>16] = ", urban_distance)
    print("Rural distance, km [>16] = ", rural_distance)
    print("Motorway distance, km [>16] = ", motorway_distance)
    
    print("Urban distance share, % [29-44] = ", np.round(urban_distance_share, 1))
    print("Rural distance share, % [23-43] = ", np.round(rural_distance_share, 1))
    print("Motorway distance share, % [23-43] = ", np.round(motorway_distance_share, 1))
    
    print("Urban average speed [kph] = ", urban_avg_speed)
    print("Rural average speed [kph] = ", rural_avg_speed)
    print("Motorway average speed [kph] = ", motorway_avg_speed)
    print("Total trip average speed [kph] = ", total_trip_avg_speed)
    
    print("Motorway speed above 145 km/h, % [<3% mot. time] = ", motorway_speed_above_145kph)
    print("Motorway speed above 100 km/h, min [>=5] = ", motorway_speed_above_100kph)
    print("Urban stop time, % [6-30] = ", urban_stop_time_pct)
    
    print("Start and end points elevation absolute difference, m [<=100m] = ", elev_abs_diff_start_end)    #print("Start and end points elevation absolute difference averaged, m [<=100m] = ", round((d_alt1 + d_elev)/2, 2))
    print("Total cumulative positive elevation gain, m/100km [<1200m/100km] = ", total_cumul_elev_pos_gain)
    print("Urban cumulative positive elevation gain, m/100km [<1200m/100km] = ", urban_cumul_elev_pos_gain)
    
    if time_engine_idle >= 300:
        print("Idling event(s) longer than 300 s, Yes/No = ", "Yes")
        engine_idle_gt_5min = 'Yes'
    else:
        print("Idling event(s) longer than 300 s, Yes/No = ", "No")
        engine_idle_gt_5min = 'Yes'
        
    print("Idling event(s) longer than 300 s = ", time_engine_idle)
    print("Initial idling duration, s [<=15] = ", init_idle_duration)
    
    print("Cold start average speed, km/h [15-40] = ", cold_start_avg_speed)
    print("Cold start maximum speed, km/h [< 60] = ", cold_start_max_speed)
    print("Cold start stop time, s [<=90] = ", cold_stop_time)

    irow = 0
    rpt['DATA QC'][irow] = filename_pems + '.xlsx';  irow = irow + 1
    rpt['DATA QC'][irow] = filename_obd;  irow = irow + 1
    rpt['DATA QC'][irow] = '';  irow = irow + 1
    rpt['EPA'][irow] = round(FE_mpg, 1);  irow = irow + 1
    rpt['EPA'][irow] = round(CO2_gpm_test, 1);  irow = irow + 1

    irow = 7
    rpt['EPA'][irow] = round(1.60934*CO_RDE_mgpkm, 2);  irow = irow + 1
    rpt['EPA'][irow] = round(1.60934*NOx_RDE_mgpkm, 2);  irow = irow + 1
    rpt['EPA'][irow] = '';  irow = irow + 1
    rpt['EPA'][irow] = round(1.60934*CO_RDEu_mgpkm, 2);  irow = irow + 1
    rpt['EPA'][irow] = round(1.60934*NOx_RDEu_mgpkm, 2);  irow = irow + 1

    irow = 19
    rpt['EPA'][irow] = trip_distance;  irow = irow + 1
    rpt['EPA'][irow] = trip_duration;  irow = irow + 1
    rpt['EPA'][irow] = cold_start_duration;  irow = irow + 1
    rpt['EPA'][irow] = urban_distance;  irow = irow + 1
    rpt['EPA'][irow] = rural_distance;  irow = irow + 1
    rpt['EPA'][irow] = motorway_distance;  irow = irow + 1
    rpt['EPA'][irow] = np.round(urban_distance_share, 1);  irow = irow + 1
    rpt['EPA'][irow] = np.round(rural_distance_share, 1);  irow = irow + 1
    rpt['EPA'][irow] = np.round(motorway_distance_share, 1);  irow = irow + 1
    rpt['EPA'][irow] = urban_avg_speed;  irow = irow + 1
    rpt['EPA'][irow] = rural_avg_speed;  irow = irow + 1
    rpt['EPA'][irow] = motorway_avg_speed;  irow = irow + 1
    rpt['EPA'][irow] = total_trip_avg_speed;  irow = irow + 1
    rpt['EPA'][irow] = motorway_speed_above_145kph;  irow = irow + 1
    rpt['EPA'][irow] = motorway_speed_above_100kph;  irow = irow + 1
    rpt['EPA'][irow] = urban_stop_time_pct;  irow = irow + 1
    
    rpt['EPA'][irow] = elev_abs_diff_start_end;  irow = irow + 1
    rpt['EPA'][irow] = total_cumul_elev_pos_gain;  irow = irow + 1
    rpt['EPA'][irow] = urban_cumul_elev_pos_gain;  irow = irow + 1
    
    rpt['EPA'][irow] = engine_idle_gt_5min;  irow = irow + 1
    rpt['EPA'][irow] = init_idle_duration;  irow = irow + 1
    rpt['EPA'][irow] = cold_start_avg_speed;  irow = irow + 1
    rpt['EPA'][irow] = cold_start_max_speed;  irow = irow + 1
    rpt['EPA'][irow] = cold_stop_time;  irow = irow + 1
    
    irow = 48
    rpt['EPA'][irow] = round(GHG_WLTC_threshold/1000, 3);  irow = irow + 1
    
    irow = 56
    rpt['EPA'][irow] = nMAWt;  irow = irow + 1
    rpt['EPA'][irow] = nMAWt;  irow = irow + 1
    rpt['EPA'][irow] = 0;  irow = irow + 1
    rpt['EPA'][irow] = nMAWu;  irow = irow + 1
    rpt['EPA'][irow] = nMAWr;  irow = irow + 1
    rpt['EPA'][irow] = nMAWm;  irow = irow + 1

    rpt['EPA'][irow] = nMAWu;  irow = irow + 1
    rpt['EPA'][irow] = nMAWr;  irow = irow + 1
    rpt['EPA'][irow] = nMAWm;  irow = irow + 1

    rpt['EPA'][irow] = 100;  irow = irow + 1
    rpt['EPA'][irow] = 100;  irow = irow + 1
    rpt['EPA'][irow] = 100;  irow = irow + 1
    
    irow = 70
    rpt['EPA'][irow] = round(RPA_avg_EPA[0], 3);  irow = irow + 1
    rpt['EPA'][irow] = round(RPA_avg_EPA[1], 3);  irow = irow + 1
    rpt['EPA'][irow] = round(RPA_avg_EPA[2], 3);  irow = irow + 1
    rpt['EPA'][irow] = round(VA_avg_95pct_EPA[0], 2);  irow = irow + 1
    rpt['EPA'][irow] = round(VA_avg_95pct_EPA[1], 2);  irow = irow + 1
    rpt['EPA'][irow] = round(VA_avg_95pct_EPA[2], 2);  irow = irow + 1

    rpt['EPA'][irow] = nVu_ACC;  irow = irow + 1
    rpt['EPA'][irow] = nVr_ACC;  irow = irow + 1
    rpt['EPA'][irow] = nVm_ACC;  irow = irow + 1

    outFilename = os.path.join(output_folder, os.path.basename(filename_pems))
    outFilename= os.path.splitext(outFilename)[0] + '_rpt' + now_hms + suffix
    rpt.to_html(outFilename)

def RWE_report (rpt, filename_pems, filename_obd, output_folder, now_hms, suffix, trip_distance, trip_duration, cold_start_duration, urban_distance, rural_distance, motorway_distance,
                urban_distance_share, rural_distance_share, motorway_distance_share, urban_avg_speed, rural_avg_speed, motorway_avg_speed, total_trip_avg_speed,
                motorway_speed_above_145kph, motorway_speed_above_100kph, urban_stop_time_pct, elev_abs_diff_start_end, total_cumul_elev_pos_gain, urban_cumul_elev_pos_gain,
                init_idle_duration, cold_start_avg_speed, cold_start_max_speed, cold_stop_time, vehspd_avg_EPA, RPA_avg_EPA, VA_avg_95pct_EPA, nVu, nVr, nVm, 
                nMAWt, nMAWu, nMAWu_out, nMAWr, nMAWr_out, nMAWm, nMAWm_out, CO2_RDE_mgpkm, CO_RDE_mgpkm, NOx_RDE_mgpkm, NMHC_RDE_mgpkm, CO2_RDEu_mgpkm, CO_RDEu_mgpkm, NOx_RDEu_mgpkm, NMHC_RDEu_mgpkm, FE_mpg,
                CO2_gpm_test, engNOx_mgpm_test, engNMHC_mgpm_test, CO_mgpm_test, GHG_WLTC_threshold):
    
    nMAWt = nMAWu + nMAWr + nMAWm

    print("Total trip distance [km] = ", trip_distance)
    print("Total trip duration min. [90-120] = ", trip_duration)
    print("Cold start duration min. [<=5] = ", cold_start_duration)
    
    print("Urban distance, km [>16] = ", urban_distance)
    print("Rural distance, km [>16] = ", rural_distance)
    print("Motorway distance, km [>16] = ", motorway_distance)
    
    print("Urban distance share, % [29-44] = ", np.round(urban_distance_share, 1))
    print("Rural distance share, % [23-43] = ", np.round(rural_distance_share, 1))
    print("Motorway distance share, % [23-43] = ", np.round(motorway_distance_share, 1))
    
    print("Urban average speed [kph] = ", urban_avg_speed)
    print("Rural average speed [kph] = ", rural_avg_speed)
    print("Motorway average speed [kph] = ", motorway_avg_speed)
    print("Total trip average speed [kph] = ", total_trip_avg_speed)
    
    print("Motorway speed above 145 km/h, % [<3% mot. time] = ", motorway_speed_above_145kph)
    print("Motorway speed above 100 km/h, min [>=5] = ", motorway_speed_above_100kph)
    print("Urban stop time, % [6-30] = ", urban_stop_time_pct)
    
    print("Start and end points elevation absolute difference, m [<=100m] = ", elev_abs_diff_start_end)
    print("Total cumulative positive elevation gain, m/100km [<1200m/100km] = ", total_cumul_elev_pos_gain)
    print("Urban cumulative positive elevation gain, m/100km [<1200m/100km] = ", urban_cumul_elev_pos_gain)
    
    if time_engine_idle >= 300:
        print("Idling event(s) longer than 300 s, Yes/No = ", "Yes")
        engine_idle_gt_5min = 'Yes'
    else:
        print("Idling event(s) longer than 300 s, Yes/No = ", "No")
        engine_idle_gt_5min = 'Yes'
        
    print("Idling event(s) longer than 300 s = ", time_engine_idle)
    print("Initial idling duration, s [<=15] = ", init_idle_duration)
    
    print("Cold start average speed, km/h [15-40] = ", cold_start_avg_speed)
    print("Cold start maximum speed, km/h [< 60] = ", cold_start_max_speed)
    print("Cold start stop time, s [<=90] = ", cold_stop_time)

    # 'PEMS FILE NAME' in list(rpt.loc[:, 'DATA '])
    if filename_pems != '': 
        target_label = 'PEMS FILE NAME'
        mask = rpt['DATA'].str.contains(target_label, case=False)
        _index = np.where(mask)[0].tolist()
        rpt.loc[_index, 'Type'] = filename_pems
    if filename_obd != '': 
        target_label = 'ECU FILE NAME'
        mask = rpt['DATA'].str.contains(target_label, case=False)
        _index = np.where(mask)[0].tolist()
        rpt.loc[_index, 'Type'] = filename_obd
                
    # mask = rpt['DATA'].str.contains('RDE Fuel Economy - Total', case=False)
    # _index = np.where(mask)[0].tolist()
    # rpt.loc[_index, 'RWE Results'] = FE_mpg

    _strings = ['Fuel Economy - Total', 'CO2 emissions - Total', 'CO2 MAW emissions - Total', 'CO2 MAW emissions - Urban', 'CO emissions - Total', 'CO MAW emissions - Total', 'CO MAW emissions - Urban', 
                'NOx emissions - Total', 'NOx MAW emissions - Total', 'NOx MAW emissions - Urban', 'NMHC emissions - Total', 'NMHC MAW emissions - Total', 'NMHC MAW emissions - Urban']
    CO2_RDE_mgpkm, CO2_RDEu_mgpkm, CO_RDE_mgpkm, CO_RDEu_mgpkm = 1.60934*CO2_RDE_mgpkm/1000, 1.60934*CO2_RDEu_mgpkm/1000, 1.60934*CO_RDE_mgpkm, 1.60934*CO_RDEu_mgpkm
    NOx_RDE_mgpkm, NOx_RDEu_mgpkm, NMHC_RDE_mgpkm, NMHC_RDEu_mgpkm = 1.60934*NOx_RDE_mgpkm, 1.60934*NOx_RDEu_mgpkm, 1.60934*NMHC_RDE_mgpkm, 1.60934*NMHC_RDEu_mgpkm
    _values = [FE_mpg, CO2_gpm_test, CO2_RDE_mgpkm, CO2_RDEu_mgpkm, CO_mgpm_test, CO_RDE_mgpkm, CO_RDEu_mgpkm, engNOx_mgpm_test, NOx_RDE_mgpkm, NOx_RDEu_mgpkm, engNMHC_mgpm_test, NMHC_RDE_mgpkm, NMHC_RDEu_mgpkm]
    for i in range(len(_strings)):
        mask = rpt['DATA'].str.contains(_strings[i], case=False)
        _index = np.where(mask)[0].tolist()
        rpt.loc[_index[0], 'RWE Results'] = np.round(_values[i], 2)

    _strings = ['Total trip distance', 'Total trip duration', 'Cold start duration', 'Data in extreme conditions', 'Urban distance', 'Rural distance', 'Motorway distance', 'Urban distance share', 'Rural distance share', 'Motorway distance share', 
                'Urban average speed']
    _values = [trip_distance, trip_duration, cold_start_duration, 0, urban_distance, rural_distance, motorway_distance, urban_distance_share, rural_distance_share, motorway_distance_share, urban_avg_speed] 
    for i in range(len(_strings)):
        mask = rpt['DATA'].str.contains(_strings[i], case=False)
        _index = np.where(mask)[0].tolist()
        rpt.loc[_index[0], 'RWE Results'] = np.round(_values[i], 2)

    _strings = ['Motorway speed above 145 km/h', 'Motorway speed above 100 km/h', "Idling event longer than 300s", 'Initial idling duration', 'Urban stop time', 'Start-end elevation absolute difference', 
                'Total cumulative positive elevation gain', 'Urban cumulative positive elevation gain', 'Cold start average speed', 'Cold start maximum speed', 'Cold start stop time']
    _values = [motorway_speed_above_145kph, motorway_speed_above_100kph, time_engine_idle, init_idle_duration, urban_stop_time_pct, elev_abs_diff_start_end, 
               total_cumul_elev_pos_gain, urban_cumul_elev_pos_gain, cold_start_avg_speed, cold_start_max_speed, cold_stop_time]
    for i in range(len(_strings)):
        mask = rpt['DATA'].str.contains(_strings[i], case=False)
        _index = np.where(mask)[0].tolist()
        if len(_index) == 0:
            print(_strings[i])
            continue
        try:
            rpt.loc[_index[0], 'RWE Results'] = np.round(_values[i], 2)
        except ValueError:
            print(_values[i], "Invalid values for np.round")
    
    _strings = ['Share of normal urban windows', 'Share of normal rural windows', 'Share of normal motorway windows', 'Urban RPA', 'Rural RPA', 'Motorway RPA', 'Urban 95th percentile Speed', 'Rural 95th percentile Speed', 
                'Motorway 95th percentile Speed', 'Number of counts - Urban', 'Number of counts - Rural', 'Number of counts - Motorway', 'Rural average speed', 'Motorway average speed', 'Total trip average speed',
                'Reference WLTC CO2 mass', 'Total number of windows', 'Number of urban windows', 'Number of rural windows', 'Number of motorway windows']
    nMAWu_pct, nMAWr_pct, nMAWm_pct = max(0, (1 - nMAWu_out/nMAWu)*100), (1 - nMAWr_out/nMAWr)*100, (1 - nMAWm_out/nMAWm)*100
    _values = [nMAWu_pct, nMAWr_pct, nMAWm_pct, RPA_avg_EPA[0], RPA_avg_EPA[1], RPA_avg_EPA[2], VA_avg_95pct_EPA[0], VA_avg_95pct_EPA[1], VA_avg_95pct_EPA[2], nVu, nVr, nVm, rural_avg_speed, motorway_avg_speed, 
               total_trip_avg_speed, GHG_WLTC_threshold/1000, nMAWt, nMAWu, nMAWr, nMAWm]
    # ≤
    for i in range(len(_strings)):
        mask = rpt['DATA'].str.contains(_strings[i], case=False)
        _index = np.where(mask)[0].tolist()
        if len(_index) == 0:
            print(_strings[i])
            continue
        try:
            rpt.loc[_index[0], 'RWE Results'] = np.round(_values[i], 2)
        except ValueError:
            print(_values[i], "Invalid values for np.round")

    outFilename = os.path.join(output_folder, os.path.basename(filename_pems))
    outFilename= os.path.splitext(outFilename)[0] + '_rpt' + now_hms + suffix
    rpt.to_html(outFilename)

    return rpt

'''
https://stackoverflow.com/questions/38511373/change-the-color-of-text-within-a-pandas-dataframe-html-table-python-using-style
'''        
def vehmove_start_end(df, dt, vehmove_start_ini, Inst_Mass_GHG, GHG_WLTC_threshold):
    
    nel = len(df)
    j = k = 0
    vehmove_start_kph = 0.1
    vehmove_stop_kph = 0.01
    vehmove_start_min = 30
    if dt == 1:
        vehmove_start_min = vehmove_start_min
    else:
        vehmove_start_min = vehmove_start_min/dt
        
    vehmove_start = []
    vehmove_end = []
    for i in range(vehmove_start_ini+1, nel):
        if (df['vehspd_ecu'][i] > vehmove_start_kph) and (df['vehspd_ecu'][i-1]) <= vehmove_stop_kph and np.sum(Inst_Mass_GHG[i:-1]) > GHG_WLTC_threshold:
            if (i >= k+vehmove_start_min) and (i > vehmove_start_min) and ((len(vehmove_start) == 0) or (len(vehmove_start) == len(vehmove_end))): 
                vehmove_start.append(i)
                j = i
        elif (df['vehspd_ecu'][i] <= vehmove_stop_kph) and (df['vehspd_ecu'][i-1] > vehmove_start_kph) and (np.sum(Inst_Mass_GHG[i:-1]) > GHG_WLTC_threshold):
            if (i > j) and len(vehmove_end) < len(vehmove_start): 
                if (i - j) < vehmove_start_min: i = i+vehmove_start_min
                vehmove_end.append(i-1)
                k = i
                
    vehmove_end_check = False
    for j in range(len(vehmove_start)-1):
        if vehmove_end[j] <= vehmove_start[j]:
            vehmove_end_check = True
            break

    if (len(vehmove_end) < len(vehmove_start)):
        _idx_ends = df.loc[(df.index > vehmove_start[-1]) & (df['vehspd_ecu'] > vehmove_stop_kph) & (Inst_Mass_GHG.iloc[::-1].cumsum().iloc[::-1] < GHG_WLTC_threshold)].index
        if len(_idx_ends) > 0: 
            vehmove_end = np.concatenate((vehmove_end, _idx_ends[0] - 1), axis=None)
        else:
            vehmove_end = np.concatenate((vehmove_end, vehmove_start[-1] + int(300/dt)), axis=None)
    elif (len(vehmove_end) > len(vehmove_start)) or (vehmove_end_check):
        vehmove_end1 = []
        for k in range(len(vehmove_start)):
            if k == len(vehmove_start) - 1:
                _idx_ends = df.loc[(df.index > vehmove_start[k]) & (df['vehspd_ecu'] > 0.0) & (Inst_Mass_GHG.iloc[::-1].cumsum().iloc[::-1] < GHG_WLTC_threshold)].index
            else:
                _idx_ends = df.loc[(df.index > vehmove_start[k]) & (df.index < vehmove_start[k+1]-1) & (df['vehspd_ecu'] > vehmove_stop_kph) & (Inst_Mass_GHG.iloc[::-1].cumsum().iloc[::-1] < GHG_WLTC_threshold)].index
                
            if len(_idx_ends) > 0:
                vehmove_end1.append(_idx_ends[0]-1)
            else:
                if k < len(vehmove_start) - 1: vehmove_end1.append(vehmove_start[k+1]-1)

            vehmove_end = vehmove_end1
            del vehmove_end1
    num_vehmove_start = len(vehmove_start)    

    return vehmove_start, vehmove_end, num_vehmove_start

def read_pp_files(df, pp_files):
    # npp_files = len(pp_files)
    for i in range(len(pp_files)):
        pp_file = str(pp_files.fname[i])
        if i == 0:
            # ECT_unit = iVEH_SPEED_unit = imVEH_SPEED_unit = iGPS_GROUND_SPEED_unit = Tamb_unit = Tamb_cname = '' 
            df = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape', low_memory=False) # add "flot" in the SG_fuel

            data_columns_units = dict(zip(df.columns, list(df.loc[0, :])))
            df.drop(df.index[0], inplace=True) # delete the units row
            df.iloc[:, 3:] = df.iloc[:, 3:].apply(pd.to_numeric, errors='coerce').fillna(0)
            column_names, column_units = [], []
            columns_check = ['sTIME', 'iSCB_LAP', 'iSCB_LAT', 'AmbientTemp', 'icMASS_FLOW', 'EV_std', 'iFLOW_EX_TEMP', 'iCOOL_TEMP', 'imCOOL_TEMP', 'iMAP', 'iENG_SPEED', 'iENG_LOAD', 'iVEH_SPEED', 'imVEH_SPEED', \
                'CATEMP11', 'iREL_THROT_POS', 'Lambda', 'iWfgps', 'iWf', 'iPPTorq', 'iBhp', 'iCALCRT_CO2m', 'iCALCRT_COm', 'iCALCRT_NOm', 'iCALCRT_NO2m', 'iCALCRT_NOxm', 'iCALCRT_kNOm', 'iCALCRT_kNO2m', 'iCALCRT_kNOxm', \
                    'iCALCRT_HCm', 'iCALCRT_CH4m', 'iCALCRT_NMHCm', 'iCALCRT_O2m', 'iHUM_TEMP', 'iFLOW_MASS_FLOW', 'iGPS_LAT', 'iGPS_LON', 'iGPS_ALT', 'iGPS_GROUND_SPEED', 'sDATE', 'sSTATUS_PATH']
            common_fev_epa_var = df.columns.intersection(columns_check)        
            df = df[common_fev_epa_var]            
            
            try: 
                for text in common_fev_epa_var:
                    if text not in columns_check: continue
                    unit = _unit = data_columns_units.get(text)
                    if 'sTIME' in text: _unit = 'hhmmss' if (unit == 'hh:mm:ss.xxx') or (unit == 'hh:mm:ss') else 's'
                    elif 'iCOOL_TEMP' in text: _unit = 'deg F' if (unit == 'deg F') or (unit == 'degF') or (unit == 'Deg F') else 'deg C'
                    elif 'imCOOL_TEMP' in text: _unit = 'deg F' if (unit == 'deg F') or (unit == 'degF') or (unit == 'Deg F') else 'deg C'
                    elif 'iWfgps' in text: _unit = 'gal/s' if (unit == 'gal/s') else 'g/s'
                    elif 'iWf' in text: _unit = 'g/s' if (unit == 'g/s') else 'g/s'
                    elif ('CATEMP11') in text: _unit = 'deg F' if (unit == 'Deg F') or (unit == 'degF') or (unit == 'deg F') else 'deg C'
                    elif 'imVEH_SPEED' in text: _unit = 'kph' if (unit == 'kph') or (unit == 'km/h') or (unit == 'km/hr') else 'mph'
                    elif 'iVEH_SPEED' in text: _unit = 'mph' if (unit == 'mph') or (unit == 'mi/h') else 'kph' 
                    elif 'iGPS_GROUND_SPEED' in text: _unit = 'mph' if (unit == 'mph') or (unit == 'MPH') else 'kph'

                    elif 'iENG_SPEED' in text: _unit = 'rpm' if (unit == 'RPM') or (unit == 'rpm') or (unit == '1/min') else 'rps'
                    elif 'iENG_LOAD' in text: _unit = '%' if (unit == '%') else 'flot'              
                    elif 'iMAP' in text: _unit = 'kPa' if (unit == 'kPaA') or (unit == 'kpa') else 'psi'              
                    elif ('iSCB_LAT' or 'AmbientTemp') in text: _unit = 'deg F' if (unit == 'Deg F') or (unit == 'degF') or (unit == 'deg F') else 'deg C'

                    if (text == 'imCOOL_TEMP'): 
                        if max(df[text]) < 1: continue
                        text = "ECT"
                        df = df.rename(columns={'imCOOL_TEMP': text})
                    elif (text == 'iPPTorq'):
                        if max(df[text]) < 50: continue
                        text = "iENG_TORQ"
                        df = df.rename(columns={'iPPTorq': text})
                    elif (text == 'iBhp'): 
                        if max(df[text]) < 50: 
                            continue
                        text = "iENG_HP"
                        df = df.rename(columns={'iBhp': text})
                                                                        
                    if len(df[text]) > 0:
                        column_names.append(text)
                        column_units.append(_unit)
                
                EPA_PEMS_columns_units = dict(zip(column_names, column_units))     
                del column_names, column_units               
            except ImportError:
                iVEH_SPEED_cname = ''
                iGPS_GROUND_SPEED_cname = ''
                iENG_SPEED_cname = ''
                iENG_LOAD_cname = ''

            # df = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape') # add "flot" in the SG_fuel
            # df = df.drop(df.index[0])
            df = df.reset_index(drop = True)
            df = df.replace(np.nan, '', regex=True)

            itmp = df[df['sDATE'].str.match('Vehicle Description:')].index[0] # df[df['model'].str.contains('ac')]
            vehicle_title = df.sTIME[itmp] + ', '
            itmp = df[df['sDATE'].str.match('Summary Information:')].index[0] - 3 # df[df['model'].str.contains('ac')]
            
            # nel_df1 = len(df.sTIME)
            df = df.drop(df.index[itmp : len(df)])
            df = df.drop(df[df.sSTATUS_PATH == ''].index)
            if 'iGPS_GROUND_SPEED' in EPA_PEMS_columns_units.keys(): 
                df = df.drop(df[df['iGPS_GROUND_SPEED'] == ''].index)
            df = df.drop(df[df.sSTATUS_PATH == 'STANDBY'].index)
            df = df.drop(df[df.sSTATUS_PATH == 'CALIBRATION'].index)
            df = df.reset_index(drop = True)
            # nel_df1 = len(df)
            # df1 = df 
            # df1_file = pp_file
            df2 = []
        else:
            dft = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape', low_memory=False) #encoding = "ISO-8859-1")
            dft = dft.drop(dft.index[0])
            dft.iloc[:, 3:] = dft.iloc[:, 3:].apply(pd.to_numeric, errors='coerce').fillna(0)
            dft = dft.reset_index(drop = True)
            dft = dft.replace(np.nan, '', regex=True)

            itmp = dft[dft['sDATE'].str.match('Summary Information:')].index[0] - 3 # df[df['model'].str.contains('ac')]        
            nel_dft = len(dft.sTIME)

            dft = dft.drop(dft.loc[itmp : nel_dft].index)
            dft = dft.drop(dft[dft.sSTATUS_PATH == ''].index)
            if 'iGPS_GROUND_SPEED' in EPA_PEMS_columns_units.keys(): 
                dft = dft.drop(dft[dft['iGPS_GROUND_SPEED'] == ''].index)
            dft = dft.drop(dft[dft.sSTATUS_PATH == 'STANDBY'].index)
            dft = dft.drop(dft[dft.sSTATUS_PATH == 'CALIBRATION'].index)
            #dft = dft.sort_values(['sTIME'], ascending=[True])
            dft = dft.reset_index(drop = True)
            dft.index = range(len(df), len(df)  + len(dft))
            if i == 1: df2 = dft 
            # df2_file = pp_file
                
            df = pd.concat([df, dft])

    # df = df.drop(df[df.sSTATUS_PATH == 'STANDBY'].index)
    # df = df.drop(df[df.sSTATUS_PATH == 'CALIBRATION'].index)
    # df = df.reset_index(drop=True)
    
    return df, EPA_PEMS_columns_units, vehicle_title, pp_file, df, df2

def engine_torques(engine_speed_rpm, rpm_table, torque_table):
    """
    Estimates engine torque for a given engine speed using linear interpolation.

    Args:
        engine_speed_rpm (float or array_like): The current engine speed(s) in RPM.
        rpm_table (list or array_like): The list of known RPM points (x-coordinates).
        torque_table (list or array_like): The list of corresponding torque values (y-coordinates).

    Returns:
        float or array_like: The estimated torque value(s) in N·m.
    """
    # np.interp performs linear interpolation
    # It also handles extrapolation by using the constant values at the edges of the table if the input is outside the range.
    estimated_torque = np.interp(engine_speed_rpm, rpm_table, torque_table)
    return estimated_torque

def df_SMA_by_second(df, time_column, time_SMA, EPA_PEMS_columns_units):
    
    dt = df[time_column][1] - df[time_column][0]
    factors = int(round(time_SMA/dt, 0))
    df_rolling = df.rolling(window=int(factors)).mean().dropna().reset_index(drop = True)
    df_rolling['sTIME'] = df_rolling['sTIME'] - df_rolling['sTIME'][0]

    _time_by_resampling = np.arange(df_rolling[time_column][0], df_rolling[time_column][len(df_rolling)-1], time_SMA)
    mass_flowrate_units = ["kg/hr",	"SCFM",	"l/s",	"gal/s", "g/s"]    
    df_new = pd.DataFrame()
    for i, text in enumerate(EPA_PEMS_columns_units.keys()):
        _unit = EPA_PEMS_columns_units.get(text)
        f_linear = interpolate.interp1d(df_rolling[time_column], df_rolling[text], kind='linear')
        df_new[text] = pd.DataFrame(f_linear(_time_by_resampling))
        if _unit in mass_flowrate_units:
            df_new[text] = df_new[text] * factors
    
    # plt.plot(df['sTIME'], df['iCALCRT_NOxm'], 'b-')
    # plt.plot(df_new['sTIME'], df_new['iCALCRT_NOxm']/factors, 'r--')
    # print('NOx =', df['iCALCRT_NOxm'].sum()*dt,  df_new['iCALCRT_NOxm'].sum()*dt)
    del df_rolling, df
    
    return df_new

def df_resample_by_second(df, time_column, time_resampling, EPA_PEMS_columns_units):
    
    df['sTIME'] = df['sTIME'] - df['sTIME'][0]
    dt = df[time_column][1] - df[time_column][0]
    factors = time_resampling/dt
    
    mass_flowrate_units = ["kg/hr",	"SCFM",	"l/s",	"gal/s", "g/s"]
    _time_by_resampling = np.arange(df[time_column][0], df[time_column][len(df)-1], time_resampling)

    df_new = pd.DataFrame()
    for i, text in enumerate(EPA_PEMS_columns_units.keys()):
        # if text == time_column:
        #     df_new[text] = pd.DataFrame(_time_by_resampling)
        #     continue
        
        _unit = EPA_PEMS_columns_units.get(text)
        f_linear_mass_flows = interpolate.interp1d(df[time_column], df[text], kind='linear')
        df_new[text] = pd.DataFrame(f_linear_mass_flows(_time_by_resampling))
        if _unit in mass_flowrate_units:
            df_new[text] = df_new[text] * factors

    # print('CO2 total = ',  df0['iCALCRT_CO2m'].sum(), df_new['iCALCRT_CO2m'].sum())
    # print('CO total = ',  df0['iCALCRT_COm'].sum(), df_new['iCALCRT_COm'].sum())
    del df

    return df_new
    
def read_FEV_PEMS(df, pp_files, fev_to_epa_var, epa_units, fev_units):
    for i in range (len(pp_files)):
        pp_file = str(pp_files.fname[i])
        
        if i == 0:            
            df = pd.read_csv(pp_file, skiprows=0, escapechar='\\', encoding = 'unicode_escape', low_memory=False)
            data_columns_units = dict(zip(df.columns, list(df.loc[0, :])))
            df.drop(df.index[0], inplace=True) # delete the units row
            # df = df.iloc[1:].reset_index(drop=True) # drop the first row containing units
            column_names, column_units = [], []
            df.dropna(how='all', inplace=True)
            common_fev_epa_var = df.columns.intersection(list(fev_units.keys()))          
            df = df[common_fev_epa_var]
            df.dropna(how='all', inplace=True)
            df.reset_index(drop=True, inplace=True)
            df = df.apply(pd.to_numeric, errors='coerce').fillna(0)
            dt = df['TMB_REC'][1] - df['TMB_REC'][0] if ('TMB_REC' in df.columns) else 0.1

            try: 
                for i, text in enumerate(common_fev_epa_var):
                    fev_unit = fev_units.get(text)
                    text_RWE = fev_to_epa_var.get(text)
                    epa_unit = epa_units.get(text_RWE)
                    epa_unit = "deg C" if (epa_unit == "degC") or (epa_unit == "Deg C") else epa_unit
                
                    if (fev_unit == 's') and (epa_unit == 'hh:mm:ss.xxx'):
                        if (text_RWE == "sTIME"): _unit = fev_unit
                    elif (fev_unit == 'hh:mm:ss.xxx') and (epa_unit == 'hh:mm:ss.xxx'): 
                        if (text_RWE == "sTIME"): _unit = 'hhmmss' 
                    elif (fev_unit == 'kPa') and (epa_unit == 'mbar'): _unit, df[text] = 'kPa', 10 * df[text]
                    elif (fev_unit == 'mph') and (epa_unit == 'mph') or (fev_to_epa_var.get(text) == "iGPS_GROUND_SPEED"):
                        if (text_RWE == "iVEH_SPEED"): _unit = epa_unit
                        elif (text_RWE == "iGPS_GROUND_SPEED"): _unit = epa_unit
                    elif (fev_unit == '°C') and ((epa_unit == 'deg F') or (epa_unit == 'Deg F')) or (text_RWE == "iCOOL_TEMP"):
                        df[text] = 9/5 * df[text] + 32 # convert to deg F 
                        if (epa_unit == 'Deg F'): epa_unit = 'deg F'
                        if (text_RWE == "iCOOL_TEMP"): _unit = epa_unit
                    elif (text_RWE == "CATEMP11"): 
                        _unit = epa_unit
                        if epa_unit == ((epa_unit == 'deg F') or (epa_unit == 'Deg F')):
                            df[text] = (9/5) * df[text] + 32 # convert back to deg F
                    elif (fev_unit == '°C') and ((epa_unit == 'deg C') or (epa_unit == "degC") or (epa_unit == "Deg C")) or (text_RWE == "imCOOL_TEMP"):
                        if (text_RWE == "iSCB_LAT") or (text_RWE == 'AAT') or (text_RWE == 'iAAT' or text_RWE == 'AmbientTemp'): 
                            _unit = epa_unit
                        else: 
                            _unit = epa_unit                       
                    elif (fev_unit == 'm³/s') and (epa_unit == 'SCFM'):
                        df[text], _unit = 2118.8800032 * df[text] * dt, epa_unit
                    elif (fev_unit == 'km/h') and (epa_unit == 'mph'):
                        df[text], _unit = 0.621371 * df[text], epa_unit
                    elif (fev_unit == 'km/h') and ((epa_unit == 'kph') or (epa_unit == 'km/h')) or (text_RWE == 'imVEH_SPEED'): _unit = 'kph'
                    elif (fev_unit == 'kw') and (epa_unit == 'bhp'):
                        df[text], _unit = 1.34102 * df[text], epa_unit
                    elif (fev_unit == 'mg/s') and (epa_unit == 'g/s'): 
                        # if text == "MF_CO_EGH_TP":
                        #     print(text, text_RWE)
                        df[text], _unit = 0.001 * df[text] * dt, epa_unit                    
                    elif (fev_unit == 'g/s') and (epa_unit == 'gal/s') and (text_RWE == 'iWfgps'):
                        df[text], _unit = 0.000358 * df[text] * dt, epa_unit
                    elif (fev_unit == 'kg/h' and epa_unit == 'kg/hr') or (fev_unit == 'g/s' and epa_unit == 'g/s'):
                        # if text == "MF_CO2_EGH_TP":
                        #     print(text, text_RWE)
                        df[text], _unit = df[text] * dt, epa_unit
                    elif (fev_unit == '1/min') and (epa_unit == 'rpm'):
                        _unit = epa_unit
                    elif (fev_unit == 'Nm') and (epa_unit == 'lb-ft'): df[text], _unit = 0.7375621493 * df[text], epa_unit
                    elif (fev_unit == '1') and (epa_unit == 'flot'): _unit = epa_unit
                    else:
                        _unit = epa_unit                     

                    if (text == "TRQ_EN_ECU") and ("PWR_E_EN" not in df.columns): 
                        if ("iENG_SPEED" in df.columns) and (fev_unit == 'Nm') and (epa_units.get("iENG_SPEED") == 'rpm'):
                            df["iENG_HP"] = 1.34102 * (df["TRQ_EN_ECU"] * df["iENG_SPEED"])/9550
                        elif ("iENG_SPEED" in df.columns) and (fev_unit == 'lb-ft') and (epa_units.get("iENG_SPEED") == 'rpm'):
                            df["iENG_HP"] = (df["TRQ_EN_ECU"] * df["iENG_SPEED"])/5252
                        elif ("RS_CKS_ECU" in df.columns) and (fev_unit == 'Nm') and ((epa_units.get("iENG_SPEED") == 'rpm') or (fev_units.get("RS_CKS_ECU") == '1/min')):
                            df["iENG_HP"] = 1.34102 * (df["TRQ_EN_ECU"] * df["iENG_SPEED"])/9550
                        _unit = "bhp"
                        # PWR_E_EN_calc = TRUE
                    # if text == "TMB_REC":
                    #     df = df.loc[df[text] >= 0.0, :].reset_index(drop=True)

                    elif (text_RWE == 'iBhp'): 
                        if max(df[text]) < 50: continue
                        text_RWE = "iENG_HP"

                    if (text_RWE == 'iPPTorq'): text_RWE = 'iENG_TORQ'
                    if (text_RWE == 'imCOOL_TEMP'): text_RWE = 'ECT'
                    df = df.rename(columns={text: text_RWE}) #  , inplace=True)
                                                                                                
                    if len(df[text_RWE]) > 0:
                        column_names.append(text_RWE)
                        column_units.append(_unit)

                    # if 'iGPS_GROUND_SPEED' in text: # Ground Speed	iGPS_GROUND_SPEED	mph
                    #     iGPS_GROUND_SPEED_cname = 'iGPS_GROUND_SPEED'
                    #     icol_iGPS_GROUND_SPEED = columns.index(iGPS_GROUND_SPEED_cname)
                    #     if ('mph' in units[icol_iGPS_GROUND_SPEED]) or ('MPH' in units[icol_iGPS_GROUND_SPEED]):
                    #         iGPS_GROUND_SPEED_unit = 'mph'
                    #     elif ('kph' in units[icol_iGPS_GROUND_SPEED]) or ('KPH' in units[icol_iGPS_GROUND_SPEED]):
                    #         iGPS_GROUND_SPEED_unit = 'kph'
                    #     else:
                    #         iGPS_GROUND_SPEED_unit = 'mph'
                                                
                EPA_PEMS_columns_units = dict(zip(column_names, column_units))     
                del column_names, column_units     
            except ImportError:
                iVEH_SPEED_cname = ''
                iGPS_GROUND_SPEED_cname = ''
                iENG_SPEED_cname = ''
                iENG_LOAD_cname = ''
            
            # df = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape')# add "flot" in the SG_fuel
            df = df.reset_index(drop = True)
            df = df.replace(np.nan, '', regex=True)
            df = df
            df2 = []
            df1_file = pp_file
            vehicle_title = []
        else:
            dft = pd.read_csv(pp_file, skiprows=1, escapechar='\\', encoding = 'unicode_escape', low_memory=False) #encoding = "ISO-8859-1")
            dft = df
            dft = dft.drop(dft.index[0])
            dft = dft.reset_index(drop = True)
            dft = dft.apply(pd.to_numeric, errors='coerce').fillna(0)
            # dft = dft.replace(np.nan, '', regex=True)
            dft.index = range(len(df), len(df)  + len(dft))
            if i == 1:
                df2 = dft
            df2_file = pp_file
            df = pd.concat([df, dft])

    time_resampling = time_SMA = 1
    df = df_resample_by_second(df, 'sTIME', time_resampling, EPA_PEMS_columns_units)
    # df = df_SMA_by_second(df, 'sTIME', time_SMA, EPA_PEMS_columns_units)
    
    return [df, EPA_PEMS_columns_units, vehicle_title, pp_file, df, df2]

t = time.time()
path = os.getcwd() + '/'

'''
MAW CO2 curve page 156
1) p1  to p2
    Mco2 = a1 x vk + b1
    where, a1 = (MCo2 at p2 - Mcos2 at p1)/(vk at p2 - vk at p1)
    b1 = Mco2 at p1 - a1 x vk at p1
    
2) p2 to p3
    a2 = (Mco2 at p3 - Mco2 at p2) / (vk at p3 - vk at p2) 
    b2 = Mco2 at p2 - a2 * vk at p2    
    
    Mco2 = a2 vk + b2    
'''
    
RDE_report_settings_file = select_file('RDE Report & Settings Excel file')
xls = pd.ExcelFile(RDE_report_settings_file)
rpt = pd.read_excel(xls, sheet_name='RDE REPORT', index_col=None, header=0, skiprows=1, dtype=object)
rpt = rpt.replace(np.nan, '', regex=True)

# rpt2020 = pd.read_excel(xls, sheet_name='RWE REPORT-new', index_col=None, header=0, skiprows=1, dtype=object)
rpt2020 = pd.read_excel(xls, sheet_name='RWE REPORT-2020', index_col=None, header=0, skiprows=1, dtype=object)
# Drop any columns that contain the substring "Unnamed"
rpt2020 = rpt2020.drop(rpt2020.columns[rpt2020.columns.str.contains('Unnamed', case=False)], axis=1)
rpt2020 = rpt2020.replace(np.nan, '', regex=True)

df = pd.read_excel(xls, sheet_name='Settings', index_col=None, header=0, skiprows=0, dtype=object)
df = df.replace(np.nan, '', regex=True)

settings_MAw_CO2_pct = 50; etime_reset = FALSE; vehicle_type = 1; OBD = FALSE 
LAB_WLTC_NOx_Fit = FALSE; Exhaust_Emissions_Standard = 'T2Bin4'; WLTC_GHG_gpkm = ''; WLTC_GHGu_gpkm = ''

if df['PASS-FAIL CALCULATIONS'][0] == 'CO2 window [%]': settings_MAw_CO2_pct = float(df['Default'][0])
if df['PASS-FAIL CALCULATIONS'][1] == 'etime_reset': etime_reset = df['Default'][1]
if df['PASS-FAIL CALCULATIONS'][2] == 'VEH_TYPE': vehicle_type = int(df['Default'][2])
if df['PASS-FAIL CALCULATIONS'][3] == 'sync_OBD': OBD = df['Default'][3]
if df['PASS-FAIL CALCULATIONS'][4] == 'LAB_WLTC_NOx_Fit': LAB_WLTC_NOx_Fit = df['Default'][4]
if df['PASS-FAIL CALCULATIONS'][5] == 'Exhaust_Emissions_Standard': Exhaust_Emissions_Standard = df['Default'][5]
if df['PASS-FAIL CALCULATIONS'][6][0:17] == 'RDE Mode (LD/HD)': RDE_Mode = df['Default'][6]
if df['PASS-FAIL CALCULATIONS'][7][0:21] == 'Calc NOx (g/bhp-hr)': calc_NOx_gbhp_hr = df['Default'][7]
if df['PASS-FAIL CALCULATIONS'][8][0:18] == 'Peak Torque (NM)': engine_peak_trq = float(df['Default'][8])
if df['PASS-FAIL CALCULATIONS'][9][0:14] == 'WLTC Urban CO2': WLTC_GHGu_gpkm = float(df['Default'][9])/1.60934
FEV_to_EPA_PEMS = True if (df['PASS-FAIL CALCULATIONS'][10] == "FEV_to_EPA_PEMS") and (df.loc[10, "Default"] == True) else False

WLTC_CO2_gpkm = pd.to_numeric(df['WLTC_CO2_gpkm'][0:4], errors='coerce') # df['WLTC_CO2_gpmi'][0:4].values/1.60934
WLTC_GHG_gpkm = WLTC_CO2_comp_gpkm = float(df['WLTC_CO2_gpkm'][4]) # /1.60934

WLTC_CO2_gpmi = pd.to_numeric(df['WLTC_CO2_gpmi'][0:4], errors='coerce') # df['WLTC_CO2_gpmi'][0:4].values/1.60934
WLTC_NMHC_gpkm = pd.to_numeric(df['NMHC_WLTC_gpmi'][0:4], errors='coerce') # df['WLTC_CO2_gpmi'][0:4].values/1.60934
WLTC_CO2_comp_gpmi = float(df['WLTC_CO2_gpmi'][4])
WLTC_NMHC_comp_gpmi = float(df['NMHC_WLTC_gpmi'][4])
NMHC_STD_mgpmi = 1000 * WLTC_NMHC_comp_gpmi
# WLTC_CO2_gpkm[0] = WLTC_CO2_gpkm[0] 
# WLTC_CO2_gpkm[1] = WLTC_CO2_gpkm[1] 
# WLTC_CO2_gpkm[2] = WLTC_CO2_gpkm[2] 
# WLTC_CO2_gpkm[3] = WLTC_CO2_gpkm[3] 

WLTC_vehspd_kph = pd.to_numeric(df['vehspd_WLTC_kph'][0:4], errors='coerce') # df['vehspd_WLTC_kph'][0:4].values
WLTC_vehspd_comp_kph = float(df['vehspd_WLTC_kph'][4]) # WLTC vehicle speed composite
WLTC_dist_km = pd.to_numeric(df['WLTC_dist_km'][0:4], errors='coerce') # df['WLTC_dist_km'][0:4].values
WLTC_dist_comp_km = float(df['WLTC_dist_km'][4]) # WLTC vehicle speed composite

settings_avg_speed = pd.to_numeric(df['MAW_vehspd_kph'][0:16], errors='coerce') # df['MAW_vehspd_kph'][0:16].values
settings_CO2= pd.to_numeric(df['MAW_CO2_gpkm'][0:16], errors='coerce') # df['MAW_CO2_gpkm'][0:16].values
settings_TOL1_minus_pct = pd.to_numeric(df['TOL1-'][0:16], errors='coerce') # df['TOL1-'][0:16].values
settings_TOL1_plus_pct = pd.to_numeric(df['TOL1+'][0:16], errors='coerce') # df['TOL1+'][0:16].values

vehspd_FTP3bags_kph = pd.to_numeric(df['vehspd_FTP3bags_kph'][0:3], errors='coerce') # df['vehspd_FTP3bags_kph'][0:3].values
FTP3_CO2_gpmi = pd.to_numeric(df['CO2_FTP3bags_gpmi'][0:3], errors='coerce') # df['CO2_FTP3bags_gpmi'][0:3].values
FTP3_NOx_gpmi = pd.to_numeric(df['NOx_FTP3bags_gpmi'][0:3], errors='coerce') # df['NOx_FTP3bags_gpmi'][0:3].values
FTP3_CO_gpmi = pd.to_numeric(df['CO_FTP3bags_gpmi'][0:3], errors='coerce') # df['CO_FTP3bags_gpmi'][0:3].values

vehspd_FTP3bags_comp_kph = float(df['vehspd_FTP3bags_kph'][3])
FTP3_CO2_comp_gpmi = float(df['CO2_FTP3bags_gpmi'][3])
FTP_NOx_comp_gpmi = float(df['NOx_FTP3bags_gpmi'][3])
FTP_CO_comp_gpmi = float(df['CO_FTP3bags_gpmi'][3])

vehspd_US06_kph = pd.to_numeric(df['vehspd_US06_kph'][0:2], errors='coerce') # .values
US06_CO2_gpmi = pd.to_numeric(df['CO2_US06_gpmi'][0:2], errors='coerce') # df['CO2_US06_gpmi'][0:2].values
US06_NOx_gpmi = pd.to_numeric(df['NOx_US06_gpmi'][0:2], errors='coerce') # df['NOx_US06_gpmi'][0:2].values
US06_CO_gpmi = pd.to_numeric(df['CO_US06_gpmi'][0:2], errors='coerce') # df['CO_US06_gpmi'][0:2].values

vehspd_US06_comp_kph = float(df['vehspd_US06_kph'][2])
US06_CO2_comp_gpmi = float(df['CO2_US06_gpmi'][2])
US06_NOx_comp_gpmi = float(df['NOx_US06_gpmi'][2])
US06_CO_comp_gpmi = float(df['CO_US06_gpmi'][2])

WLTC_NOx_gpmi = pd.to_numeric(df['NOx_WLTC_gpmi'][0:4], errors='coerce') # .values
WLTC_CO_gpmi = pd.to_numeric(df['CO_WLTC_gpmi'][0:4], errors='coerce') # .values
WLTC_NOx_comp_gpmi = float(df['NOx_WLTC_gpmi'][4])
WLTC_CO_comp_gpmi = float(df['CO_WLTC_gpmi'][4])

vehspd_HWFET_kph = pd.to_numeric(df['vehspd_HWFET_kph'][0:1], errors='coerce') # .values
HwFET_CO2_gpmi = pd.to_numeric(df['CO2_HwFET_gpmi'][0:1], errors='coerce') # .values
HwFET_NOx_gpmi = pd.to_numeric(df['NOx_HwFET_gpmi'][0:1], errors='coerce') # .values
HwFET_CO_gpmi = pd.to_numeric(df['CO_HwFET_gpmi'][0:1], errors='coerce') # .values

if (Exhaust_Emissions_Standard == 'T2Bin4') or (Exhaust_Emissions_Standard == 'T2BIN4'): 
    NOx_STD_mgpmi = T2Bin4_NOx*1000
    CO_STD_gpmi = T2Bin4_CO
    NMHC_STD_mgpmi = T2Bin4_NMOG*1000
elif (Exhaust_Emissions_Standard =='T2Bin5') or (Exhaust_Emissions_Standard =='T2BIN5'): 
    NOx_STD_mgpmi = T2Bin5_NOx*1000
    CO_STD_gpmi = T2Bin5_CO
    NMHC_STD_mgpmi = T2Bin5_NMOG*1000
elif (Exhaust_Emissions_Standard =='T3Bin70') or (Exhaust_Emissions_Standard =='T3BIN70'): 
    NOx_STD_mgpmi = T3Bin70_NOx*1000
    CO_STD_gpmi = T3Bin70_CO
    NMHC_STD_mgpmi = T3Bin70_NMOG*1000
elif (Exhaust_Emissions_Standard =='T3Bin85') or (Exhaust_Emissions_Standard =='T3BIN85'): 
    NOx_STD_mgpmi = T3Bin85_NOx*1000
    CO_STD_gpmi = T3Bin85_CO
    NMHC_STD_mgpmi = T3Bin85_NMOG*1000
elif (Exhaust_Emissions_Standard =='T3Bin125') or (Exhaust_Emissions_Standard =='T3BIN125'): 
    NOx_STD_mgpmi = T3Bin125_NOx*1000
    CO_STD_gpmi = T3Bin125_CO
    NMHC_STD_mgpmi = T3Bin125_NMOG*1000
elif (Exhaust_Emissions_Standard == 'T3Bin160') or (Exhaust_Emissions_Standard == 'T3BIN160'): 
    NOx_STD_mgpmi = T3Bin160_NOx*1000
    CO_STD_gpmi = T3Bin160_CO
    NMHC_STD_mgpmi = T3Bin160_NMOG*1000
elif (Exhaust_Emissions_Standard == 'EURO6D_SI') or (Exhaust_Emissions_Standard =='EURO6D_SI'): 
    NOx_STD_mgpmi = EURO6D_SI_NOx*1000
    NMHC_STD_mgpmi = EURO6D_SI_NMHC*1000
    CO_STD_gpmi = EURO6D_SI_CO
elif (Exhaust_Emissions_Standard == 'EURO6D_CI') or (Exhaust_Emissions_Standard =='EURO6D_CI'): 
    NOx_STD_mgpmi = EURO6D_CI_NOx*1000
    NMHC_STD_mgpmi = EURO6D_CI_NMHC*1000
    CO_STD_gpmi = EURO6D_CI_CO
    
if ('engtrq_RPM' in df.columns) and ('engtrqWOT_NM' in df.columns):
    engtrq_RPMs = df.loc[df['engtrq_RPM'] != '', 'engtrq_RPM'].astype(float)
    engtrqWOT_NMs = df.loc[df['engtrqWOT_NM'] != '', 'engtrqWOT_NM'].astype(float)

if FEV_to_EPA_PEMS: 
    mapping_sheet = 'FEV_to_EPA_PEMS_map'
    df_mapping = pd.read_excel(mapping_sheet + '.xlsx', sheet_name=mapping_sheet)

    # Create dictionaries for mapping
    fev_to_epa_var = dict(zip(df_mapping['FEV_VAR'], df_mapping['EPA_VAR']))
    epa_units = dict(zip(df_mapping['EPA_VAR'], df_mapping['EPA_UNITS']))
    fev_units = dict(zip(df_mapping['FEV_VAR'], df_mapping['FEV_UNITS']))
    # del df_mapping
    # print(fev_to_epa_var.keys())
    # print(fev_to_epa_var.values())
    # print(fev_to_epa_var.get('TMB_REC'))
    
    # print(fev_to_epa_units.keys())
    # print(fev_to_epa_units.values())
    # print(fev_to_epa_units.get('s'))
    
    # print(fev_to_epa_var)
    # print(fev_to_epa_units)

df_setting = df
del xls

GHG_WLTC_threshold = (settings_MAw_CO2_pct/100 + 0.00001)*np.sum(WLTC_dist_km * WLTC_CO2_gpkm)
# GHG_WLTC_threshold1 = (settings_MAw_CO2_pct/100 + 0.00001)*(WLTC_dist_comp_km  *WLTC_CO2_comp_gpkm)
# GHG_WLTC_threshold = min(GHG_WLTC_threshold, GHG_WLTC_threshold1)

script_inputs_file = ''
infile_folder = save_folder('Select PEMS and OBD input file folder')

suffix = ".html"

now=datetime.datetime.now()
hhmmss=datetime.time(now.hour, now.minute, now.second) #'{:02d}:{:02d}:{:02d}'.format(now.hour, now.minute, now.second)
hms = str(hhmmss).split(':')
now_hms = str(hms[0]) + str(hms[1]) + str(hms[2])

odir = os.getcwd()    
os.chdir(infile_folder)

elapsed = time.time() - t
print("\n*************\nElapsed Time: {:5.0f} Seconds after selecting the template and file folder" .format(elapsed))

extension = 'csv'
pp_filenames = [i for i in glob.glob('pp_*.{}'.format(extension))]
csv_filenames = [i for i in glob.glob('*.{}'.format(extension))]
nel_df1 = 0
vehicle_title = ''

pp_files = pd.DataFrame(pp_filenames, columns=['fname']).sort_values('fname')
pp_files = pp_files.reset_index(drop = True)

if FEV_to_EPA_PEMS: 
    [df, EPA_PEMS_columns_units, vehicle_title, pp_file, df1, df2] = read_FEV_PEMS(df, pp_files, fev_to_epa_var, epa_units, fev_units)
else:
    [df, EPA_PEMS_columns_units, vehicle_title, pp_file, df1, df2] = read_pp_files(df, pp_files)

# if (("sDATE" in df.columns)) and (sDATE_cname == ''): sDATE_cname = "sDATE"
sTIME_cname = "sTIME" if ("sTIME" in df.columns) else ''
sDATE_cname = "sDATE" if ("sDATE" in df.columns) else ''
sSTATUS_PATH_cname = "sSTATUS_PATH" if ("sSTATUS_PATH" in df.columns) else ''
iGPS_LAT_cname = "iGPS_LAT" if ("iGPS_LAT" in df.columns) else ''
iGPS_LON_cname = "iGPS_LON" if ("iGPS_LAT" in df.columns) else ''
iGPS_ALT_cname = "iGPS_ALT" if ("iGPS_ALT" in df.columns) else ''
iGPS_GROUND_SPEED_cname = "iGPS_GROUND_SPEED" if ("iGPS_GROUND_SPEED" in df.columns) else ''

iVEH_SPEED_cname = "iVEH_SPEED" if ("iVEH_SPEED" in df.columns) else ''
imVEH_SPEED_cname = "imVEH_SPEED" if ("imVEH_SPEED" in df.columns) else ''
iENG_SPEED_cname = "iENG_SPEED" if ("iENG_SPEED" in df.columns) else ''
iENG_LOAD_cname = "iENG_LOAD" if ("iENG_LOAD" in df.columns) else ''
iENG_TORQ_cname = "iENG_TORQ" if ("iENG_TORQ" in df.columns) else ''
iENG_HP_cname = "iENG_HP" if ("iENG_HP" in df.columns) else ''

iWf_cname = "iWf" if ("iWf" in df.columns) else ''
iWfgps_cname = "iWfgps" if ("iWfgps" in df.columns) else ''
ECT_cname = "ECT" if ("ECT" in df.columns) else ''
iCOOL_TEMP_cname = "iCOOL_TEMP" if ("iCOOL_TEMP" in df.columns) else ECT_cname
CatalystTemp_cname = "CATEMP11" if ("CATEMP11" in df.columns) else "iCATEMP11"
iFLOW_EX_TEMP_cname = "iFLOW_EX_TEMP" if ("iFLOW_EX_TEMP" in df.columns) else ''
iHUM_TEMP_cname = "iHUM_TEMP" if ("iHUM_TEMP" in df.columns) else ''
Tamb_cname = "iSCB_LAT" if ("iSCB_LAT" in df.columns) else ''

iCALCRT_CO2m_cname = "iCALCRT_CO2m" if ("iCALCRT_CO2m" in df.columns) else ''
iCALCRT_COm_cname = "iCALCRT_COm" if ("iCALCRT_COm" in df.columns) else ''
iCALCRT_NOxm_cname = "iCALCRT_NOxm" if ("iCALCRT_NOxm" in df.columns) else ''
iCALCRT_kNOxm_cname = "iCALCRT_kNOxm" if ("iCALCRT_kNOxm" in df.columns) else ''
iCALCRT_NMHC_cname = "iCALCRT_NMHCm" if ("iCALCRT_NMHCm" in df.columns) else ''
iSCB_LAP_cname = "iSCB_LAP" if ("iSCB_LAP" in df.columns) else ''
Lambda_cname = "Lambda" if ("Lambda" in df.columns) else ''

# df = df[list(EPA_PEMS_columns_units.keys())]
# df.iloc[:, 3:] = df.iloc[:, 3:].apply(pd.to_numeric, errors='coerce') # except date, time and sSTATUS_PATH
# df = df.drop(columns=df.filter(regex='Unnamed:').columns)

# cols = df.columns
# df = df.replace(np.nan, '', regex=True)
# rows = df.shape[0]
# cols = df.shape[1]
# if cols > 185:
#     max_cols = 185
# else:
#     max_cols = cols
# for icol in range(max_cols-1):
#     cname = df.columns[icol]
#     # if cname == 'sDATE': sDATE_cname = cname
#     # elif cname == 'sSTATUS_PATH': sSTATUS_PATH_cname = cname
#     # elif cname == 'iGPS_ALT' : iGPS_ALT_cname = cname
#     # elif cname == ('iCALCRT_CO2m' or 'iCALCRT_CO2m'): iCALCRT_CO2m_cname = cname
#     # elif cname == 'iCALCRT_COm': iCALCRT_COm_cname = cname
#     # elif cname == 'iCALCRT_NOxm': iCALCRT_NOxm_cname = cname # instantenous NOx
#     # elif cname == 'iCALCRT_kNOxm': iCALCRT_kNOxm_cname = cname # use the corrected instantenous NOx
#     # elif cname == 'iSCB_LAP': iSCB_LAP_cname = cname # local ambient pressure
#     # elif cname == 'Lambda': Lambda_cname = cname # local ambient temperature
#     elif cname[0:8] == 'Unnamed:':
#         df = df.drop(columns=cname)
#         cols -= 1
#         iUnnamed = 1
        
# print("Total number of data points before cleaning: ", len(df))
df = df.reset_index(drop=True)
print("Total number of data points after dropna(all): ", len(df))

# df = df.rename(columns={'iBhp': 'iENG_HP'})
if (eTIME_DATA_PLOT_flag) and (FEV_to_EPA_PEMS): df = df.loc[df['sTIME'] <= eTIME_DATA_PLOT, :]
if iENG_HP_cname != '': engpwr_HP =  df['iENG_HP']
engine_RPM = df['iENG_SPEED'].astype(float)
        
if iENG_TORQ_cname != '':   
    if EPA_PEMS_columns_units.get('iENG_TORQ') == 'Nm': 
        engine_trq_lbft = 0.7375621493 * df['iENG_TORQ'].astype(float)
    else:
        engine_trq_lbft = df['iENG_TORQ'].astype(float)   
    engpwr_HP = (engine_RPM * engine_trq_lbft)/5252
elif (not FEV_to_EPA_PEMS) and (iENG_LOAD_cname != '') and (len(engtrq_RPMs) > 0) and (len(engtrqWOT_NMs) > 0):
    engine_WOT_trq = engine_torques(engine_RPM, engtrq_RPMs, engtrqWOT_NMs)
    engine_trq = engine_WOT_trq * df[iENG_LOAD_cname].astype(float)/100
    engine_trq_lbft = 0.7375621493 * engine_trq
    df['iENG_TORQ'] = engine_trq_lbft if (EPA_PEMS_columns_units.get('iENG_TORQ') == 'lb-ft') or (iENG_TORQ_cname == '') else engine_trq
    engpwr_HP = df['iENG_HP'] = (engine_RPM * engine_trq_lbft)/5252

if (iVEH_SPEED_cname == '') and (imVEH_SPEED_cname != ''):
    iVEH_SPEED_cname = "iVEH_SPEED"
    df[iVEH_SPEED_cname] = 0.621371 * df[imVEH_SPEED_cname]

nel = len(df)
if (FEV_to_EPA_PEMS):
    etime = df.sTIME.astype(float)
else:
    if EPA_PEMS_columns_units.get(sTIME_cname) != 'hhmmss': 
        for i in range (nel):
            df.sTIME[i] = time.strftime('%H:%M:%S', time.gmtime(int(df.sTIME[i])))
    [last_h, last_m, last_s] =str(df.sTIME[len(df.sTIME)-1]).split(':')
    last_s = float(last_s.split('.')[0])

    time0 = timePEMS_to_second(str(df.sTIME[0]))
    time1 = timePEMS_to_second(str(df.sTIME[1]))
    timen = timePEMS_to_second(str(df.sTIME[nel-1]))
    dt = (time1 - time0) # (timen - time0)/(len(df))
    etime = np.arange(0, dt*nel, dt)
df['etime'] = etime

# Create figure folder if it doesn't exist
fig_folder = os.path.join(infile_folder, 'figures')
if not os.path.exists(fig_folder):
    os.makedirs(fig_folder)
    
fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6)) 
# min_val = min(len(engpwr_HP_calc), len(engpwr_HP))
ax.plot(etime, engpwr_HP, 'r-')
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Measured Engine Power (bhp)')
ax.set_title('FEV to EPA PEMS Engine Power Comparison')
plt.grid(True)
# plt.legend()
plt.tight_layout()

fig_path = os.path.join(fig_folder, 'engine-hp.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6)) 
ax.plot(etime, df['iENG_SPEED'], 'b-')
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Engine Speed (rpm)')
plt.grid(True)
# plt.legend()
plt.tight_layout()
fig_path = os.path.join(fig_folder, 'engine-rpm.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6))
ax.plot(df['iENG_SPEED'], df['iENG_TORQ'],     
    marker='o',             # Use circle markers
    markersize=5,          # Set the size of the markers (e.g., 15 points)
    markerfacecolor='None', # Make the marker interior empty/transparent
    markeredgecolor='r',    # Set the marker outline color (e.g., blue)
    markeredgewidth=0.5,    # Set the thickness of the marker outline
    linestyle='None'        # No connecting lines between markers
    )

ax.set_xlabel('Engine Speed (rpm)')
ax.set_ylabel('Engine Torque (lb-ft)')
plt.grid(True)
# plt.legend()
plt.tight_layout()

fig_path = os.path.join(fig_folder, 'engine-rpm-torque.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6)) 
ax.plot(etime, df['iVEH_SPEED'], 'b-') 
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Vehicle Speed (mph)')
plt.grid(True)
# plt.legend()
plt.tight_layout()
fig_path = os.path.join(fig_folder, 'vehspd_mph.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6))
ax.plot(etime, df['imVEH_SPEED'], 'b-') 
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Vehicle Speed (kph)')
plt.grid(True)
# plt.legend()
plt.tight_layout()
fig_path = os.path.join(fig_folder, 'vehspd_kph.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6))
ax.plot(etime, df['iENG_SPEED'], 'b-')
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Engine Speed (rpm)')
plt.grid(True)
# plt.legend()
plt.tight_layout()
fig_path = os.path.join(fig_folder, 'engine-rpm.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

plt.close('all')

nel = len(df)

if iVEH_SPEED_cname == '' and iGPS_GROUND_SPEED_cname != '': df['iVEH_SPEED'] = df[iGPS_GROUND_SPEED_cname]
if iVEH_SPEED_cname != '': iGPS_GROUND_SPEED_mode = 0

if (iVEH_SPEED_cname != '') and (EPA_PEMS_columns_units.get(iVEH_SPEED_cname) == "kph"): 
    df['vehspd_ecu'] = df.iVEH_SPEED.astype(float)
elif ("imVEH_SPEED" in df.columns) and (EPA_PEMS_columns_units.get(imVEH_SPEED_cname) == "kph"): 
    df['vehspd_ecu'] = df.imVEH_SPEED.astype(float)
elif (iVEH_SPEED_cname == '') and (iGPS_GROUND_SPEED_cname != ''):
    df['vehspd_ecu'] = df.iGPS_GROUND_SPEED.astype(float) if EPA_PEMS_columns_units.get(iGPS_GROUND_SPEED_cname) == "kph" else 1.60934 * df.iGPS_GROUND_SPEED.astype(float)
    iGPS_GROUND_SPEED_mode = 1

df['eng_Combustion_ON'] = eng_Combustion_ON = np.where(df[iENG_SPEED_cname].astype(float) >= 250, 1, 0)
fig = plt.figure(facecolor=(1, 1, 1))
plt.plot(df[iENG_SPEED_cname].astype(float))
fig = plt.figure(facecolor=(1, 1, 1))
plt.plot(df['eng_Combustion_ON'])

if (not FEV_to_EPA_PEMS):    
    vehspd_move_start = df.loc[(df['vehspd_ecu'].shift(2) < 0.01) & (df['vehspd_ecu'].shift(1) < 0.1) & (df['vehspd_ecu'] > 0.5) &  (df['vehspd_ecu'].shift(-15)  > 0) & (df['vehspd_ecu'].shift(-30)  <= vehspd_urban_kph)].index[0] - 10
    _idx_engine_on = df.loc[(df['eng_Combustion_ON'].shift(5) < 0.5) & (df['eng_Combustion_ON'] > 0.5)].index
    if len(_idx_engine_on) > 0:
        if _idx_engine_on[0] > 60: 
            vehspd_move_start = max(0, min(vehspd_move_start, _idx_engine_on[0]-60))
        elif _idx_engine_on[0] > 30: 
            vehspd_move_start = max(0, min(vehspd_move_start, _idx_engine_on[0]-30))
        else:
            vehspd_move_start = max(0, min(vehspd_move_start, _idx_engine_on[0]-15))
            
    if vehspd_move_start > 0: 
        df = df.drop(df.index[:vehspd_move_start])
        df = df.reset_index(drop = True)
    # print('len(df) = at line 1364', len(df))

    US_max_vehspd_mph_cutoff = 1.60934 * US_max_vehspd_kph
    vehspd_move_stop = df.loc[(df['vehspd_ecu'].shift(10) <= 0.01) & (df['vehspd_ecu'].shift(1) <= 0.1) & (df['vehspd_ecu'] > 0.5) &  (df['vehspd_ecu'].shift(-30)  > US_max_vehspd_mph_cutoff)].index
    if (len(vehspd_move_stop) == 1):
        vehspd_move_stop = vehspd_move_stop[0]-30 if len(vehspd_move_stop) > 0 and max(df.loc[vehspd_move_stop[0]-30: vehspd_move_stop[0]-1, 'vehspd_ecu']) <= 0.1 else (vehspd_move_stop[0]-1 if len(vehspd_move_stop) > 0 else 0)  

    if (len(vehspd_move_stop) == 1) and (vehspd_move_stop >= 0.95*len(df)):
        _idx1 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index >= max(0, vehspd_move_stop - 50)) & (df.index <= vehspd_move_stop)].index
        df = df.drop(df.index[_idx1[-1]:])
    elif (len(vehspd_move_stop) > 0):
        for k, idx in enumerate(vehspd_move_stop):
            if (k == len(vehspd_move_stop)-1) and (idx <= len(df)): # and (len(vehspd_145kph_stop) < len(vehspd_move_stop)):
                _idx1_145kph = df.loc[(df['vehspd_ecu'] >= US_max_vehspd_mph_cutoff)].index
                if len(_idx1_145kph) > 0:
                    _idx1 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index >= max(0, _idx1_145kph[0] - 50)) & (df.index <= _idx1_145kph[0])].index
                else:
                    _idx1 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index >= max(0, idx - 50)) & (df.index <= idx)].index
                df = df.drop(df.index[_idx1[-1]:])
            else:
                if (idx >= 0.95*len(df)) and (idx < len(df)):
                    _idx1_145kph = df.loc[(df['vehspd_ecu'] > 1.05* US_max_vehspd_kph)& (df.index >= max(0, idx - 100)) & (df.index <= idx) ].index
                    if len(_idx1_145kph) > 0:
                        _idx1 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index >= max(0, _idx1_145kph[0] - 50)) & (df.index <= _idx1_145kph[0])].index
                    else:
                        _idx1 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index >= max(0, idx - 50)) & (df.index <= idx)].index
                    df = df.drop(df.index[_idx1[-1]:]) if len(_idx1) <= 30 else df.drop(df.index[_idx1[0]:])
                    df = df.reset_index(drop = True)
                elif (idx <= len(df)):
                    _idx1 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index >= max(0, idx - 50)) & (df.index <= idx)].index
                    _idx2_145kph = df.loc[(df['vehspd_ecu'] >= US_max_vehspd_mph_cutoff) & (df.index >= idx - 50) & (df.index <= idx + 100)].index
                    _idx2 = df.loc[(df['vehspd_ecu'] <= 0.01) & (df.index > _idx2_145kph[-1]) & (df.index <= idx+100)].index
                    if (len(_idx1) > 0) and (len(_idx2) > 0):
                        df = df.drop(df.index[_idx1[-1]:_idx2[0]])
                    else:
                        df = df.drop(df.index[max(0,idx-10):idx+50])
                    df = df.reset_index(drop = True)
   
    df = df.reset_index(drop = True)

if (iGPS_GROUND_SPEED_cname != ''): 
    if EPA_PEMS_columns_units.get(iGPS_GROUND_SPEED_cname) == 'mph': df['iGPS_GROUND_SPEED_kph'] = 1.60934 * df.iGPS_GROUND_SPEED.astype(float)
    else: df.iGPS_GROUND_SPEED = df.iGPS_GROUND_SPEED.astype(float) 

nel = len(df)
# print('len(df) = at line 1522', len(df))

sample_start, sample_end = 0, -1
if sample_end == -1: sample_end = len(df)    
if script_inputs_file == '': script_inputs_file = pp_file
    
pp_files = pd.DataFrame(pp_filenames, columns=['fname']).sort_values('fname')
pp_files = pp_files.reset_index(drop = True)
pp_file1 = pp_files.fname[0]
pp_filen = pp_files.fname[len(pp_files.fname)-1]
pp_filem = max(pp_filenames, key=len)
if len(pp_filem) > len(pp_filen): pp_filen = pp_filem
        
troute = pp_file1.split('.')[0]
route1 = (troute.split('-')[0]).split('_')
route1 = route1[len(route1)-1]

troute = pp_filen.split('.')[0]
route2 = troute.split('-')[1]
route = route1 + '-' + route2

if not FEV_to_EPA_PEMS:
    testdate =''
    if df.sDATE[0] != '':
        testdate = str(df.sDATE[0]) #'03/29/2018'
    else:
        testdate = str(df.sDATE[1]) #'03/29/2018'
    vehicle_title = vehicle_title + route + ' routes @ ' + testdate
else:
    vehicle_title = (pp_file.split('.')[0]).split('pp_')[1] if 'pp_' in pp_file else pp_file.split('.')[0]

SAE_vehicle_title = ''
vehicle_title = SAE_vehicle_title

filename_combined = pp_file1.split('-')[0] + '-' + route2
if OBD == True: 
	filename_combined_obd = 'OBD_' + pp_file1.split('-')[0] + '-' + route2 + '.xlsx'
else:
	filename_combined_obd = ''

del troute, route1, route2, pp_file1, pp_filen

if (OBD):
    if iCALCRT_CO2m_cname != '':
        for i in range (sample_start, nel-3):
            if df[iCALCRT_CO2m_cname][i] != '' and (df['vehspd_ecu'][i] < US_max_vehspd_mph and df['vehspd_ecu'][i+1] > 0 and df['vehspd_ecu'][i+2] > 0 and df['vehspd_ecu'][i+3] > 0) and \
            (df['vehspd_ecu'][i+1] < US_max_vehspd_mph and df['vehspd_ecu'][i+2] < US_max_vehspd_mph and df['vehspd_ecu'][i+3] < US_max_vehspd_mph):
                sample_start = i
                break

    if np.max(df['vehspd_ecu']) > EU_peak_vehspd_kph:
        for i in range(sample_end-2, sample_start+1, -1):
            if df['vehspd_ecu'][i] > EU_peak_vehspd_kph and  df['vehspd_ecu'][i-15] <10 and df['vehspd_ecu'][i-16] < 0.5:
                for j in range (500, 0, -1):
                    if df['vehspd_ecu'][i-j] == 0 and df['vehspd_ecu'][i-j-1] > 0.5 and  df['vehspd_ecu'][i-j-2] > 1.5:
                        sample_end = i-j+2
                        break

    if sample_start > 15: nel = sample_end - sample_start + 1
    df = df[sample_start:sample_end, :]

    df = df.reset_index(drop = True)
    nel = sample_end = len(df)

# VMT = np.zeros(nel)
vehspd_flt = np.zeros(nel)
eVeh_accelD = np.zeros(nel)
eVeh_accelE = np.zeros(nel)
eVeh_accelP = np.zeros(nel)

etime = df['etime'] - df['etime'][0]
Inst_Mass_GHG = df[iCALCRT_CO2m_cname].astype(float) if iCALCRT_CO2m_cname != '' else np.zeros(nel)
Cumul_Mass_GHG = df[iCALCRT_CO2m_cname].cumsum() if iCALCRT_CO2m_cname != '' else np.zeros(nel)
engine_RPM = df[iENG_SPEED_cname].astype(float) if iENG_LOAD_cname != '' else np.zeros(nel)
engine_load = df[iENG_LOAD_cname].astype(float) if iENG_LOAD_cname != '' else np.zeros(nel)
ECT = df[ECT_cname].astype(float) if ECT_cname != '' else np.zeros(nel)

Lambda = df[Lambda_cname].astype(float) if Lambda_cname != '' else np.zeros(nel)
CatalystTemp = Texh = df[CatalystTemp_cname] if (CatalystTemp_cname != '') and (EPA_PEMS_columns_units.get(CatalystTemp_cname) == 'deg C') else 5/9 *  (df[CatalystTemp_cname] - 32)
Tamb = df[Tamb_cname].astype(float) if Tamb_cname != '' else np.zeros(nel)

if (iWfgps_cname != ''): fuelflow_gals = df[iWfgps_cname].astype(float) / (1 if EPA_PEMS_columns_units.get(iWfgps_cname) == 'gal/s' else 2834.89)

Inst_Mass_CO = df[iCALCRT_COm_cname].astype(float) if iCALCRT_COm_cname != '' else np.zeros(nel)
Inst_Mass_NOx = df[iCALCRT_NOxm_cname].astype(float) if iCALCRT_NOxm_cname != '' else np.zeros(nel)
Inst_Mass_NMHC = df[iCALCRT_NMHC_cname].astype(float) if iCALCRT_NMHC_cname != '' else np.zeros(nel)
Corrected_Inst_Mass_NOx = df[iCALCRT_kNOxm_cname].astype(float) if iCALCRT_kNOxm_cname != '' else np.zeros(nel)
if (iENG_SPEED_cname != '') and (iGPS_ALT_cname != ''):
    df.loc[(df[iENG_SPEED_cname].astype(float) <= 0.5) & df['imVEH_SPEED'].astype(float) < 0.2, iGPS_ALT_cname] = df[iGPS_ALT_cname].shift(1)

if (iGPS_GROUND_SPEED_cname != ''):
    df.loc[(df[iGPS_GROUND_SPEED_cname].astype(float) < 50) & df['imVEH_SPEED'].astype(float) > 10, 'vehspd_ecu'] = 0
else:
    iGPS_GROUND_SPEED_mode = 0

if vehicle_type == 1 and iENG_SPEED_cname != '' and OBD == FALSE and iGPS_GROUND_SPEED_mode == 1:
    df.loc[(df[iENG_SPEED_cname].astype(float) < 50) & df['vehspd_ecu'].astype(float) > 1, 'vehspd_ecu'] = 0

df.loc[(df['vehspd_ecu'].astype(float) < vehspd_1_kph) & (df['vehspd_ecu'].shift(1).astype(float) < vehspd_1_kph) & (df['vehspd_ecu'].shift(-1).astype(float) < vehspd_1_kph) & (df['vehspd_ecu'].shift(-2).astype(float) < vehspd_1_kph), 'vehspd_ecu'] = 0
df.loc[(df['vehspd_ecu'] < 0.01) & (df[iCALCRT_CO2m_cname]  > 0.5) & (df['vehspd_ecu'].shift(1) > 0.5) & (df['vehspd_ecu'].shift(-1) > 0.5), 'vehspd_ecu'] = 0.5 * (df['vehspd_ecu'].shift(1) + df['vehspd_ecu'].shift(-1))
        
# for j in range (len(etime)-1):
# 	if (etime[j+1] - etime[j]) > 120: 
# 		if iENG_SPEED_cname != '' and (etime[j+1] - etime[j]) > 300 and engine_RPM[j] > 0 : engine_RPM[j] = 0;
# 		if iENG_LOAD_cname != '' and (etime[j+1] - etime[j]) > 300 and engine_load[j+1] >= 0 : engine_load[j+1] = 0; engine_load[j] = 0;
# 		if iCOOL_TEMP_cname != '' and (etime[j+1] - etime[j]) > 300 and ECT[j+1] > 10 : ECT[j+1] = 0; ECT[j] = 0;
# 		if iCOOL_TEMP_cname != '' and (etime[j+1] - etime[j]) > 300 and ECT[j] > 10 and ECT[j+1] < 10 : ECT[j+1] = 0; ECT[j] = 0;
# 		if iENG_SPEED_cname != '' and (etime[j+1] - etime[j]) > 120 and engine_RPM[j+1] == 0  and engine_RPM[j] > 0: engine_RPM[j] = 0;
# 		if iENG_SPEED_cname != '' and (etime[j+1] - etime[j]) > 120 and engine_RPM[j] == 0  and engine_RPM[j+1] > 0: engine_RPM[j+1] = 0;

df = df.reset_index(drop = True)
nel = len(df)
df.to_excel( filename_combined + '.xlsx', index=False) 

Altitude = raw_Altitude = df[iGPS_ALT_cname].astype(float) if (iGPS_ALT_cname != '') else np.zeros(nel)
if (max(Altitude) > 0.01): 
    Altitude = savgol_filter(Altitude, window_length=61, polyorder=3)
else:
    Altitude = raw_Altitude = np.zeros(nel)
dt = etime[1] -  etime[0]
dt2 = etime[2] -  etime[0]

df['dVMT'] = (dt/3.6) * df['vehspd_ecu']
df['VMT'] = df['dVMT'].cumsum()

veh_ACC = np.nan_to_num(np.where(df['vehspd_ecu'].notna(), (df['vehspd_ecu'].shift(-1) - df['vehspd_ecu'].shift(1)) / (dt2 * 3.6), 0))
veh_accel = np.nan_to_num(np.where(df['vehspd_ecu'].notna(), (df['vehspd_ecu'].shift(-1) - df['vehspd_ecu'].shift(1)) / (dt2 * 3.6), 0))

RPA = np.nan_to_num(np.where(veh_accel > 0.1, veh_accel, 0.1))
if (iGPS_ALT_cname != ''):
    dAltitude = np.where(df['vehspd_ecu'] > MAW_vehspd_0_kph, df[iGPS_ALT_cname] - df[iGPS_ALT_cname].shift(1), df[iGPS_ALT_cname].shift(1))
else:
    dAltitude = Altitude = np.zeros(nel)

if OBD == TRUE:
    [h2, m2, s2] = str(df2.sTIME[df2.index[0]]).split(':')
    h2, m2 = float(h2), float(m2)
    s2 = float(s2.split('.')[0])
    
    dff_obd = pd.DataFrame(csv_filenames, columns = ['OBD'])
    dff_obd = dff_obd.sort_values('OBD')
    dff_obd = dff_obd.reset_index(drop=True)
    nobd = len (dff_obd)
    for i in range (nobd):
        if i <= len (dff_obd) :
            if (dff_obd.OBD[i][0:3] == 'TA_'):
                dff_obd = dff_obd.drop(dff_obd.index[i])
                dff_obd = dff_obd.reset_index(drop = True)
            elif (dff_obd.OBD[i][0:3] == 'pp_'):
                dff_obd = dff_obd.drop(dff_obd.index[i:len(dff_obd)])
                if dff_obd.OBD[i-1][0:3] == 'pp_': dff_obd = dff_obd.drop(dff_obd.index[i-1])
                dff_obd = dff_obd.reset_index(drop = True)

    for i in range (len(dff_obd)):
        obd_file = str(dff_obd.OBD[i])
        if i == 0:
            df_obd = pd.read_csv(obd_file, skiprows=0, encoding = 'unicode_escape') #encoding = "ISO-8859-1")
            df_obd = df_obd.replace(np.nan, '', regex=True)
            if i == 0: df1_obd = df_obd
            df1_obd_file = obd_file

            nel_obd1 = len(df_obd)
        else:
            dft = pd.read_csv(obd_file, skiprows=0, encoding = 'unicode_escape') #encoding = "ISO-8859-1")
            dft = dft.replace(np.nan, '', regex=True)            
            dft.index = range(len(df_obd), len(df_obd)  + len(dft))
            df_obd = pd.concat([df_obd, dft])
            if i == 1: df2_obd = dft 
            df2_obd_file = obd_file
    
    df_obd = df_obd.sort_values(['Clock_T(hms)'], ascending=[True])
    df_obd = df_obd.reset_index(drop=True)
        
    ovehspd = df_obd['SPD(km/h)'].astype(float)
    oengtrq = df_obd['TQRFIN(Nm)'].astype(float)
    oengrpm = df_obd['NE(rpm)'].astype(float)
    oengect = df_obd['THW(C)'].astype(float)
    oambairT= df_obd['THA(C)'].astype(float)
    oexhtmp = df_obd['THEXH2(C)'].astype(float)
    hms_obd = df_obd['Clock_T(hms)']

    [hh0, mm0, ss0, otime0] = timeOBD_to_second(str(hms_obd[0]))

    last_obd_found = 0
    nhms_obd = len(hms_obd)
    last_obd_index = nhms_obd
    otime = np.zeros(nhms_obd)
    for i in range (nhms_obd):
        [hh, mm, ss, otimei] = timeOBD_to_second(str(hms_obd[i]))
        otime[i] = otimei - otime0
        if (last_obd_found == 0) and float(hh) > float(last_h) and (float(mm) +30) >= float(last_m) and float(ss) >= float(last_s): 
            last_obd_index, last_obd_found = i, i
            break

    if nhms_obd > last_obd_index: df_obd = df_obd.drop(df_obd.index[last_obd_index:nhms_obd])

    vehspd1 = df['vehspd_ecu']
    etime1 = etime
    ovehspd1 = ovehspd
    otime1 = otime
            
    timeoffset_obd = 0
    timeoffset_pems = 0
    itime_start = 0
    ndoe = 500
    nsamples_obd_sync = 600
    [timeoffset_pems, timeoffset_obd] = vehspd_obd_sync(ndoe, nsamples_obd_sync, itime_start, etime, df['vehspd_ecu'], otime, ovehspd.values)

    timeoffset_pems0 = timeoffset_pems
    timeoffset_obd0 = timeoffset_obd
    
    vehspd2 = 1.60394*df2.iGPS_GROUND_SPEED.astype(float)        
    ovehspd2 = df2_obd['SPD(km/h)'].astype(float)

    etime10 = timePEMS_to_second(str(df1.sTIME[df1.index[0]]))
    etime11 = timePEMS_to_second(str(df1.sTIME[df1.index[len(df1)-1]]))
    etime20 = timePEMS_to_second(str(df2.sTIME[df2.index[0]]))
    etime21 = timePEMS_to_second(str(df2.sTIME[df2.index[len(df2)-1]]))
    delta_etime21 = etime20 - etime10

    [hh10, mm10, ss10, otime10] = timeOBD_to_second(str(df1_obd['Clock_T(hms)'][df1_obd.index[0]]))
    [hh11, mm11, ss11, otime11] = timeOBD_to_second(str(df1_obd['Clock_T(hms)'][df1_obd.index[len(df1_obd)-1]]))
    [hh20, mm20, ss20, otime20] = timeOBD_to_second(str(df2_obd['Clock_T(hms)'][df2_obd.index[0]]))
    [hh21, mm21, ss21, otime21] = timeOBD_to_second(str(df2_obd['Clock_T(hms)'][df2_obd.index[len(df2_obd)-1]]))
    delta_otime21 = otime20 - otime10
            
    etime = etime + timeoffset_pems
    otime = otime + timeoffset_obd
    
    vehspd2 = vehspd2.reset_index(drop=True)
    ovehspd2 = ovehspd2.reset_index(drop=True)

    iovehspd_GT0 = 0
    ivehspd_GT0 = 0
    for i in range (len(ovehspd2)):
        if ovehspd2[i] > 0 and iovehspd_GT0 == 0: iovehspd_GT0 = i
    for i in range (len(vehspd2)):
        if vehspd2[i] > 0 and ivehspd_GT0 == 0: ivehspd_GT0 = i
    
    ivehspd_ecu = 0        
    for i in range (nel_df1-500, nel_df1+1000):
        if ivehspd_ecu == 0 and abs(df['vehspd_ecu'][i])> 0.5 and abs(df['vehspd_ecu'][i+1])  > 0.5 and abs(df['vehspd_ecu'][i+2])  > 0.5:
            ivehspd_ecu = i
            
    ovehspd = df_obd['SPD(km/h)'].astype(float)
    oengtrq = df_obd['TQRFIN(Nm)'].astype(float)
    oengrpm = df_obd['NE(rpm)'].astype(float)
    oengect = df_obd['THW(C)'].astype(float)
    if max(oengect) > 300: oengect-= 273.5
    oambairT= df_obd['THA(C)'].astype(float)
    oexhtmp = df_obd['THEXH2(C)'].astype(float)
    hms_obd = df_obd['Clock_T(hms)'].astype(float)
    otime = otime[0:len(df_obd)]

    for j in range (len(otime)-1):
    	if (otime[j+1] - otime[j]) > 120: 
    		if oengtrq[j+1] > 0: oengtrq[j+1] = 0
    		if (otime[j+1] - otime[j]) > 300 and oengrpm[j] > 0 : oengrpm[j] = 0
    		if (otime[j+1] - otime[j]) > 300 and oengect[j+1] > 10 : oengect[j+1] = 0; oengect[j] = 0
    		if (otime[j+1] - otime[j]) > 120 and oengrpm[j+1] == 0  and oengrpm[j] > 0: oengrpm[j] = 0
    		if (otime[j+1] - otime[j]) > 120 and oengrpm[j] == 0  and oengrpm[j+1] > 0: oengrpm[j+1] = 0

    df_obd['OBD_time(s)'] = otime
    df_obd['OBD_engrpm'] = oengrpm
    df_obd['OBD_engtrq_NM'] = oengtrq
    df_obd['OBD_ECT_C'] = oengect
    df_obd.to_excel(filename_combined_obd, index=False, encoding='utf-8-sig')

###################

elapsed = time.time() - t
print("\n*************\nElapsed Time: {:5.0f} Seconds before plotting and saving to PDF" .format(elapsed))

here = os.path.dirname(os.path.realpath("__file__"))

vehspd_ecu = df['vehspd_ecu']

vehmove_start_ini = 0
time_engine_start = 0
time_engine_idle_int = 0
time_cold_start = 0
_index_ECT_cold_start_end = 0
ECT_cold_start_low = 95 if EPA_PEMS_columns_units.get(ECT_cname) == "deg F" else 35
ECT_cold_start_end = 158 if EPA_PEMS_columns_units.get(ECT_cname) == "deg F" else 70
ECT_diff_cold_start = 44.6 if EPA_PEMS_columns_units.get(ECT_cname) == "deg F" else 7
cold_start= 1  
cold_start_end=0

time_engine_idle_int = 0
time_cold_start = 0
_index_time_engine_start = 0
_index_ECT_cold_start_end = 0

_index_ECT_cold_start_end = (df[ECT_cname] - ECT_cold_start_end).abs().idxmin()
ECT_at_cold_start = df.loc[_index_ECT_cold_start_end, ECT_cname]
time_cold_start_end = etime[_index_ECT_cold_start_end]

if (_index_ECT_cold_start_end == len(df[ECT_cname]) -1) or (iENG_SPEED_cname == ''):
    _index_time_engine_start = 1
    time_engine_idle_int = 0
    cold_start_end = 1
    _index_ECT_cold_start_end = int(300/dt)

if (EPA_PEMS_columns_units.get(Tamb_cname) != 'deg C'): Tamb = 5/9*(Tamb - 32)
_idx_engine_start = df.loc[(df['eng_Combustion_ON'].shift(-1) == 1) & (df['eng_Combustion_ON'] == 0)].index
if len(_idx_engine_start) > 0:
    time_engine_start = etime[_idx_engine_start[0]]
ECT_engine_start = df.loc[(df['eng_Combustion_ON'].shift(-1) == 1) & (df['eng_Combustion_ON'] == 0), ECT_cname].values[0]
_index_time_engine_start = df.loc[(df['eng_Combustion_ON'].shift(-1) == 1) & (df['eng_Combustion_ON'] == 0), 'etime'].index[0]
if (ECT_engine_start <= engine_cold_start_ECT): time_cold_start = etime[_index_time_engine_start]
# if (ECT_engine_start <= engine_cold_start_ECT): time_cold_start = df.sTIME[_index_time_engine_start]

df.loc[(df[ECT_cname] <= engine_cold_start_ECT), 'cold_start'] = 1
df.loc[((df[ECT_cname] - Tamb) <= 7.0) & (df[ECT_cname] <= 35) & (vehspd_ecu > 0.01) & (df["cold_start"] == 1), 'cold_start'] = 0
vehmove_start = df.loc[(vehspd_ecu.shift(-1) >= 0.5) & (vehspd_ecu < 0.1), 'etime']
vehmove_start_ini = vehmove_start.index[0] if len(vehmove_start) > 0 else 0

_index_time_engine_idle = df.loc[(vehspd_ecu > 0.5) & (vehspd_ecu.shift(-1) <= 0.1) & (df['eng_Combustion_ON'] == 1), 'etime']
_index_time_engine_idle_int = _index_time_engine_idle.index[0] if len(_index_time_engine_idle) > 0 else 0
        
nVu = df.loc[(vehspd_ecu <= vehspd_urban_kph) & (vehspd_ecu >= 0)].shape[0] # urban driving below 60 kph

nVm = df.loc[vehspd_ecu > vehspd_rural_kph & (vehspd_ecu <= vehspd_motorway_kph)].shape[0] # motoway driving above 90 kph
nVr = df.loc[(vehspd_ecu > vehspd_urban_kph) & (vehspd_ecu <= vehspd_rural_kph)].shape[0] # rural driving 60 to 90 kph

print('nVu = ', nVu, ", nVm = ", nVm, ", nVr = ", nVr)
if (iCOOL_TEMP_cname == '') and (max(df[ECT_cname]) == 0): _index_ECT_cold_start_end = int(300/dt)

dt2 = etime[2] -  etime[0]
veh_ACC = np.nan_to_num(np.where(df['vehspd_ecu'].notna(), (df['vehspd_ecu'].shift(-1) - df['vehspd_ecu'].shift(1)) / (dt2 * 3.6), 0))
vehspd_accel = np.nan_to_num(df['vehspd_ecu'] * veh_ACC)

_idx_Vu = df.loc[(df['vehspd_ecu'] <=  vehspd_urban_kph) & (df['vehspd_ecu'] >= 0)].index
if len(_idx_Vu) > 0:
    Vu = df.loc[_idx_Vu, 'vehspd_ecu'].reset_index(drop=True)
    nVu = len(Vu)
    dVMTu = df.loc[_idx_Vu, 'dVMT'].reset_index(drop=True)
    TIMEu = etime[_idx_Vu]
    Vu_ACC = np.where(Vu >= 1, (Vu.shift(-1) - Vu.shift(1))/(dt2*3.6), 0)
    vehspd_ACCu = Vu * Vu_ACC/3.6
else:
    nVu, Vu, dVMTu, TIMEu = 0, np.zeros(nVu), np.zeros(nVu), np.zeros(nVu)

time_engine_idle = len(df.loc[(df[iENG_SPEED_cname] <= 500)]) * dt
num_vehspd_zero = len(df.loc[(df['vehspd_ecu'] <= vehspd_1_kph/20)])
_idx_Vm = df.loc[(df['vehspd_ecu'] > vehspd_rural_kph) & (vehspd_ecu <= vehspd_motorway_kph)].index
if len(_idx_Vm) > 0:
    Vm = df.loc[_idx_Vm, 'vehspd_ecu'].reset_index(drop=True)
    nVm = len(Vm)
    dVMTm = list(df.loc[_idx_Vm, 'dVMT']) # .reset_index(drop=True)
    TIMEm = list(etime[_idx_Vm])
    Vm_ACC = np.where(Vm >= 1, (Vm.shift(-1) - Vm.shift(1))/(dt2*3.6), 0)
    vehspd_ACCm = Vm * Vm_ACC/3.6
else:
    nVm, Vm, dVMTm, TIMEm = 0, np.zeros(nVm), np.zeros(nVm), np.zeros(nVm)

num_vehspd_100kph = df.loc[(df['vehspd_ecu'] > EU_vehspd_100kph)].shape[0]
num_vehspd_145kph = df.loc[(df['vehspd_ecu'] > EU_peak_vehspd_kph)].shape[0]

_idx_Vr = df.loc[(df['vehspd_ecu'] <=  vehspd_rural_kph) & (df['vehspd_ecu'] > vehspd_urban_kph)].index
if len(_idx_Vr) > 0:
    Vr = df.loc[_idx_Vr, 'vehspd_ecu'] # .reset_index(drop=True)
    nVr = len(Vr)
    dVMTr = df.loc[_idx_Vr, 'dVMT'] # .reset_index(drop=True)
    TIMEr = etime[_idx_Vr]
    Vr_ACC = np.where(Vr >= 1, (Vr.shift(-1) - Vr.shift(1))/(dt2*3.6), 0)
    vehspd_ACCr = Vr * Vr_ACC/3.6
else:
    nVr, Vr, dVMTr, TIMEr = 0, np.zeros(nVr), np.zeros(nVr), np.zeros(nVr)

Inst_Mass_GHG[df['eng_Combustion_ON'] < 0.5] = 0

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6)) 
ax.plot(etime, vehspd_accel, 'r-')
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Vehicle Speed x Acceleration (kph x m/s2)')
ax.set_title('FEV to EPA PEMS Engine Power Comparison')
plt.grid(True)
# plt.legend()
plt.tight_layout()

fig_path = os.path.join(fig_folder, 'vehspd_accel.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig, ax = plt.subplots(facecolor=(1, 1, 1), figsize=(8, 6)) 
ax.plot(etime, veh_ACC, 'b-')
ax.set_xlabel('Elapsed Time (s)')
ax.set_ylabel('Vehicle Acceleration (m/s2)')
ax.set_title('FEV to EPA PEMS Engine Power Comparison')
plt.grid(True)
# plt.legend()
plt.tight_layout()

fig_path = os.path.join(fig_folder, 'veh_ACC.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

plt.close('all')
if dt == 0: dt = etime[1] -  etime[0]
[vehmove_start, vehmove_end, num_vehmove_start] = vehmove_start_end(df, dt, vehmove_start_ini, Inst_Mass_GHG, GHG_WLTC_threshold)

GHG_start = np.zeros(nel)
GHG_end = np.zeros(nel)
GHG_exclude_pts = np.zeros(nel)

GHG_MAW = np.zeros(nel)  
CO_MAW = np.zeros(nel)  
NOx_MAW = np.zeros(nel)  
NMHC_MAW = np.zeros(nel)  
tMAW_engRPM = np.zeros(nel)
tMAW_vehspd = np.zeros(nel)
tMAW_GHG_gpkm = np.zeros(nel)
tMAW_GHG_g = np.zeros(nel)
tMAW_CO_gpkm = np.zeros(nel)
tMAW_NOx_gpkm = np.zeros(nel)
tMAW_NMHC_gpkm = np.zeros(nel)
tMAW_ECT = np.zeros(nel)
tMAW_Texh = np.zeros(nel)
tMAW_VA_95pct = np.zeros(nel)

if max(df['vehspd_ecu']) < 2: 
    print('max(vehspd) < 2')
    sys.exit(0)

vehspd_ecu = df['vehspd_ecu']
kel = 0       
for i in range(0, num_vehmove_start):
    k = int(vehmove_start[i])
    l, MAW_end = int(vehmove_end[i]), int(vehmove_end[i]) + 1
    for iel in range(k, l+1):
        start = GHG_start[iel] = iel
        total_Inst_Mass_GHG = 0
        total_Inst_Mass_CO = 0
        total_Inst_Mass_NOx = 0
        total_Inst_Mass_NMHC = 0
        for jel in range (start, nel): # nel):      
            if df['vehspd_ecu'][jel] >= 0.1 and jel >= _index_ECT_cold_start_end:
                total_Inst_Mass_GHG = total_Inst_Mass_GHG + Inst_Mass_GHG[jel]
                total_Inst_Mass_CO = total_Inst_Mass_CO + Inst_Mass_CO[jel]
                total_Inst_Mass_NOx = total_Inst_Mass_NOx + Inst_Mass_NOx[jel]
                total_Inst_Mass_NMHC = total_Inst_Mass_NMHC + Inst_Mass_NMHC[jel]
                if total_Inst_Mass_GHG > GHG_WLTC_threshold :
                    GHG_MAW[iel] = total_Inst_Mass_GHG
                    CO_MAW[iel] = total_Inst_Mass_CO
                    NOx_MAW[iel] = total_Inst_Mass_NOx
                    NMHC_MAW[iel] = total_Inst_Mass_NMHC
                    if jel == GHG_end[iel-1]:
                        end = GHG_end[iel] = jel+1
                    else:
                        end = GHG_end[iel] = jel                    
                        break
        if max(GHG_end) == 0: 
            print('Max (GHG_end) = 0 or max(vehspd) < 10')
            sys.exit(0)
        # total_Inst_Mass_GHG1 = np.where(df.loc[(df.index >= start) & (df.index >= _index_ECT_cold_start_end) & (vehspd_ecu >= 0.1) & \
        #                                       (Inst_Mass_GHG.iloc[::-1].cumsum().iloc[::-1] > GHG_WLTC_threshold)], Inst_Mass_GHG[start:-1].sum(), 0)
        for ipts in range (start, end):
            if df['vehspd_ecu'][ipts] <= 0: npts_exclude = GHG_exclude_pts[iel] = GHG_exclude_pts[iel] +1

        if end < start or end==0:
            GHG_start[iel] = 0
            GHG_exclude_pts[iel] = 0
        elif iel >= _index_ECT_cold_start_end and np.sum(df['vehspd_ecu'][start:end]) > 0:   
            tMAW_vehspd[kel] = np.sum(df['vehspd_ecu'][start:end])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_engRPM[kel] = np.sum(engine_RPM[start:end])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_ECT[kel] = np.sum(df.loc[start:end, ECT_cname])/(end-start-GHG_exclude_pts[iel]+1)
            tMAW_Texh[kel] = np.sum(Texh[start:end])/(end-start-GHG_exclude_pts[iel]+1)
            if np.sum(df.loc[start:end, 'dVMT']) == 0:
                tMAW_GHG_gpkm[kel] = 0
                tMAW_CO_gpkm[kel] = 0
                tMAW_NOx_gpkm[kel] = 0
                tMAW_NMHC_gpkm[kel] = 0
            else:
                tMAW_GHG_gpkm[kel] = GHG_MAW[iel]*1000/ np.sum(df.loc[start:end, 'dVMT'])
                tMAW_CO_gpkm[kel] = CO_MAW[iel]*1000/ np.sum(df.loc[start:end, 'dVMT'])
                tMAW_NOx_gpkm[kel] = NOx_MAW[iel]*1000/ np.sum(df.loc[start:end, 'dVMT'])
                tMAW_NMHC_gpkm[kel] = NMHC_MAW[iel]*1000/ np.sum(df.loc[start:end, 'dVMT'])

            tMAW_GHG_g[kel] = total_Inst_Mass_GHG
            kel += 1

num_MAW_vehspd = kel
MAW_vehspd = tMAW_vehspd[:num_MAW_vehspd]
MAW_GHG_gpkm = tMAW_GHG_gpkm[:num_MAW_vehspd]
MAW_CO_gpkm = tMAW_CO_gpkm[:num_MAW_vehspd]
MAW_NOx_gpkm = tMAW_NOx_gpkm[:num_MAW_vehspd]
MAW_NMHC_gpkm = tMAW_NMHC_gpkm[:num_MAW_vehspd]

MAWu_vehspd = np.zeros(num_MAW_vehspd)
MAWu_GHG_gpkm = np.zeros(num_MAW_vehspd)
MAWu_CO_gpkm = np.zeros(num_MAW_vehspd)
MAWu_NOx_gpkm = np.zeros(num_MAW_vehspd)
MAWu_NMHC_gpkm = np.zeros(num_MAW_vehspd)

MAWr_vehspd = np.zeros(num_MAW_vehspd)
MAWr_GHG_gpkm = np.zeros(num_MAW_vehspd)
MAWr_CO_gpkm = np.zeros(num_MAW_vehspd)
MAWr_NOx_gpkm = np.zeros(num_MAW_vehspd)
MAWr_NMHC_gpkm = np.zeros(num_MAW_vehspd)

MAWm_vehspd = np.zeros(num_MAW_vehspd)
MAWm_GHG_gpkm = np.zeros(num_MAW_vehspd)
MAWm_CO_gpkm = np.zeros(num_MAW_vehspd)
MAWm_NOx_gpkm = np.zeros(num_MAW_vehspd)
MAWm_NMHC_gpkm = np.zeros(num_MAW_vehspd)

j = nMAWu = nMAWr = nMAWm = 0 
for i in range(num_MAW_vehspd):
    if MAW_vehspd[i] <= MAW_urban_vehspd_kph:
        MAWu_vehspd[nMAWu] = MAW_vehspd[i]
        MAWu_GHG_gpkm[nMAWu] = MAW_GHG_gpkm[i]
        MAWu_CO_gpkm[nMAWu] = MAW_CO_gpkm[i]
        MAWu_NOx_gpkm[nMAWu] = MAW_NOx_gpkm[i]
        MAWu_NMHC_gpkm[nMAWu] = MAW_NMHC_gpkm[i]
        nMAWu = nMAWu + 1
    elif MAW_vehspd[i] > MAW_urban_vehspd_kph and MAW_vehspd[i] <= MAW_rural_vehspd_kph:
        MAWr_vehspd[nMAWr] = MAW_vehspd[i]
        MAWr_GHG_gpkm[nMAWr] = MAW_GHG_gpkm[i]
        MAWr_CO_gpkm[nMAWr] = MAW_CO_gpkm[i]
        MAWr_NOx_gpkm[nMAWr] = MAW_NOx_gpkm[i]        
        MAWr_NMHC_gpkm[nMAWr] = MAW_NMHC_gpkm[i]        
        nMAWr = nMAWr + 1
    elif MAW_vehspd[i] > MAW_rural_vehspd_kph:
        MAWm_vehspd[nMAWm] = MAW_vehspd[i]
        MAWm_GHG_gpkm[nMAWm] = MAW_GHG_gpkm[i]
        MAWm_CO_gpkm[nMAWm] = MAW_CO_gpkm[i]
        MAWm_NOx_gpkm[nMAWm] = MAW_NOx_gpkm[i]        
        MAWm_NMHC_gpkm[nMAWm] = MAW_NMHC_gpkm[i]         
        nMAWm = nMAWm + 1

TIMEt = nVu + nVr + nVm
MAWu_vehspd = MAWu_vehspd[:nMAWu]
MAWu_GHG_gpkm = MAWu_GHG_gpkm[:nMAWu]
MAWu_CO_gpkm = MAWu_CO_gpkm[:nMAWu]
MAWu_NOx_gpkm = MAWu_NOx_gpkm[:nMAWu]
MAWu_NMHC_gpkm = MAWu_NMHC_gpkm[:nMAWu]

MAWr_vehspd = MAWr_vehspd[:nMAWr]
MAWr_GHG_gpkm = MAWr_GHG_gpkm[:nMAWr]
MAWr_CO_gpkm = MAWr_CO_gpkm[:nMAWr]
MAWr_NOx_gpkm = MAWr_NOx_gpkm[:nMAWr]
MAWr_NMHC_gpkm = MAWr_NMHC_gpkm[:nMAWr]

MAWm_vehspd = MAWr_vehspd[:nMAWm]
MAWm_GHG_gpkm = MAWr_GHG_gpkm[:nMAWm]
MAWm_CO_gpkm = MAWr_CO_gpkm[:nMAWm]
MAWm_NOx_gpkm = MAWr_NOx_gpkm[:nMAWm]
MAWm_NMHC_gpkm = MAWr_NMHC_gpkm[:nMAWm]

mean_Vu = np.mean(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'vehspd_ecu'])

rms_dAltitude = 'RMS dAltitude = ' + str(round(np.sqrt(np.mean(dAltitude**2)), 2)) + 'm'
if (max(Altitude) > 0.01):
    rgrade = np.zeros(nel)
    slope = np.zeros(nel)
    for i in range(0, nel):
        if i == 0: 
            slope[0] = 0 
            rgrade[0] = 0
        else:
            if (df.loc[i, 'dVMT'] < 0.01):
                slope[i] = 0
                rgrade[i] = 0
            else:
                slope[i] = dAltitude[i]/df.loc[i, 'dVMT']
                if df['vehspd_ecu'][i] < 0.1: slope[i] = 0
                
                if slope[i] > 0.7:
                    slope[i] = 0.7
                elif slope[i] < -0.7:
                    slope[i] = -0.7
            
                rgrade[i] = 100* np.tan(np.arcsin(slope[i]))
                if rgrade[i] > max_grade_pct:
                    rgrade[i] = max_grade_pct
                elif rgrade[i] < -max_grade_pct:
                    rgrade[i] = -max_grade_pct
else: 
    rgrade = np.zeros(nel)
            
if (max(Altitude) > 0.01):
    rgrade_savgol = savgol_filter(rgrade, window_length=61, polyorder=3)
else:
    rgrade_savgol = np.zeros(nel)

rms_road_grade = 'EPA RMS road grade = ' + str(round(np.sqrt(np.mean(rgrade**2)), 2)) + '%'
rms_road_grade_pos = 'EPA RMS road grade = ' + str(round(np.sqrt(np.mean(rgrade[rgrade>0]**2)), 2)) + '%'
  
rms_road_grade_savgol  =  str(round(np.sqrt(np.mean(rgrade_savgol**2)), 3)) + '%'
rms_road_grade_savgol_pos =  'EPA RMS Road Grade ' + str(round(np.sqrt(np.mean(rgrade_savgol[rgrade_savgol>0]**2)), 2)) + '%'

ehspd_ACCu_95pct = vehspd_ACCr_95pct = vehspd_ACCm_95pct = 0
Vu_ACC_95pct = Vr_ACC_95pct = Vm_ACC_95pct = 0
RPAu = RPAr = RPAm = 0 

[Vu_mean, Vr_mean, Vm_mean] = Vehspd_mean(Vu, Vr, Vm)
[vehspd_ACCu_95pct, vehspd_ACCr_95pct, vehspd_ACCm_95pct] = VehSpd_Accel_95pct(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, Vu_ACC, Vr_ACC, Vm_ACC)
[Vu_ACC_95pct, Vr_ACC_95pct, Vm_ACC_95pct] = VehSpd_Accel_95pct(Vu*Vu_ACC/3.6, Vr*Vr_ACC/3.6, Vm*Vm_ACC/3.6, Vu_ACC, Vr_ACC, Vm_ACC)
[RPAu, RPAr, RPAm] = Relative_Positive_Acceleration(vehspd_ACCu, vehspd_ACCr, vehspd_ACCm, dVMTu, dVMTr, dVMTm, Vu_ACC, Vr_ACC, Vm_ACC)

vehspd_avg_EPA = [Vu_mean, Vr_mean, Vm_mean]
VA_avg_95pct_EPA = [vehspd_ACCu_95pct, vehspd_ACCr_95pct, vehspd_ACCm_95pct]
VA_urm_95pct_EPA = [Vu_ACC_95pct, Vr_ACC_95pct, Vm_ACC_95pct]

RPA_avg_EPA = [RPAu, RPAr, RPAm]

trip_distance = round(df['VMT'][nel-1]/1000, 2)
trip_duration = round(TIMEt/60, 2)
if ((_index_ECT_cold_start_end-_index_time_engine_start)/60 < 5):
    cold_start_duration = 300/60
else:
    cold_start_duration = round((_index_ECT_cold_start_end-_index_time_engine_start)/60, 2)

urban_distance = round(np.sum(dVMTu)/1000, 2)
rural_distance = round(np.sum(dVMTr)/1000, 2)
motorway_distance = round(np.sum(dVMTm)/1000, 2)

urban_distance_share = round(urban_distance/trip_distance*100, 2)
rural_distance_share = round(rural_distance/trip_distance*100, 2)
motorway_distance_share = round(motorway_distance/trip_distance*100, 2)

Vt = (np.sum(Vu) + np.sum(Vr) + np.sum(Vm))/(nVu + nVr + nVm)
Vtm = round(np.mean(df['vehspd_ecu'][df['vehspd_ecu'] >= vehspd_0_kph]), 2)

urban_avg_speed = round(Vu_mean, 2)
rural_avg_speed = round(Vr_mean, 2)
motorway_avg_speed = round(Vm_mean, 2)
total_trip_avg_speed = round(Vt, 2)

if nVm == 0: motorway_speed_above_145kph = 0
else: motorway_speed_above_145kph = round(num_vehspd_145kph/nVm*100, 2)
    
if nVr == 0: motorway_speed_above_100kph = 0
else: motorway_speed_above_100kph = round(num_vehspd_100kph/60, 2) # 60, 2)
urban_stop_time_pct = round(num_vehspd_zero/nVu*100, 2)

delta_altitude_raw = abs(raw_Altitude[nel-1]-raw_Altitude[0])
delta_altitude = abs(Altitude[nel-1]-Altitude[0])
delta_altitude_start_end = min(delta_altitude, delta_altitude_raw)
elevation_pos = np.sum(dAltitude[dAltitude>0])
pu = dAltitude * np.logical_and(df['vehspd_ecu']>=0, df['vehspd_ecu'] <= vehspd_urban_kph)
elevation_pos_u = np.sum(pu[pu>0])

elev_abs_diff_start_end = round(delta_altitude_start_end, 2)
total_cumul_elev_pos_gain_100km = round(elevation_pos*(100/trip_distance), 2)
urban_cumul_elev_pos_gain_100km = round(elevation_pos_u*(100/urban_distance), 2)

init_idle_duration = round(time_engine_idle_int, 2)
_index_time_engine_start = int(_index_time_engine_start)

if _index_ECT_cold_start_end != df.loc[df[ECT_cname] >= 70, :].index[0]:
    _index_ECT_cold_start_end = df.loc[df[ECT_cname] >= 70, :].index[0]
elif _index_time_engine_start > _index_ECT_cold_start_end and _index_ECT_cold_start_end == (300): # /dt): 
    _index_ECT_cold_start_end = _index_ECT_cold_start_end + _index_time_engine_start
    
cold_start_avg_speed = round(np.mean(df['vehspd_ecu'][ _index_time_engine_start: _index_ECT_cold_start_end]), 2)
cold_start_max_speed = round(np.max(df['vehspd_ecu'][ _index_time_engine_start:_index_ECT_cold_start_end]), 2)
# j = 0
# nstop_during_coldstart = 0
# for k in range (int(_index_ECT_cold_start_end)):
# 	if df['vehspd_ecu'][k] <= MAW_vehspd_0_kph: j = j + 1 # for EMROAD zero speed definition
# nstop_during_coldstart = j

# cold_stop_time = round(nstop_during_coldstart)
cold_stop_time = df.loc[(df['vehspd_ecu']<= MAW_vehspd_0_kph) & (df.index < _index_ECT_cold_start_end), 'etime'].count()
# cold_stop_time2 = df.loc[(df['vehspd_ecu']<= 0) & (df.index <= _index_ECT_cold_start_end), 'sTIME'].count()
# print('# of cold_stop_time', cold_stop_time, cold_stop_time1, cold_stop_time2)

nVu_ACC = len(Vu_ACC[Vu_ACC > 0.1])
nVr_ACC = len(Vr_ACC[Vr_ACC > 0.1])
nVm_ACC = len(Vm_ACC[Vm_ACC > 0.1])

nMAWt = len(MAW_vehspd)
    
CO2_RDE_mgpkm = np.sum(Inst_Mass_GHG) *1000/(np.sum(df['dVMT'])/1000)
CO_RDE_mgpkm = np.sum(Inst_Mass_CO) *1000/(np.sum(df['dVMT'])/1000)
NOx_RDE_mgpkm = np.sum(Inst_Mass_NOx) *1000/(np.sum(df['dVMT'])/1000)
NMHC_RDE_mgpkm = np.sum(Inst_Mass_NMHC) *1000/(np.sum(df['dVMT'])/1000)
    
CO2_RDEu_mgpkm = np.sum(Inst_Mass_GHG[df['vehspd_ecu'] <= vehspd_urban_kph]) *1000/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT']/1000))
CO_RDEu_mgpkm = np.sum(Inst_Mass_CO[df['vehspd_ecu'] <= vehspd_urban_kph]) *1000/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT']/1000))
NOx_RDEu_mgpkm = np.sum(Inst_Mass_NOx[df['vehspd_ecu'] <= vehspd_urban_kph]) *1000/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT']/1000))
NMHC_RDEu_mgpkm = np.sum(Inst_Mass_NMHC[df['vehspd_ecu'] <= vehspd_urban_kph]) *1000/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT']/1000))

total_dist_miles = 0.000621371 * df['VMT'][nel-1]
total_fuel_gals = np.sum(fuelflow_gals)
if total_fuel_gals <= 0.1: 
    FE_mpg = 'N/A'
    CO2_gpm = 'N/A'
else:
    FE_mpg = total_dist_miles / total_fuel_gals
    CO2_mgpm = 8887/FE_mpg
    
CO_mgpm_test = 1000 * np.sum(Inst_Mass_CO) / total_dist_miles
CO2_gpm_test = np.sum(Inst_Mass_GHG) / total_dist_miles
engNOx_mgpm_test = 1000 * np.sum(Inst_Mass_NOx) / total_dist_miles
engNMHC_mgpm_test = 1000 * np.sum(Inst_Mass_NMHC) / total_dist_miles

label_CO_test = str(round(CO_mgpm_test/1000,1)) + " g/mile"
label_CO2_gpm_test = str(round(CO2_gpm_test,1)) + " g/mile"
label_engNOx_test = str(round(engNOx_mgpm_test,1)) + " mg/mile"
label_engNMHC_test = str(round(engNMHC_mgpm_test,1)) + " mg/mile"

if FE_mpg == 'N/A':
    print (vehicle_title, ": ", "%0.1f" %GHG_WLTC_threshold, "gCO2 MAW, FE = N/A", label_CO2_gpm_test, label_engNOx_test)
else:
    if vehicle_title != '': print (vehicle_title, ": ", "%0.1f" %GHG_WLTC_threshold, "gCO2 MAW, FE = ", "%0.1f" %FE_mpg, "MPG, GHG = ", label_CO2_gpm_test, ", NOx = ", label_engNOx_test)
    if vehicle_title == '': print ("MAW CO2 WLTC threshold (g) = ", "%0.1f" %GHG_WLTC_threshold, ", FE (mpg) = ", "%0.1f" %FE_mpg, ", GHG (g/mile) = ", label_CO2_gpm_test, ", NOx (mg/mile) = ", label_engNOx_test)
        
if (FEV_to_EPA_PEMS): filename_combined = pp_file.split('.')[0]

curve_CO2 = np.zeros(len(settings_avg_speed))
for j in range (len(settings_avg_speed)):
    tmp = float(settings_avg_speed[j])
    if tmp < WLTC_vehspd_kph [0]: # extra-polation
       curve_CO2[j] =  WLTC_CO2_gpkm[1] + (tmp - WLTC_vehspd_kph[1])/(WLTC_vehspd_kph[0] - WLTC_vehspd_kph[1]) * (WLTC_CO2_gpkm[0] - WLTC_CO2_gpkm[1])
    elif tmp >= WLTC_vehspd_kph [0] and  tmp <= WLTC_vehspd_kph[1]:
       a1 =  (WLTC_CO2_gpkm[1] - WLTC_CO2_gpkm[0])/(WLTC_vehspd_kph[1] - WLTC_vehspd_kph[0])
       curve_CO2[j] = a1 * tmp + (WLTC_CO2_gpkm[0] - a1 * WLTC_vehspd_kph[0])
    elif tmp >= WLTC_vehspd_kph [1] and  tmp <= WLTC_vehspd_kph[2]:
       a1 =  (WLTC_CO2_gpkm[2] - WLTC_CO2_gpkm[1])/(WLTC_vehspd_kph[2] - WLTC_vehspd_kph[1])
       curve_CO2[j] = a1 * tmp + (WLTC_CO2_gpkm[1] - a1 * WLTC_vehspd_kph[1])
    elif tmp >= WLTC_vehspd_kph [2] and  tmp <= WLTC_vehspd_kph[3]:
       a1 =  (WLTC_CO2_gpkm[3] - WLTC_CO2_gpkm[2])/(WLTC_vehspd_kph[3] - WLTC_vehspd_kph[2])
       curve_CO2[j] = a1 * tmp + (WLTC_CO2_gpkm[2] - a1 * WLTC_vehspd_kph[2])
    elif tmp > WLTC_vehspd_kph [3]:
       curve_CO2[j] =  WLTC_CO2_gpkm[2] + (tmp - WLTC_vehspd_kph[2])/(WLTC_vehspd_kph[3] - WLTC_vehspd_kph[2]) * (WLTC_CO2_gpkm[3] - WLTC_CO2_gpkm[2])

settings_TOL1_minus_EPA = (1 + settings_TOL1_minus_pct/100) * curve_CO2
settings_TOL1_plus_EPA = (1 + settings_TOL1_plus_pct/100) * curve_CO2

MAW_GHG_gpkm_TOL_minus = np.interp(MAW_vehspd, settings_avg_speed, settings_TOL1_minus_EPA)
MAW_GHG_gpkm_TOL_plus = np.interp(MAW_vehspd, settings_avg_speed, settings_TOL1_plus_EPA)
MAW_TOL_outside = MAW_GHG_gpkm * np.logical_or(MAW_GHG_gpkm > MAW_GHG_gpkm_TOL_plus, MAW_GHG_gpkm < MAW_GHG_gpkm_TOL_minus) # , MAW_vehspd <= 45)
MAW_TOL_outside_urban = MAW_TOL_outside * np.logical_and(MAW_TOL_outside > 0, MAW_vehspd >= vehspd_1_kph, MAW_vehspd <= MAW_urban_vehspd_kph)
MAW_TOL_outside_rural = MAW_TOL_outside * np.logical_and(MAW_TOL_outside > 0, MAW_vehspd > MAW_urban_vehspd_kph, MAW_vehspd <= MAW_rural_vehspd_kph)
MAW_TOL_outside_motorway = MAW_TOL_outside * np.logical_and(MAW_TOL_outside > 0, MAW_vehspd > MAW_rural_vehspd_kph, MAW_vehspd <= vehspd_motorway_kph)
# len(MAW_TOL_outside [MAW_TOL_outside > 0])
nMAWu_out = len(MAW_TOL_outside_urban [MAW_TOL_outside_urban > 0])
nMAWr_out = len(MAW_TOL_outside_rural [MAW_TOL_outside_rural > 0])
nMAWm_out = len(MAW_TOL_outside_motorway [MAW_TOL_outside_motorway > 0])

rpt2020 = RWE_report (rpt2020, filename_combined, filename_combined_obd, infile_folder, now_hms, suffix, trip_distance, trip_duration, cold_start_duration, urban_distance, rural_distance, motorway_distance,
            urban_distance_share, rural_distance_share, motorway_distance_share, urban_avg_speed, rural_avg_speed, motorway_avg_speed, total_trip_avg_speed,
            motorway_speed_above_145kph, motorway_speed_above_100kph, urban_stop_time_pct, elev_abs_diff_start_end, total_cumul_elev_pos_gain_100km, urban_cumul_elev_pos_gain_100km,
            init_idle_duration, cold_start_avg_speed, cold_start_max_speed, cold_stop_time, vehspd_avg_EPA, RPA_avg_EPA, VA_avg_95pct_EPA, nVu, nVr, nVm,
            nMAWt, nMAWu, nMAWu_out, nMAWr, nMAWr_out, nMAWm, nMAWm_out, CO2_RDE_mgpkm, CO_RDE_mgpkm, NOx_RDE_mgpkm, NMHC_RDE_mgpkm, CO2_RDEu_mgpkm, CO_RDEu_mgpkm, NOx_RDEu_mgpkm, NMHC_RDEu_mgpkm, FE_mpg, 
            CO2_gpm_test, engNOx_mgpm_test, engNMHC_mgpm_test, CO_mgpm_test, GHG_WLTC_threshold)

# df_output = pd.DataFrame()
# df_output['MAW_vehspd'] = MAW_vehspd
# df_output['MAW_GHG_gpkm'] = MAW_GHG_gpkm
# df_output['MAW_GHG_gpkm_TOL_minus'] = MAW_GHG_gpkm_TOL_minus
# df_output['MAW_GHG_gpkm_TOL_plus'] = MAW_GHG_gpkm_TOL_plus
# df_output['MAW_outside_urban'] = MAW_outside_urban
# df_output.to_excel(os.path.splitext(outFilename_inp)[0] + '_TOL_outside' + now_hms + '.xlsx', index=None, header=True)

outFilename_inp = os.path.join(infile_folder, os.path.basename(filename_combined))
outFilename = os.path.splitext(outFilename_inp)[0] + '_fig' + now_hms + '.pdf'

vehspd_ecu= df['vehspd_ecu']
# dVMT = df['dVMT']

Duration = np.zeros(nel)
shp = (nel, 26)
data = np.zeros(shp)
    
data[:,0]=etime
data[:,1]=df['vehspd_ecu']
data[:,2]=df['VMT']
data[:,3]=Altitude
if (Altitude[nel-1] >= 0) == False: Altitude[nel-1] = Altitude[nel-2]
data[:,4]=engine_RPM
data[:,5]=engine_load
data[:,6]=df[ECT_cname]
data[:,7]=Tamb
data[:,8]=Inst_Mass_GHG
data[:,9]=Inst_Mass_NOx
data[:,10]=Inst_Mass_NMHC
data[:,11]=Inst_Mass_CO
data[:,12]=fuelflow_gals
data[:,13]= CatalystTemp

data[:,14]=tMAW_GHG_gpkm
data[:,15]=tMAW_CO_gpkm    
data[:,16]=tMAW_NOx_gpkm     
data[:,17]=tMAW_NMHC_gpkm     
data[:,18]=tMAW_vehspd     
data[:,19]=veh_accel     
data[:,20]=RPA  
data[:,21]=Lambda  

for k in range (nel):
    if GHG_end[k] == GHG_start[k] and GHG_exclude_pts[k] == 0:
        Duration[k] = 0
    else:
        Duration[k] = GHG_end[k] - GHG_start[k] - GHG_exclude_pts[k] + 1
data[:,22] = Duration
data[:,23]=GHG_start
data[:,24]=GHG_end
data[:,25]=GHG_exclude_pts

df_columns_EPA = ['TIME', 'Vehicle Speed', 'Cumul_Distance', 'GPS Altitude', 'Engine RPM', 'Engine Load', 'Coolant Temperature', 'Ambient Temperature', 'Instantaneous Mass CO2', 
                  'Instantaneous Mass NOx', 'Instantaneous Mass NMHC', 'Instantaneous Mass CO', 'Fuel Flow', 'Catalyst Temperature',
                  'Duration', 'Start', 'End', 'Excluded points', 'CO2', 'CO', 'NOx', 'NMHC', 'Mean Vehicle Speed',  'veh_accel',  'RPA',  'lambda']

df_RDE = pd.DataFrame(data, index=None, columns=df_columns_EPA)        
df_RDE['Catalyst Temperature']=df_RDE['Catalyst Temperature'].replace(0, '', regex=True)
df_RDE['Duration']=df_RDE['Duration'].replace(0, '', regex=True)
df_RDE['Start']=df_RDE['Start'].replace(0, '', regex=True)
df_RDE['End']=df_RDE['End'].replace(0, '', regex=True)
df_RDE['Excluded points']=df_RDE['Excluded points'].replace(0, '', regex=True)

df_RDE['CO2']=df_RDE['CO2'].replace(0, '', regex=True)
df_RDE['CO']=df_RDE['CO'].replace(0, '', regex=True)
df_RDE['NOx']=df_RDE['NOx'].replace(0, '', regex=True)
df_RDE['NMHC']=df_RDE['NMHC'].replace(0, '', regex=True)

df_RDE['Mean Vehicle Speed']=df_RDE['Mean Vehicle Speed'].replace(0, '', regex=True)
df_RDE['veh_accel']=df_RDE['veh_accel'].replace(0, '', regex=True)
df_RDE['RPA']=df_RDE['RPA'].replace(0, '', regex=True)
df_RDE['lambda']=df_RDE['lambda'].replace(0, '', regex=True)

VA_95pct_emload1 = np.zeros(len(VA_vehspd_mean_emload))
for j in range (len(VA_vehspd_mean_emload)):
    tmp = float(VA_vehspd_mean_emload[j])
    if tmp < 74.6:
       VA_95pct_emload1[j] =  0.136*tmp+14.44
    else:
       VA_95pct_emload1[j] =  0.0742*tmp+18.966

RPA_mean_emload1 = np.zeros(len(RPA_vehspd_mean_emload))
for j in range (len(RPA_vehspd_mean_emload)):
    tmp = float(RPA_vehspd_mean_emload[j])
    if tmp < 94.05:
       RPA_mean_emload1[j] =  -0.0016 *tmp+0.1755
    else:
       RPA_mean_emload1[j] =  0.025

pp1 = PdfPages(outFilename)

etime0 = etime
etime_reset = False
if (not FEV_to_EPA_PEMS) and (etime_reset == True) and (OBD != True):
    [etime, tvehspd_zero] = etime_resetting(etime, df['vehspd_ecu'])
#    if OBD == True: [otime, tvehspd_zero] = etime_resetting(otime, ovehspd) 

et = pd.DataFrame(etime)
et.columns={'time'}
det_t = et - et.shift(1)
det_t.columns={'delta_t'}

# #[dta.append(det_t.index[i]) for i in det_t.loc[:, 'delta_t'] if det_t.loc[i, 'delta_t'] > 10]

dta = []
for i in range (len(det_t['delta_t'])):
    if det_t.loc[i, 'delta_t'] > 10: dta.append(det_t.index[i])

ndta = len(dta)
netime = len(etime)
for i in range (ndta):
    etime[dta[i]: netime] -= det_t.loc[dta[i], 'delta_t']-1
    
MAW_GHG_gpmi = 1.60934* MAW_GHG_gpkm

sMAWu_GHG_gpkm = MAW_GHG_gpkm[MAW_vehspd<45]
sMAWu_GHG_gpkm = np.mean(sMAWu_GHG_gpkm[sMAWu_GHG_gpkm > 0])
sMAWr_GHG_gpkm = np.mean(MAWr_GHG_gpkm)
sMAWm_GHG_gpkm = np.mean(MAW_GHG_gpkm[MAW_vehspd >= 80])
sMAW_GHG_gpkm = np.mean(MAW_GHG_gpkm)
sMAWu_NOx_gpkm = np.mean(MAW_NOx_gpkm[MAW_vehspd<45])
sMAWr_NOx_gpkm = np.mean(MAWr_NOx_gpkm)
sMAWm_NOx_gpkm = np.mean(MAW_NOx_gpkm[MAW_vehspd >= 80])
sMAW_NOx_gpkm = np.mean(MAW_NOx_gpkm)
sMAW_Vu = np.mean(MAWu_vehspd)
sMAW_Vr = np.mean(MAWr_vehspd)
sMAW_Vm = np.mean(MAW_vehspd[MAW_vehspd>=80])
sMAW_V = np.mean(MAW_vehspd)

MAW_vehspd_urm = [sMAW_Vu, sMAW_Vr, sMAW_Vm]
MAW_GHG_urm = [sMAWu_GHG_gpkm, sMAWr_GHG_gpkm, sMAWm_GHG_gpkm]

df_vehspd = pd.DataFrame(MAW_vehspd_urm, columns=['MAW_vehspd_urm'])
#df_vehspd['MAW_vehspd_urm'] = MAW_vehspd_urm
df_vehspd['MAW_GHG_urm'] = MAW_GHG_urm
df_vehspd['VehSpd_urm'] = vehspd_avg_EPA
df_vehspd['VA_95p_urm'] = VA_urm_95pct_EPA
df_vehspd['RPA_urm'] = RPA_avg_EPA

fname = os.path.splitext(outFilename_inp)[0] + '_out' + now_hms + '.xlsx'

writer = pd.ExcelWriter(fname, engine = 'xlsxwriter')
rpt2020.to_excel(writer, index=None, header=True, sheet_name = 'RWE REPORT-2020')
workbook  = writer.book
worksheet = writer.sheets['RWE REPORT-2020']
worksheet.autofit()

df_RDE.to_excel(writer, index=None, header=True, sheet_name = 'df')
worksheet = writer.sheets['df']
worksheet.autofit()

if (FEV_to_EPA_PEMS): 
    df_mapping.to_excel(writer, index=None, header=True, sheet_name = 'PEMS Mapping')
    worksheet = writer.sheets['PEMS Mapping']
    worksheet.autofit()
    del df_mapping

df_vehspd.to_excel(writer, index=None, header=True, sheet_name = 'VehSpd')
writer.close()

sf=1.0

fig = plt.figure(facecolor=(1, 1, 1))
ax = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)

label_MAW_CO2 = 'MAW w/ ' + str(settings_MAw_CO2_pct) + '% WLTC CO2'
plt.scatter(MAW_vehspd, MAW_GHG_gpkm, s=100, linewidth=1, marker="o", facecolors='none', edgecolors='gray', label='MAW CO2')
plt.scatter(WLTC_vehspd_kph, WLTC_CO2_gpkm, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='c', label = 'WLTC CO2')
plt.scatter(sMAW_Vu, sMAWu_GHG_gpkm, s=200, linewidth=3, marker="s", facecolors='none', edgecolors='b', label = 'Urban Mean CO2: ' + str(round(sMAWu_GHG_gpkm, 1)))
plt.scatter(sMAW_Vr, sMAWr_GHG_gpkm, s=250, linewidth=3, marker="<", facecolors='none', edgecolors='r', label = 'Rural Mean CO2: ' + str(round(sMAWr_GHG_gpkm, 1)))
plt.scatter(sMAW_Vm, sMAWm_GHG_gpkm, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='g', label = 'Motorway Mean CO2: ' + str(round(sMAWm_GHG_gpkm, 1)))
plt.plot([45, 45], [100, 300], 'k--', linewidth=2, label = 'Urban | Rural')
plt.plot([80, 80], [100, 300], 'r--', linewidth=2, label = 'Rural | Motorway')
plt.plot(settings_avg_speed, curve_CO2, 'k-', linewidth=3, label = 'Threshold Curve')
plt.plot(settings_avg_speed, settings_TOL1_minus_EPA, 'y-.', linewidth=3, label = 'Low-Threshold')
plt.plot(settings_avg_speed, settings_TOL1_plus_EPA, 'm-.', linewidth=3, label = 'High Threshold')
plt.xlim(0, 140)
plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [km/h]', 'CO2 Emission [g/km]')

fig_path = os.path.join(fig_folder, 'MAW_GHG_gpkm.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig = plt.figure(facecolor=(1, 1, 1))
ax = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
plt.scatter(vehspd_avg_EPA, VA_avg_95pct_EPA, s=150, linewidth=4, marker="s", facecolors='none', edgecolors='r', label='VA_urm')
plt.plot(VA_vehspd_mean_emload, VA_95pct_emload1, 'y-.', linewidth=3, label ='Limited Curve')
plt.xlim(0.0, 150)
plt.ylim(0.0, 35)
plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [km/h]', '95 pecentile Speed x Acceleration > 0.1 [m2/s3]')

fig_path = os.path.join(fig_folder, 'VA_avg_95pct_EPA.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

fig = plt.figure(facecolor=(1, 1, 1))
ax = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
#plt.plot(RPA_vehspd_mean_emload, RPA_mean_emload, 'b--', linewidth=3, label ='Limited EMROAD')
plt.plot(RPA_vehspd_mean_emload, RPA_mean_emload, 'y-.', linewidth=3, label ='Limited Curve')
# plt.plot(RPA_vehspd_mean_emload, RPA_mean_emload, 'y-.', linewidth=3, label ='Limited Curve')
plt.scatter(vehspd_avg_EPA, RPA_avg_EPA, s=150, linewidth=4, marker="s", facecolors='none', edgecolors='r', label='RPA_urm')
#plt.scatter(vehspd_avg_emload, RPA_avg_emload, s=150, linewidth=4, marker="o", facecolors='none', edgecolors='k', label ='EMROAD RPA_urm @ Horia')
plt.xlim(0.0, 150)
plt_bgf_format(pp1, fig, plt, ax, '', vehicle_title, 'Vehicle Speed [km/h]', 'RPA [m/s2]')

fig_path = os.path.join(fig_folder, 'RPA_mean_emload.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

CO2 = Inst_Mass_GHG * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
CO2u = Inst_Mass_GHG * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NOx = Inst_Mass_NOx * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
NOxu = Inst_Mass_NOx * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NOxr = Inst_Mass_NOx * np.logical_and(vehspd_ecu > vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph)
NOxm = Inst_Mass_NOx * np.logical_and(vehspd_ecu > vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph)
COu = Inst_Mass_CO * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph)
urban_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph))/1000
rural_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph))/1000
highway_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph))/1000
distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_motorway_kph))/1000
total_distance_km = np.sum(df['dVMT'] * np.logical_and(vehspd_ecu>=0, vehspd_ecu>=0))/1000

CO2_gpkm = np.sum(CO2)/distance_km
CO2u_gpkm = np.sum(CO2u)/urban_distance_km
NOx_gpmi = 1000 * np.sum(NOx)/distance_km * 1.60934
NOxu_gpmi = 1000 * np.sum(NOxu)/urban_distance_km * 1.60934

NMHC = Inst_Mass_NMHC * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_motorway_kph)
NMHCu = Inst_Mass_NMHC * np.logical_and(vehspd_ecu >= 0, vehspd_ecu <= vehspd_urban_kph)
NMHCr = Inst_Mass_NMHC * np.logical_and(vehspd_ecu > vehspd_urban_kph, vehspd_ecu <= vehspd_rural_kph)
NMHCm = Inst_Mass_NMHC * np.logical_and(vehspd_ecu > vehspd_rural_kph, vehspd_ecu <= vehspd_motorway_kph)
NMHC_gpmi = 1000 * np.sum(NMHC)/distance_km * 1.60934
NMHCu_gpmi = 1000 * np.sum(NMHCu)/urban_distance_km * 1.60934

MAW_GHGu = np.mean(tMAW_GHG_gpkm [np.logical_and(vehspd_ecu>=1, vehspd_ecu <= vehspd_urban_kph)])

# Save the original stdout so you can restore it later
original_stdout = sys.stdout
current_time1 = datetime.datetime.now()
formatted_datetime = current_time1.strftime("%Y%m%d%H%M%S")

# Open a file to write the output to
with open('console_log_' + formatted_datetime + '.txt', 'w') as f:
    # Redirect standard output to the file
    sys.stdout = f
    if (SCREEN_OUTPUT): sys.stdout = original_stdout 

    if FE_mpg == 'N/A':
        print (vehicle_title, ": ", "%0.1f" %GHG_WLTC_threshold, "gCO2 MAW, FE = N/A", label_CO2_gpm_test, label_engNOx_test)
    else:
        if vehicle_title != '': print (vehicle_title, ": ", "%0.1f" %GHG_WLTC_threshold, "gCO2 MAW, FE = ", "%0.1f" %FE_mpg, "MPG, GHG = ", label_CO2_gpm_test, ", NOx = ", label_engNOx_test)
        if vehicle_title == '': print ("MAW CO2 WLTC threshold (g) = ", "%0.1f" %GHG_WLTC_threshold, ", FE (mpg) = ", "%0.1f" %FE_mpg, ", GHG (g/mile) = ", label_CO2_gpm_test, ", NOx (mg/mile) = ", label_engNOx_test)

    print('Mean Vehicle Speed [km/h] = ', np.mean(vehspd_ecu))
    print('Vehicle Accel [m/s2] = ', np.sqrt(np.mean(veh_accel**2)))
    print('Positive Accel [m/s2] = ', np.sqrt(np.mean(veh_accel[veh_accel > 0.1]**2)))
    print('Negative Accel [m/s2] = ', np.sqrt(np.mean(veh_accel[veh_accel < 0]**2)))
    print('Mean Urban Vehicle Speed > 1 km/h [km/h] = ', np.mean(Vu[Vu >= vehspd_0_kph]))
    print('Mean Urban Vehicle Speed [km/h] = ', np.mean(Vu))
    print('Mean Rural Vehicle Speed [km/h] = ', np.mean(Vr))
    print('Mean Motorway Vehicle Speed [km/h] = ', np.mean(Vm))
    print('RMS Road Grade [%] = ', round(np.sqrt(np.mean(rgrade**2)), 1))
    print('RMS Road Grade Filtered [%] = ', round(np.sqrt(np.mean(rgrade_savgol**2)),1))
    print('RMS Catalyst Temperature [C] = ', np.round(np.sqrt(np.mean(CatalystTemp**2)),1))
    print('*********************')

    print('MAW CO2 [g/km] = ', MAW_GHGu)
    print('CO2 RDE [g/km] = ', CO2_gpkm)
    print('CO2 RDE Urban [g/km] = ', CO2u_gpkm)
    print('CO2 WLTC [g/km] = ', WLTC_GHG_gpkm)
    print('CO2 WLTC Urban [g/km] = ', WLTC_GHGu_gpkm)
    print('NOx RDE [mg/mi] = ', NOx_gpmi)
    print('NOx RDE Urban [mg/mi] = ', NOxu_gpmi)
    print('NMHC RDE [mg/mi] = ', NMHC_gpmi)
    print('NMHC RDE Urban [mg/mi] = ', NMHCu_gpmi)
    print('NMHCu sum (mg) =', 1000 * np.sum(NMHCu), ', urban_distance_mile = ', urban_distance_km/1.60934)
    print('NMHCu sum (mg/mile) =', NMHCu_gpmi, ', NMHCu sum (mg/mile) =', 1000 * np.sum(NMHCu)/(urban_distance_km/1.60934))
    print('distance_km, urban = ', urban_distance_km, ', rural =', rural_distance_km, ', motorway =', highway_distance_km, ', <= 145 kph ', distance_km, ', total VMT =', np.sum(df['dVMT'])/1000)

    sys.stdout = original_stdout 

RF1 = 1.2; RF2 = 1.25
RF1 = 1.3; RF2 = 1.5
a1 = (RF2 - 1) / (RF2*(RF1 - RF2)) 
b1 = 1 - a1 * RF1

CO2f = CO2_gpkm / WLTC_GHG_gpkm
CO2uf = CO2u_gpkm/WLTC_GHGu_gpkm
MAW_GHGu = np.mean(tMAW_GHG_gpkm[np.logical_and(vehspd_ecu>=0, vehspd_ecu <= vehspd_urban_kph)])
mCO2uf = MAW_GHGu/WLTC_GHGu_gpkm
if CO2f <= RF1:
    RFt = 1.0
elif CO2f > RF1 and CO2f <= RF2:
    RFt = a1 * CO2f + b1
else:
    RFt = 1/CO2f

if CO2uf <= RF1:
    RFu = 1.0
elif CO2uf > RF1 and CO2uf <= RF2:
    RFu = a1 * CO2uf + b1
else:
    RFu = 1/CO2uf

print('CO2 ratios RDE/WLTC = ', np.round(CO2f, 2), ' RDE/WLTC Urban = ', np.round(CO2uf, 2), ' MAW/WLTC Urban = ', np.round(mCO2uf, 2))    
print('a1 = ', np.round(a1, 2), ' b1 = ', np.round(b1, 3), ' RF urban = ', np.round(RFu, 2), ' RFt = ', RFt)
   
NOx_x_ranges = [min(MAW_vehspd), max(MAW_vehspd)]
FTP_vehspd_ranges = [min(MAW_vehspd), 91.25]  # mean = 21.2 mph (34.12kph, max = 56.7 mph (91.25 kph)
vehspd_FTP3bags_mph = 1/1.60934*vehspd_FTP3bags_kph
vehspd_US06_mph = 1/1.60934*vehspd_US06_kph
vehspd_HWFET_mph = 1/1.60934*vehspd_HWFET_kph

MAW_GHG_gpmi = 1.60934 * MAW_GHG_gpkm
MAW_vehspd_mph = 1/1.60934 * MAW_vehspd
MAW_CO2_ranges = [HwFET_CO2_gpmi, np.max(MAW_GHG_gpmi)*2/3]
MAWu_CO2_ranges = [HwFET_CO2_gpmi, np.max(MAW_GHG_gpmi)*2/3]
MAW_NOx_mgpmi = 1.60934*1000*MAW_NOx_gpkm
NOx_RDEu_mgpmi = 1.60934*NOx_RDEu_mgpkm
NOx_RDE_mgpmi = 1.60934*NOx_RDE_mgpkm
MAW_NMHC_mgpmi = 1.60934*1000*MAW_NMHC_gpkm
NMHC_RDEu_mgpmi = 1.60934*NMHC_RDEu_mgpkm
NMHC_RDE_mgpmi = 1.60934*NMHC_RDE_mgpkm
FTP_NOx_mgpmi = 1000*FTP_NOx_comp_gpmi
FTP3_NOx_mgpmi = 1000* FTP3_NOx_gpmi
US06_NOx_mgpmi = 1000* US06_NOx_gpmi
HwFET_NOx_mgpmi = 1000* HwFET_NOx_gpmi
MAW_CO_gpmi = 1.60934*MAW_CO_gpkm
CO_RDEu_gpmi = 1.60934*CO_RDEu_mgpkm/1000
CO_RDE_gpmi = 1.60934*CO_RDE_mgpkm/1000

label_NOx_title = vehicle_title + ', NOx Standards: ' + str(int(NOx_STD_mgpmi)) + ' mg/mile'

label_RDEu_NOx = 'RDE Urban NOx: ' + str(round(NOx_RDEu_mgpmi, 1)) + ' mg/mile'
label_RDE_NOx = 'RDE NOx: ' + str(round(NOx_RDE_mgpmi, 1)) + ' mg/mile'
label_NOx_standards = 'NOx Standards: ' + str(int(NOx_STD_mgpmi)) + ' mg/mile'

label_RDEu_CO = 'RDE Urban CO: ' + str(round(CO_RDEu_gpmi, 1)) + ' g/mile'
label_RDE_CO = 'RDE CO: ' + str(round(CO_RDE_gpmi, 1)) + ' g/mile'
label_CO_standards = 'CO Standards: ' + str(round(CO_STD_gpmi, 1)) + ' g/mile'
vehspd_WLTC_mph = WLTC_vehspd_kph/1.6034
WLTC_CO2_gpmi = WLTC_CO2_gpkm*1.6034
WLTC_NOx_mgpmi = 1000*WLTC_NOx_gpmi

NOx_RDE_mgpmi = 1000 * np.sum(Inst_Mass_NOx)/(df['VMT'][nel-1]/1000)*1.6034
GHG_RDE_gpmi = np.sum(Inst_Mass_GHG)/(df['VMT'][nel-1]/1000)*1.6034
CO_RDE_gpmi = np.sum(Inst_Mass_CO)/(df['VMT'][nel-1]/1000)*1.6034
CO_RDEu_gpmi = 1000*np.sum(Inst_Mass_NOx[vehspd_ecu <= vehspd_urban_kph])/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT'])/1000)*1.6034
NOx_RDEu_gpmi = 1000*np.sum(Inst_Mass_NOx[vehspd_ecu <= vehspd_urban_kph])/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT'])/1000)*1.6034
CO_RDEu_gpmi = np.sum(Inst_Mass_CO[vehspd_ecu <= vehspd_urban_kph])/(np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT'])/1000)*1.6034

mean_vehspd_RDE = np.mean(vehspd_ecu)
mean_vehspd_RDEu = np.mean(vehspd_ecu[vehspd_ecu <= vehspd_urban_kph])

# 2016 Malibu Composite Emissions
# Composite_vehspd_kph = np.array([34.12, 53.5, 77.9])/1.60934
# Composite_CO2_gpmi = [264.5, 254.2, 301.8]
# Composite_NOx_mgpmi= [10, 6, 13]
# Composite_CO_gpmi = [0.352, 0.393, 1.269]
Composite_vehspd_kph = np.array([vehspd_FTP3bags_comp_kph, vehspd_US06_comp_kph])/1.60934
Composite_CO2_gpmi = np.array([FTP3_CO2_comp_gpmi, US06_CO2_comp_gpmi])
Composite_NOx_mgpmi = 1000 * np.array([FTP_NOx_comp_gpmi, US06_NOx_comp_gpmi])
Composite_CO_gpmi =  np.array([FTP_CO_comp_gpmi, US06_CO_comp_gpmi])  
    
if (LAB_WLTC_NOx_Fit == True) or (LAB_WLTC_NOx_Fit > 0):
    # Composite_vehspd_kph = np.array([vehspd_FTP3bags_comp_kph, WLTC_vehspd_comp_kph, vehspd_US06_comp_kph])/1.60934
    # Composite_CO2_gpmi = np.array([FTP3_CO2_comp_gpmi, WLTC_CO2_comp_gpmi, US06_CO2_comp_gpmi])
    # Composite_NOx_mgpmi = 1000 * np.array([FTP_NOx_comp_gpmi, WLTC_NOx_comp_gpmi, US06_NOx_comp_gpmi])
    # Composite_CO_gpmi =  np.array([FTP_CO_comp_gpmi, WLTC_CO_comp_gpmi, US06_CO_comp_gpmi])

    # x_vehspd = np.array([vehspd_FTP3bags_mph[0], vehspd_FTP3bags_mph[1], vehspd_FTP3bags_mph[2], vehspd_US06_mph[0], vehspd_US06_mph[1], vehspd_HWFET_mph[0], \
    #                       vehspd_WLTC_mph[0], vehspd_WLTC_mph[1], vehspd_WLTC_mph[2], vehspd_WLTC_mph[3], Composite_vehspd_kph[0], Composite_vehspd_kph[1], Composite_vehspd_kph[2]])
    # x_CO2 = np.array([FTP3_CO2_gpmi[0], FTP3_CO2_gpmi[1], FTP3_CO2_gpmi[2], US06_CO2_gpmi[0], US06_CO2_gpmi[1], HwFET_CO2_gpmi[0], \
    #                   WLTC_CO2_gpmi[0], WLTC_CO2_gpmi[1], WLTC_CO2_gpmi[2], WLTC_CO2_gpmi[3], Composite_CO2_gpmi[0], Composite_CO2_gpmi[1], Composite_CO2_gpmi[2]])
    # y_NOx = np.array([FTP3_NOx_mgpmi[0], FTP3_NOx_mgpmi[1], FTP3_NOx_mgpmi[2], US06_NOx_mgpmi[0], US06_NOx_mgpmi[1], HwFET_NOx_mgpmi[0], \
    #                   WLTC_NOx_mgpmi[0], WLTC_NOx_mgpmi[1], WLTC_NOx_mgpmi[2], WLTC_NOx_mgpmi[3], Composite_NOx_mgpmi[0], Composite_NOx_mgpmi[1], Composite_NOx_mgpmi[2]])
    # y_CO = np.array([FTP3_CO_gpmi[0], FTP3_CO_gpmi[1], FTP3_CO_gpmi[2], US06_CO_gpmi[0], US06_CO_gpmi[1], HwFET_CO_gpmi[0], \
    #                   WLTC_CO_gpmi[0], WLTC_CO_gpmi[1], WLTC_CO_gpmi[2], WLTC_CO_gpmi[3], Composite_CO_gpmi[0], Composite_CO_gpmi[1], Composite_CO_gpmi[2]])
        
    x_vehspd = np.array([vehspd_WLTC_mph[0], vehspd_WLTC_mph[1], vehspd_WLTC_mph[2], vehspd_WLTC_mph[3]])
    x_CO2 = np.array([WLTC_CO2_gpmi[0], WLTC_CO2_gpmi[1], WLTC_CO2_gpmi[2], WLTC_CO2_gpmi[3]])
    y_NOx = np.array([WLTC_NOx_mgpmi[0], WLTC_NOx_mgpmi[1], WLTC_NOx_mgpmi[2], WLTC_NOx_mgpmi[3]])
    y_CO = np.array([WLTC_CO_gpmi[0], WLTC_CO_gpmi[1], WLTC_CO_gpmi[2], WLTC_CO_gpmi[3]])
else:
    Composite_vehspd_kph = np.array([vehspd_FTP3bags_comp_kph, vehspd_US06_comp_kph])/1.60934
    Composite_CO2_gpmi = np.array([FTP3_CO2_comp_gpmi, US06_CO2_comp_gpmi])
    Composite_NOx_mgpmi = 1000 * np.array([FTP_NOx_comp_gpmi, US06_NOx_comp_gpmi])
    Composite_CO_gpmi =  np.array([FTP_CO_comp_gpmi, US06_CO_comp_gpmi])    
    x_vehspd = np.array([vehspd_FTP3bags_mph[0], vehspd_FTP3bags_mph[1], vehspd_FTP3bags_mph[2], vehspd_US06_mph[0], vehspd_US06_mph[1], vehspd_HWFET_mph[0], \
                          Composite_vehspd_kph[0], Composite_vehspd_kph[1]])
    x_CO2 = np.array([FTP3_CO2_gpmi[0], FTP3_CO2_gpmi[1], FTP3_CO2_gpmi[2], US06_CO2_gpmi[0], US06_CO2_gpmi[1], HwFET_CO2_gpmi[0], \
                      Composite_CO2_gpmi[0], Composite_CO2_gpmi[1]])
    y_NOx = np.array([FTP3_NOx_mgpmi[0], FTP3_NOx_mgpmi[1], FTP3_NOx_mgpmi[2], US06_NOx_mgpmi[0], US06_NOx_mgpmi[1], HwFET_NOx_mgpmi[0], \
                      Composite_NOx_mgpmi[0], Composite_NOx_mgpmi[1]])
    y_CO = np.array([FTP3_CO_gpmi[0], FTP3_CO_gpmi[1], FTP3_CO_gpmi[2], US06_CO_gpmi[0], US06_CO_gpmi[1], HwFET_CO_gpmi[0], \
                      Composite_CO_gpmi[0], Composite_CO_gpmi[1]])

# max_pts = min(len(etime), len(vehspd_ecu), len(Inst_Mass_GHG), len(df[ECT_cname]), len(Inst_Mass_GHG))
# print('max_pts = ', max_pts)

coefs_NOx_GHG = poly.polyfit(MAW_GHG_gpmi, MAW_NOx_mgpmi, 1)
x_new = np.linspace(np.min(MAW_GHG_gpmi), np.max(MAW_GHG_gpmi), 100)
ffit = poly.polyval(x_new, coefs_NOx_GHG)

print('MAW NOx/GHG Slope: ', coefs_NOx_GHG[1], ' intercept: ', coefs_NOx_GHG[0])

# '''
# ffit = coefs[0] + coefs[1] * x_new + coefs[2] * x_new**2
# plt.plot(x_new, ffit, 'b-', x_new, ffit1, 'r-.')
# '''

dfx = pd.DataFrame(x_CO2, columns = ['MAW_CO2'])
dfx['MAW_NOx'] = y_NOx
dfx = dfx.sort_values(['MAW_CO2'], ascending=[True])
dfx = dfx.reset_index(drop = True)
coefs_ftp = poly.polyfit(dfx['MAW_CO2'], dfx['MAW_NOx'], 1)
x_ftp = np.linspace(np.min(dfx['MAW_CO2']), np.max(dfx['MAW_CO2']), 100)
ffit_ftp = poly.polyval(x_ftp, coefs_ftp)
print('FTP NOx/GHG Slope: ', coefs_ftp[1], ' intercept: ', coefs_ftp[0])

fig = plt.figure(facecolor=(1, 1, 1))
ax1 = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
mean_GHG_gpmi = 1.6034*np.sum(Inst_Mass_GHG)*1000/np.sum(df['dVMT'])
Inst_Mass_GHGu = Inst_Mass_GHG[vehspd_ecu <= vehspd_urban_kph]
mean_GHGu_gpmi = 1.6034*np.sum(Inst_Mass_GHGu)*1000/np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT'])

plt.scatter(MAW_GHG_gpmi, MAW_NOx_mgpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW NOx')
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW NOx fit')
plt.plot(x_ftp, ffit_ftp,  'm--', linewidth=3, label = 'FTP NOx fit')
plt.scatter(mean_GHG_gpmi, NOx_RDE_mgpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_NOx)
plt.scatter(mean_GHGu_gpmi, NOx_RDEu_mgpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_NOx)
plt.scatter(FTP3_CO2_gpmi, FTP3_NOx_mgpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 NOx')
plt.scatter(US06_CO2_gpmi, US06_NOx_mgpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 NOx')
plt.scatter(HwFET_CO2_gpmi, HwFET_NOx_mgpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET NOx')
plt.scatter(Composite_CO2_gpmi[0], Composite_NOx_mgpmi[0], s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='FTP Composite NOx')
plt.scatter(Composite_CO2_gpmi[1], Composite_NOx_mgpmi[1], s=150, linewidth=3, marker="<", facecolors='none', edgecolors='c', label='US06 Composite NOx')
plt.scatter(mean_GHG_gpmi, WLTC_NOx_comp_gpmi*1000, s=150, linewidth=3, marker="p", facecolors='none', edgecolors='m', label='WLTC Composite NOx')

if (LAB_WLTC_NOx_Fit > 0): plt.scatter(WLTC_CO2_gpmi, WLTC_NOx_mgpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC NOx')

if np.min(MAW_NOx_mgpmi) < 0: 
    plt.ylim(np.min(MAW_NOx_mgpmi)-30, np.max(MAW_NOx_mgpmi)+30)
elif LAB_WLTC_NOx_Fit > 0 and np.max(WLTC_NOx_mgpmi) > np.max(MAW_NOx_mgpmi):
    plt.ylim(0, np.max(WLTC_NOx_mgpmi)+1)
elif NOx_RDEu_mgpmi > np.max(MAW_NOx_mgpmi):
    plt.ylim(0, NOx_RDEu_mgpmi+5)
elif np.min(MAW_NOx_mgpmi) < 10 :
    plt.ylim(0, np.max(MAW_NOx_mgpmi)+5)
elif np.min(MAW_NOx_mgpmi) < dfx['MAW_NOx'][0]:
    plt.ylim(np.min(MAW_NOx_mgpmi)-30, np.max(MAW_NOx_mgpmi)+30)
else:
    plt.ylim(dfx['MAW_NOx'][0]-30, np.max(MAW_NOx_mgpmi)+30)
plt_bgf_format(pp1, fig, plt, ax1, label_NOx_standards, vehicle_title, 'CO2 Emission [g/mile]', 'NOx Emission [mg/mile]')

fig_path = os.path.join(fig_folder, 'MAW_NOx_mgpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

coefs_NOx_vehspd = poly.polyfit(MAW_vehspd_mph, MAW_NOx_mgpmi, 1)
x_new = np.linspace(np.min(MAW_vehspd_mph), np.max(MAW_vehspd_mph), 100)
ffit = poly.polyval(x_new, coefs_NOx_vehspd)
print('MAW NOx/VehSpd Slope: ', coefs_NOx_vehspd[1], ' intercept: ', coefs_NOx_vehspd[0])

dfv = pd.DataFrame(x_vehspd, columns = ['vehspd_mph'])
dfv['MAW_NOx'] = y_NOx
dfv = dfv.sort_values(['vehspd_mph'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_ftp_vehspd = poly.polyfit(dfv['vehspd_mph'], dfv['MAW_NOx'], 1)
x_ftp_vehspd = np.linspace(np.min(dfv['vehspd_mph']), np.max(dfv['vehspd_mph']), 100)
ffit_ftp_vehspd = poly.polyval(x_ftp_vehspd, coefs_ftp_vehspd)
print('FTP NOx/VehSpd Slope: ', coefs_ftp_vehspd[1], ' intercept: ', coefs_ftp_vehspd[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_vehspd_mph, MAW_NOx_mgpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW NOx')
plt.scatter(mean_vehspd_RDE, NOx_RDE_mgpkm*1.6094, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_NOx)
plt.scatter(mean_vehspd_RDEu, NOx_RDEu_mgpkm*1.6094, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_NOx)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW NOx fit')
plt.scatter(vehspd_FTP3bags_mph, FTP3_NOx_mgpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 NOx')
plt.scatter(vehspd_US06_mph, US06_NOx_mgpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 NOx')
plt.scatter(vehspd_HWFET_mph, HwFET_NOx_mgpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET NOx')
# plt.scatter(Composite_vehspd_kph, Composite_NOx_mgpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite NOx')
plt.scatter(Composite_vehspd_kph[0], Composite_NOx_mgpmi[0], s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='FTP Composite NOx')
plt.scatter(Composite_vehspd_kph[1], Composite_NOx_mgpmi[1], s=150, linewidth=3, marker="<", facecolors='none', edgecolors='c', label='US06 Composite NOx')
plt.scatter(mean_vehspd_RDE, WLTC_NOx_comp_gpmi*1000, s=150, linewidth=3, marker="p", facecolors='none', edgecolors='r', label='WLTC Composite NOx')

if (LAB_WLTC_NOx_Fit > 0): plt.scatter(vehspd_WLTC_mph, WLTC_NOx_mgpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC NOx')
plt.plot(x_ftp_vehspd, ffit_ftp_vehspd,  'm--', linewidth=3, label = 'FTP NOx fit')
plt_bgf_format(pp1, fig, plt, ax1, label_NOx_standards, vehicle_title, 'Vehicle Speed [km/h]', 'NOx Emission [mg/mile]')

fig_path = os.path.join(fig_folder, 'vehspd_MAW_NOx_mgpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

coefs_NMHC_GHG = poly.polyfit(MAW_GHG_gpmi, MAW_NMHC_mgpmi, 1)
x_new = np.linspace(np.min(MAW_GHG_gpmi), np.max(MAW_GHG_gpmi), 100)
ffit = poly.polyval(x_new, coefs_NMHC_GHG)

print('MAW NMHC/GHG Slope: ', coefs_NMHC_GHG[1], ' intercept: ', coefs_NMHC_GHG[0])

# '''
# ffit = coefs[0] + coefs[1] * x_new + coefs[2] * x_new**2
# plt.plot(x_new, ffit, 'b-', x_new, ffit1, 'r-.')
# '''

# y_NOx = np.array([FTP3_NOx_mgpmi[0], FTP3_NOx_mgpmi[1], FTP3_NOx_mgpmi[2], US06_NOx_mgpmi[0], US06_NOx_mgpmi[1], HwFET_NOx_mgpmi[0], \
#                     Composite_NOx_mgpmi[0], Composite_NOx_mgpmi[1]])
    
dfx = pd.DataFrame(x_CO2, columns = ['MAW_CO2'])
dfx['MAW_NMHC'] = y_NOx
dfx = dfx.sort_values(['MAW_CO2'], ascending=[True])
dfx = dfx.reset_index(drop = True)
coefs_ftp = poly.polyfit(dfx['MAW_CO2'], dfx['MAW_NMHC'], 1)
x_ftp = np.linspace(np.min(dfx['MAW_CO2']), np.max(dfx['MAW_CO2']), 100)
ffit_ftp = poly.polyval(x_ftp, coefs_ftp)
print('FTP NOx/GHG Slope: ', coefs_ftp[1], ' intercept: ', coefs_ftp[0])

label_RDEu_NMHC = 'RDE Urban NMHC: ' + str(round(NMHC_RDEu_mgpmi, 1)) + ' mg/mile'
label_RDE_NMHC = 'RDE NMHC: ' + str(round(NMHC_RDE_mgpmi, 1)) + ' mg/mile'
label_NMHC_standards = 'NMHC Standards: ' + str(int(NMHC_STD_mgpmi)) + ' mg/mile'

fig = plt.figure(facecolor=(1, 1, 1))
ax1 = fig.add_subplot(1,1,1)
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
mean_GHG_gpmi = 1.6034*np.sum(Inst_Mass_GHG)*1000/np.sum(df['dVMT'])
Inst_Mass_GHGu = Inst_Mass_GHG[df['vehspd_ecu'] <= vehspd_urban_kph]
mean_GHGu_gpmi = 1.6034*np.sum(Inst_Mass_GHGu)*1000/np.sum(df.loc[(df['vehspd_ecu'] <= vehspd_urban_kph), 'dVMT'])

plt.scatter(MAW_GHG_gpmi, MAW_NMHC_mgpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW NMHC')
# plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW NMHC fit')
# plt.plot(x_ftp, ffit_ftp,  'm--', linewidth=3, label = 'FTP NMHC fit')
plt.scatter(mean_GHG_gpmi, NMHC_RDE_mgpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_NMHC)
plt.scatter(mean_GHGu_gpmi, NMHC_RDEu_mgpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_NMHC)
plt.scatter(FTP3_CO2_gpmi, FTP3_NOx_mgpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 NOx')
plt.scatter(US06_CO2_gpmi, US06_NOx_mgpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 NOx')
plt.scatter(HwFET_CO2_gpmi, HwFET_NOx_mgpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET NOx')
# plt.scatter(Composite_CO2_gpmi, Composite_NOx_mgpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite NOx')
plt.scatter(Composite_CO2_gpmi[0], Composite_NOx_mgpmi[0], s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='FTP Composite NOx')
plt.scatter(Composite_CO2_gpmi[1], Composite_NOx_mgpmi[1], s=150, linewidth=3, marker="<", facecolors='none', edgecolors='c', label='US06 Composite NOx')
plt.scatter(mean_GHG_gpmi, WLTC_NMHC_comp_gpmi*1000, s=150, linewidth=3, marker="p", facecolors='none', edgecolors='r', label='WLTC Composite NMHC')

if (LAB_WLTC_NOx_Fit > 0): plt.scatter(WLTC_CO2_gpmi, WLTC_NOx_mgpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC NOx')

if np.min(MAW_NMHC_mgpmi) < 0: 
    plt.ylim(np.min(MAW_NMHC_mgpmi)-30, np.max(MAW_NMHC_mgpmi)+30)
elif LAB_WLTC_NOx_Fit > 0 and np.max(WLTC_NOx_mgpmi) > np.max(MAW_NOx_mgpmi):
    plt.ylim(0, np.max(WLTC_NOx_mgpmi)+1)
elif NMHC_RDEu_mgpmi > np.max(MAW_NMHC_mgpmi):
    plt.ylim(0, NMHC_RDEu_mgpmi+5)
elif np.min(MAW_NMHC_mgpmi) < 10 :
    plt.ylim(0, np.max(MAW_NMHC_mgpmi)+5)
elif np.min(MAW_NMHC_mgpmi) < dfx['MAW_NMHC'][0]:
    plt.ylim(np.min(MAW_NMHC_mgpmi)-30, np.max(MAW_NMHC_mgpmi)+30)
else:
    plt.ylim(dfx['MAW_NMHC'][0]-30, np.max(MAW_NMHC_mgpmi)+30)
plt_bgf_format(pp1, fig, plt, ax1, label_NMHC_standards, vehicle_title, 'CO2 Emission [g/mile]', 'NMHC Emission [mg/mile]')

fig_path = os.path.join(fig_folder, 'MAW_NMHC_mgpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

coefs_NMHC_vehspd = poly.polyfit(MAW_vehspd_mph, MAW_NMHC_mgpmi, 1)
x_new = np.linspace(np.min(MAW_vehspd_mph), np.max(MAW_vehspd_mph), 100)
ffit = poly.polyval(x_new, coefs_NMHC_vehspd)
print('MAW NMHC/VehSpd Slope: ', coefs_NMHC_vehspd[1], ' intercept: ', coefs_NMHC_vehspd[0])

dfv = pd.DataFrame(x_vehspd, columns = ['vehspd_mph'])
dfv['MAW_NMHC'] = y_NOx
dfv = dfv.sort_values(['vehspd_mph'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_ftp_vehspd = poly.polyfit(dfv['vehspd_mph'], dfv['MAW_NMHC'], 1)
x_ftp_vehspd = np.linspace(np.min(dfv['vehspd_mph']), np.max(dfv['vehspd_mph']), 100)
ffit_ftp_vehspd = poly.polyval(x_ftp_vehspd, coefs_ftp_vehspd)
print('FTP NMHC/VehSpd Slope: ', coefs_ftp_vehspd[1], ' intercept: ', coefs_ftp_vehspd[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_vehspd_mph, MAW_NMHC_mgpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW NMHC')
plt.scatter(mean_vehspd_RDE, NMHC_RDE_mgpkm*1.6094, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_NMHC)
plt.scatter(mean_vehspd_RDEu, NMHC_RDEu_mgpkm*1.6094, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_NMHC)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW NMHC fit')
plt.scatter(vehspd_FTP3bags_mph, FTP3_NOx_mgpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 NOx')
plt.scatter(vehspd_US06_mph, US06_NOx_mgpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 NOx')
plt.scatter(vehspd_HWFET_mph, HwFET_NOx_mgpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET NOx')
plt.scatter(Composite_vehspd_kph, Composite_NOx_mgpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite NOx')
plt.scatter(Composite_vehspd_kph[0], Composite_NOx_mgpmi[0], s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='FTP Composite NOx')
plt.scatter(Composite_vehspd_kph[1], Composite_NOx_mgpmi[1], s=150, linewidth=3, marker="<", facecolors='none', edgecolors='c', label='US06 Composite NOx')
plt.scatter(mean_vehspd_RDE, WLTC_NMHC_comp_gpmi*1000, s=150, linewidth=3, marker="p", facecolors='none', edgecolors='r', label='WLTC Composite NMHC')

if (LAB_WLTC_NOx_Fit > 0): plt.scatter(vehspd_WLTC_mph, WLTC_NOx_mgpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC NOx')
plt.plot(x_ftp_vehspd, ffit_ftp_vehspd,  'm--', linewidth=3, label = 'FTP NOx fit')
plt_bgf_format(pp1, fig, plt, ax1, label_NMHC_standards, vehicle_title, 'Vehicle Speed [km/h]', 'NMHC Emission [mg/mile]')

fig_path = os.path.join(fig_folder, 'vehspd_MAW_NMHC_mgpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

coefs_CO_GHG = poly.polyfit(MAW_GHG_gpmi, MAW_CO_gpmi, 1)
x_new = np.linspace(np.min(MAW_GHG_gpmi), np.max(MAW_GHG_gpmi), 100)
ffit = poly.polyval(x_new, coefs_CO_GHG)
print('MAW CO/CO2 Slope: ', coefs_CO_GHG[1], ' intercept: ', coefs_CO_GHG[0])

dfv = pd.DataFrame(x_CO2, columns = ['MAW_CO2'])
dfv['MAW_CO'] = y_CO
dfv = dfv.sort_values(['MAW_CO2'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_CO_ftp_GHG = poly.polyfit(dfv['MAW_CO2'], dfv['MAW_CO'], 1)
x_ftp_GHG = np.linspace(np.min(dfv['MAW_CO2']), np.max(dfv['MAW_CO2']), 100)
ffit_ftp_GHG = poly.polyval(x_ftp_GHG, coefs_CO_ftp_GHG)
print('FTP CO/CO2 Slope: ', coefs_CO_ftp_GHG[1], ' intercept: ', coefs_CO_ftp_GHG[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_GHG_gpmi, MAW_CO_gpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW CO')
plt.scatter(mean_GHG_gpmi, CO_RDE_gpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_CO)
plt.scatter(mean_GHGu_gpmi, CO_RDEu_gpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_CO)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW CO fit')
plt.scatter(FTP3_CO2_gpmi, FTP3_CO_gpmi, s=150, linewidth=3, marker="d", facecolors='none', edgecolors='g', label='FTP3 CO')
plt.scatter(US06_CO2_gpmi, US06_CO_gpmi, s=150, linewidth=3, marker="h", facecolors='none', edgecolors='b', label='US06 CO')
plt.scatter(HwFET_CO2_gpmi, HwFET_CO_gpmi, s=150, linewidth=3, marker="o", facecolors='none', edgecolors='c', label='HwFET CO')
# plt.scatter(Composite_CO2_gpmi, Composite_CO_gpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite CO')
plt.scatter(Composite_CO2_gpmi[0], Composite_CO_gpmi[0], s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='FTP Composite CO')
plt.scatter(Composite_CO2_gpmi[1], Composite_CO_gpmi[1], s=150, linewidth=3, marker="<", facecolors='none', edgecolors='c', label='US06 Composite CO')
plt.scatter(mean_GHG_gpmi, WLTC_CO_comp_gpmi, s=150, linewidth=3, marker="p", facecolors='none', edgecolors='r', label='WLTC Composite CO')

if (LAB_WLTC_NOx_Fit > 0): plt.scatter(WLTC_CO2_gpmi, WLTC_CO_gpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC CO')
plt.plot(x_ftp_GHG, ffit_ftp_GHG,  'm--', linewidth=3, label = 'FTP CO fit')
plt_bgf_format(pp1, fig, plt, ax1, label_CO_standards, vehicle_title, 'CO2 Emission [g/mile]', 'CO Emission [g/mile]')

fig_path = os.path.join(fig_folder, 'MAW_CO_gpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

coefs_CO_vehspd = poly.polyfit(MAW_vehspd_mph, MAW_CO_gpmi, 1)
x_new = np.linspace(np.min(MAW_vehspd_mph), np.max(MAW_vehspd_mph), 100)
ffit = poly.polyval(x_new, coefs_CO_vehspd)
print('MAW CO/VehSpd Slope: ', coefs_CO_vehspd[1], ' intercept: ', coefs_CO_vehspd[0])

dfv = pd.DataFrame(x_vehspd, columns = ['vehspd_mph'])
dfv['MAW_CO'] = y_CO
dfv = dfv.sort_values(['vehspd_mph'], ascending=[True])
dfv = dfv.reset_index(drop = True)
coefs_CO_ftp_vehspd = poly.polyfit(dfv['vehspd_mph'], dfv['MAW_CO'], 1)
x_ftp_vehspd = np.linspace(np.min(dfv['vehspd_mph']), np.max(dfv['vehspd_mph']), 100)
ffit_ftp_vehspd = poly.polyval(x_ftp_vehspd, coefs_CO_ftp_vehspd)
print('FTP CO/VehSpd Slope: ', coefs_CO_ftp_vehspd[1], ' intercept: ', coefs_CO_ftp_vehspd[0])

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
ax1 = fig.add_subplot(1, 1, 1)  # create an axes object in the figure
plt.scatter(MAW_vehspd_mph, MAW_CO_gpmi, s=50, linewidth=1, marker="o", facecolors='none', edgecolors=MAW_CO2_plot_color, label='MAW CO')
plt.scatter(mean_vehspd_RDE, CO_RDE_gpmi, s=250, linewidth=3, marker="*", facecolors='none', edgecolors='r', label = label_RDE_CO)
plt.scatter(mean_vehspd_RDEu, CO_RDEu_gpmi, s=250, linewidth=3, marker="^", facecolors='none', edgecolors='r', label = label_RDEu_CO)
plt.plot(x_new, ffit,  'g--', linewidth=3, label = 'MAW CO fit')
plt.scatter(vehspd_FTP3bags_mph, FTP3_CO_gpmi, s=150, linewidth=5, marker="d", facecolors='none', edgecolors='g', label='FTP3 CO')
plt.scatter(vehspd_US06_mph, US06_CO_gpmi, s=150, linewidth=5, marker="h", facecolors='none', edgecolors='b', label='US06 CO')
plt.scatter(vehspd_HWFET_mph, HwFET_CO_gpmi, s=150, linewidth=5, marker="o", facecolors='none', edgecolors='c', label='HwFET CO')
# plt.scatter(Composite_vehspd_kph, Composite_CO_gpmi, s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='Composite CO')
plt.scatter(Composite_vehspd_kph[0], Composite_CO_gpmi[0], s=150, linewidth=3, marker=">", facecolors='none', edgecolors='c', label='FTP Composite CO')
plt.scatter(Composite_vehspd_kph[1], Composite_CO_gpmi[1], s=150, linewidth=3, marker="<", facecolors='none', edgecolors='c', label='US06 Composite CO')
plt.scatter(mean_vehspd_RDE, WLTC_CO_comp_gpmi, s=150, linewidth=3, marker="p", facecolors='none', edgecolors='r', label='WLTC Composite CO')

if (LAB_WLTC_NOx_Fit > 0): plt.scatter(vehspd_WLTC_mph, WLTC_CO_gpmi, s=150, linewidth=3, marker="v", facecolors='none', edgecolors='k', label='WLTC CO')
plt.plot(x_ftp_vehspd, ffit_ftp_vehspd,  'm--', linewidth=3, label = 'FTP CO fit')
plt_bgf_format(pp1, fig, plt, ax1, label_CO_standards, vehicle_title, 'Vehicle Speed [km/h]', 'CO Emission [g/mile]')

fig_path = os.path.join(fig_folder, 'vehspd_CO_gpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
# print(f"Figure saved to: {fig_path}")

fuelflow_gps = df[iWf_cname] if (iWf_cname != "") else fuelflow_gals*2834.89 # https://www.aqua-calc.com/calculate/volume-to-weight
print('RMS Fuel Flow [g/s] = ', np.sqrt(np.mean(fuelflow_gps**2)))

# fuelflow_gps = fuelflow_gps
Inst_Mass_GHG = Inst_Mass_GHG
coefs_fuelflow = poly.polyfit(fuelflow_gps, Inst_Mass_GHG, 1)
x_new = np.linspace(np.min(fuelflow_gps), np.max(fuelflow_gps), 100)
ffit = poly.polyval(x_new, coefs_fuelflow)
print('MAW CO2/Fuel Flow Slope: ', coefs_fuelflow[1], ' intercept: ', coefs_fuelflow[0])

# '''
# ffit1 = coefs[0] + coefs[1] * x_new + coefs[2] * x_new**2
# plt.plot(x_new, ffit, 'b-', x_new, ffit1, 'r-.')
# '''

sf = 1.0
fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
subplot1=fig.add_subplot(111)
#subplot1.scatter(fuelflow_gals*2834.89, Inst_Mass_GHG, 'b-.', label = 'GPS')
plt.plot(x_new, ffit,  'k--', linewidth=5, label = 'RDE CO2 fit')
subplot1.scatter(fuelflow_gps, Inst_Mass_GHG, s=150, linewidth=2, marker="d", facecolors='none', edgecolors='r', label = 'CO2')
plt_bgf_format(pp1, fig, plt, subplot1, '', vehicle_title, 'Fuel Flow [g/s]', 'CO2 Emission [g/s]')

fig_path = os.path.join(fig_folder, 'fuelfow_CO2_gpmi.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

fig = plt.figure()
fig.set_size_inches(11*sf, 8.5*sf, forward=True)
subplot1=fig.add_subplot(211)
subplot1.plot(etime, vehspd_ecu, 'b-', linewidth=1, label = 'Vehicle Speed')
#subplot1.set_xlim(0, len(etime))
subplot1.set_ylabel('Vehicle Speed [km/hr]')
subplot1.grid()
#subplot1.set_xticks([])

subplot2 = fig.add_subplot(212)
subplot2.plot(etime, rgrade, 'b-', linewidth=1, label ='Raw')
if (max(Altitude) > 0.01): subplot2.plot(etime, rgrade_savgol, 'r--', linewidth=1, label ='w/ SG filter')
#subplot2.set_xlim(0.0, len(etime))
subplot2.set_xlabel('Elapsed Time [s]')
subplot2.set_ylabel('Road Grade [%]')
subplot2.grid()
# subplot2.get_shared_x_axes().join(subplot1, subplot2)
subplot2.sharex(subplot1)
subplot1.set_xticklabels([])
pp1.savefig(fig, dpi=600)

fig_path = os.path.join(fig_folder, 'vehspd_rgrade.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')
plt.close('all')

# print('after rgrade_savgol plot')
sum_Corrected_Inst_Mass_NOx = np.sum(Corrected_Inst_Mass_NOx)
sum_Inst_Mass_GHG = np.sum(Inst_Mass_GHG)
sum_Inst_Mass_NOx = np.sum(Inst_Mass_NOx)
label_Inst_Mass_GHG = 'CO2: ' + str(round(sum_Inst_Mass_GHG/1000,1)) + ' Kg'
label_Corrected_Inst_Mass_NOx = 'NOx: ' + str(round(sum_Corrected_Inst_Mass_NOx,1)) + ' g'
label_Inst_Mass_NOx = 'Inst Mass NOx: ' + str(round(sum_Inst_Mass_NOx,1)) + ' g'
label_Inst_Mass_CO = 'CO: ' + str(round(np.sum(Inst_Mass_CO),1)) + ' g'

if OBD == True: 
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Inst_Mass_GHG, 'g-', '', 'CO2 [g/s]', 'Inst CO2', \
                            etime, 1000*Inst_Mass_NOx, 'r', '', 'NOx [mg/s]', label_Inst_Mass_NOx, \
                            etime, 1000*Inst_Mass_CO, 'b', 'Elapsed Time [s]', 'CO [mg/s]', label_CO_test, \
                            otime, oengect, 'r--', 'Coolant Temperature [C]', 'Coolant Temperature')

elif OBD == False and iCOOL_TEMP_cname != '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Inst_Mass_GHG, 'g-', '', 'CO2 [g/s]', 'Inst CO2', \
                            etime, 1000*Inst_Mass_NOx, 'r', '', 'NOx [mg/s]', label_Inst_Mass_NOx, \
                            etime, 1000*Inst_Mass_CO, 'b', 'Elapsed Time [s]', 'CO [mg/s]', label_Inst_Mass_CO, \
                            etime, df[ECT_cname], 'r--', 'Coolant Temperature [C]', 'Coolant Temperature')

# if (max(Altitude) <= 0.0): rgrade_savgol = np.zeros(etime)
subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'b-', '', 'Vehicle Speed [km/h]', 'PEMS', \
                        etime, Altitude, 'r', '', 'Altitude [m]', rms_dAltitude, \
                        etime, rgrade, 'b-', 'Elapsed Time [s]', 'Grade [%]', 'Raw', \
                        etime, rgrade_savgol, 'r--', '', 'w/ SG filter')

fig_path = os.path.join(fig_folder, 'CO2_NOx_CO.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

if (iENG_SPEED_cname != '' and iENG_LOAD_cname != '' and iCOOL_TEMP_cname != '') or \
    (OBD == True and max(oengrpm) > 500 and max(oengtrq) > 0 and max(oengect) > 0):
    if OBD == True:
        subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, otime, oengrpm, 'b-', '', 'Engine Speed [RPM]', 'OBD', \
                                otime, oengtrq, 'r', '', 'Engine Torque [Nm]', 'OBD', \
                                etime, df[ECT_cname], 'r-', 'Elapsed Time [s]', 'Engine Coolant Temperature [C]', 'ECT - PEMS', \
                                otime, oengect, 'b--', 'Coolant Temperature [C]', 'ECT - OBD')
    else:
        subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, engine_RPM, 'b-', '', 'Engine Speed [RPM]', 'PEMS', \
                                etime, engine_load, 'r', '', 'Engine Load [%]', 'PEMS', \
                                etime, df[ECT_cname], 'r-', 'Elapsed Time [s]', 'Engine Coolant Temperature [C]', 'PEMS', \
                                '', '', 'r--', '', '')
        
if iENG_SPEED_cname != '' and iENG_LOAD_cname == '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, 1000*Inst_Mass_NOx, 'b-', '', 'NOx [mg/s]', label_Inst_Mass_NOx, \
                            etime, 1000*Inst_Mass_CO, 'r', '', 'CO [mg/s]', 'Inst Mass CO: ' + label_CO_test, \
                            etime, engine_RPM, 'g-', 'Elapsed Time [s]', 'Engine RPM', 'PEMS', \
                            '', '', 'r--', '', '')
elif iENG_SPEED_cname != '' and iENG_LOAD_cname != '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, 1000*Inst_Mass_NOx, 'b-', '', 'NOx [mg/s]', label_Inst_Mass_NOx, \
                            etime, 1000*Inst_Mass_CO, 'r', '', 'CO [mg/s]', 'Inst Mass CO: ' + label_CO_test, \
                            etime, engine_RPM, 'b-', 'Elapsed Time [s]', 'Engine RPM', 'Engine Speed', \
                            etime, engine_load, 'r--', 'Engine Load [%]', 'Engine Load')

if OBD == True: 
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, Altitude, 'r', '', 'Altitude [m]', rms_dAltitude, \
                            etime, dAltitude, 'r', '', 'delta Altitude', 'delta Altitude @ t=0', \
                            etime, vehspd_ecu, 'b-', 'Elapsed Time [s]', 'Vehicle Speed [km/h]', 'PEMS', \
                            otime, ovehspd, 'r--', '', 'OBD')
else:
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'b', '', 'Vehicle Speed [km/h]', 'OBD', \
                            etime, dAltitude, 'g', '', 'delta Altitude', 'delta Altitude @ t=0', \
                            etime, Altitude, 'b-.', 'Elapsed Time [s]', 'Altitude [m]', 'Filtered', \
                            etime, raw_Altitude, 'r--', '', 'Raw')

# print('before nVu > 0 and nVr > 0 and nVm >  plot')

fig_path = os.path.join(fig_folder, 'vehspd_alt_rgrade.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')


if nVu > 0 and nVr > 0 and nVm > 0:
    xVu = np.arange(nVu)
    xVr = np.arange(nVr)
    xVm = np.arange(nVm)
    subplot_shared_x_axes(sf, pp1, 'noshare', '', xVm, Vm, 'r', '', '', 'Motorway', \
                            xVr, Vr, 'g', '', 'Vehicle Speed [kph]', 'Rural', \
                            xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
                            '', '', 'r--', '', '')
    # subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', xVm, Vm, 'r', '', '', 'Motorway', \
    #                         xVr, Vr, 'g', '', 'Vehicle Speed [kph]', 'Rural', \
    #                         xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
    #                         '', '', 'r--', '', '')
elif nVu > 0 and nVr > 0 and nVm == 0:
    xVu = np.arange(nVu)
    xVr = np.arange(nVr)
    subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', \
                            xVr, Vr, 'g', '', 'Vehicle Speed [kph]', 'Rural', \
                            '', '', 'b-', '', '', '', \
                            xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm > 0:
    xVu = np.arange(nVu)
    xVm = np.arange(nVm)
    subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', \
                            xVm, Vm, 'g', '', 'Vehicle Speed [kph]', 'Motorway', \
                            '', '', 'b-', '', '', '', \
                            xVu, Vu, 'b-', 'Sampling Points', '', 'Urban', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm == 0:
    xVu = np.arange(nVu)
    subplot_shared_x_axes(sf, pp1, 'noshare', 'Vehicle Speeds @ Ubran, Rural & Motorway', \
                            xVu, Vu, 'g', 'Sampling Points', 'Vehicle Speed [kph]', 'Urban', \
                            '', '', 'b-', '', '', '', \
                            '', '', 'b-', '', '', '', \
                            '', '', 'r--', '', '')

subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, engine_RPM, 'g-', '', 'Engine Speed [RPM]', 'Engine RPM', \
                        etime, engine_load, 'b-', '', 'Engine Load [%]', 'Engine Load', \
                        etime, fuelflow_gps, 'r-', 'Elapsed Time [s]', 'Fuel Flow [g/s]', 'Fuel Flow', \
                        '', '', 'r--', '', '')

fig_path = os.path.join(fig_folder, 'engspd_load_fuel.jpg')
fig.savefig(fig_path, dpi=600, bbox_inches='tight')

if nVu > 0 and nVr > 0 and nVm > 0:
    subplot_shared_x_axes(sf, pp1, 'noshare', vehicle_title, xVu, Vu * Vu_ACC, 'g', '', '', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            xVr, Vr * Vr_ACC, 'b', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Rural, ' + str(round(vehspd_ACCr_95pct,1)), \
                            xVm, Vm * Vm_ACC, 'r-', 'Sampling Points', '', 'Motorway, ' + str(round(vehspd_ACCm_95pct,1)), \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm == 0:
    xVu = np.arange(nVu)
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, xVu, Vu * Vu_ACC, 'g', 'Sampling Points', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            '', '', 'b', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Rural, ', \
                            '', '', 'r-', 'Sampling Points', '', 'Motorway, ', \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr > 0 and nVm == 0:
    xVu = np.arange(nVu)
    xVr = np.arange(nVr)
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, xVu, Vu * Vu_ACC, 'g', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            '', '', 'b', '', '', '', \
                            xVr, Vr * Vr_ACC, 'b', 'Sampling Points', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Rural, ' + str(round(vehspd_ACCr_95pct,1)), \
                            '', '', 'r--', '', '')
elif nVu > 0 and nVr == 0 and nVm > 0:
    xVu = np.arange(nVu)
    xVm = np.arange(nVm)
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, xVu, Vu * Vu_ACC, 'g', '', '95% of Vehicle Speed x Acceleration [m2/s3]', 'Urban, ' + str(round(vehspd_ACCu_95pct,1)), \
                            '', '', 'b', '', '', '', \
                            xVm, Vm * Vm_ACC, 'r-', 'Sampling Points', '', 'Motorway, ' + str(round(vehspd_ACCm_95pct,1)), \
                            '', '', 'r--', '', '')
    
subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'g', 'Elapsed Time [s]', 'Vehicle Speed [KPH]', '', \
                        etime, veh_ACC, 'b', '', 'Vehicle Acceleration [m/s2]', 'Veh_ACC ', \
                        etime, vehspd_ecu * veh_ACC, 'r-', 'Elapsed Time [s]', 'Vehspd * ACC', 'Vehspd x veh_ACC', \
                        '', '', 'r--', '', '')
subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'g', 'Elapsed Time [s]', 'Vehicle Speed [KPH]', '', \
                        etime, Tamb, 'b', '', 'Ambient Temperature', 'Tamb [C] ', \
                        etime, CatalystTemp, 'r-', 'Elapsed Time [s]', 'Catalyst Temperature', 'CAT. Temp [C]', \
                        '', '', 'r--', '', '')
    
subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'b', '', 'Vehicle Speed [km/h]', 'OBD', \
                        etime, dAltitude, 'g', '', 'delta Altitude', 'delta Altitude @ t=0', \
                        etime, Altitude, 'b-.', 'Elapsed Time [s]', 'Altitude [m]', 'Filtered', \
                        etime, raw_Altitude, 'r--', '', 'Raw')
plt.plot(etime, raw_Altitude, 'r--', etime, Altitude, 'b-.')

if iENG_SPEED_cname != '':
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'g', 'Elapsed Time [s]', 'Vehicle Speed [KPH]', '', \
                        etime, veh_ACC, 'b', '', 'Acceleration [m/s2]', 'Veh_ACC ', \
                        etime, df.iENG_SPEED, 'r-', 'Elapsed Time [s]', 'Engine Speed', 'engSpd', \
                        '', '', 'r--', '', '')
else:
    subplot_shared_x_axes(sf, pp1, 'share', vehicle_title, etime, vehspd_ecu, 'g', 'Elapsed Time [s]', 'Vehicle Speed [KPH]', '', \
                        etime, veh_ACC, 'b', '', 'Vehicle Acceleration [m/s2]', 'Veh_ACC ', \
                        etime, vehspd_ecu * veh_ACC, 'r-', 'Elapsed Time [s]', 'Vehspd * ACC', 'Vehspd x veh_ACC', \
                        '', '', 'r--', '', '')

if OBD == True:   
    sf = 1.0
    fig = plt.figure()
    fig.set_size_inches(11*sf, 8.5*sf, forward=True)
    subplot1=fig.add_subplot(111)
    subplot1.plot(etime, vehspd_ecu, 'b-.', label = 'GPS')
    if OBD == True: subplot1.plot(otime, ovehspd, 'r--', label = 'OBD')
    subplot1.set_xlim(0, round(np.max(etime)))
#    subplot1.set_xlim(0, 1000)
    subplot1.set_ylim(000, 125)
    subplot1.grid()
    subplot1.set_title(vehicle_title)
    subplot1.set_ylabel('Vehicle Speed [km/h]')
    subplot1.legend(loc='best')      
    pp1.savefig(fig, dpi=600)

plt.close('all')
pp1.close()
#plt.show()
os.chdir(odir)

elapsed = time.time() - current_time
print("\n*************\nElapsed Time: {:5.0f} Seconds after plotting and saving to PDF" .format(elapsed))